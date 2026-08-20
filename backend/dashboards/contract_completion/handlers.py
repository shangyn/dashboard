"""
Excel解析处理器 — 合同完成情况表

每个 handler 函数签名: handler(file_path: str) -> dict
返回: {"success": True/False, "message": "...", "rows": 0}
"""
import os
import re
from datetime import datetime, date
import xlrd
import openpyxl
from models import db
from dashboards.contract_completion.models import (
    LedgerContract, CountryMapping, PaymentCollection, ScheduleTracking
)


# ── 工具函数 ──────────────────────────────────────────────

def _safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def _safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('', '-', '#DIV/0!', '#N/A', '#REF!', '#VALUE!'):
        return 0.0
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return 0.0


def _safe_int(v):
    val = _safe_float(v)
    return int(round(val))


def _safe_date_xlrd(book, value):
    """xlrd 日期转换"""
    if value is None:
        return None
    if isinstance(value, float) and value > 0:
        try:
            dt = xlrd.xldate_as_datetime(value, book.datemode)
            return dt.date()
        except Exception:
            pass
    s = str(value).strip()
    if not s or s == '-':
        return None
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _safe_date_openpyxl(value):
    """openpyxl 日期转换"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s or s == '-':
        return None
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Handler 1: 台账合同 ───────────────────────────────────

def parse_ledger_contracts(file_path: str) -> dict:
    """解析国贸合同标的台账.xlsx → cc_ledger_contract（全量替换）
    只导入2025年及之后的合同（用于当年统计+去年同期对比）
    """
    try:
        # 使用 data_only=True 读取（不用 read_only，此文件 XML 维度标记有误）
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 读取表头 → 1-indexed 列号映射
        header_map = {}
        for c in range(1, ws.max_column + 1):
            v = _safe_str(ws.cell(row=1, column=c).value)
            if v:
                header_map[v] = c  # 1-indexed column number

        # 多表头名查找辅助（支持不同版本台账的列名差异）
        def _h(*names):
            """在 header_map 中按优先级查找列号"""
            for n in names:
                col = header_map.get(n)
                if col is not None:
                    return col
            return None

        # 关键列索引
        col_contract_no = _h('合同编号', '合同号')
        col_sign_date = header_map.get('签订日期')
        col_schedule_date = header_map.get('排产日期')
        col_delivery_date = _h('组A日期', '实际发货日期', '实际发运日期')

        # 所有需要读取的列 → (Excel列号, 模型字段, 类型)
        # 支持两种表头命名（不同版本台账可能不同）
        col_field_map = [
            (_h('合同编号', '合同号'),            'contract_no',           'str'),
            (_h('标的编号', '梯号'),               'ladder_no',             'str'),
            (header_map.get('项目名称'),            'project_name',          'str'),
            (header_map.get('币种'),                'currency',              'str'),
            (header_map.get('合同额（原币）'),      'contract_amount_orig',   'float'),
            (header_map.get('台数'),                'unit_count',            'int'),
            (header_map.get('汇率'),                'exchange_rate',         'float'),
            (header_map.get('合同额（人民币）'),    'contract_amount_rmb',   'float'),
            (header_map.get('梯种'),                'elevator_type',         'str'),
            (_h('梯型', '类型'),                    'elevator_class',        'str'),
            (header_map.get('产品型号'),            'product_type',          'str'),
            (header_map.get('载重'),                'capacity',              'float'),
            (header_map.get('速度'),                'speed',                 'float'),
            (header_map.get('代理商'),              'agent',                 'str'),
            (header_map.get('产品状态'),            'product_status',        'str'),
            (header_map.get('业务员'),              'salesperson',           'str'),
            (header_map.get('国家'),                'country',               'str'),
            (_h('大区域', '大区'),                   'business_region',       'str'),
            (header_map.get('小区域'),              'sub_region',            'str'),
        ]
        # 日期列
        date_cols = [
            ('sign_date',      header_map.get('签订日期')),
            ('schedule_date',  header_map.get('排产日期')),
            ('delivery_date',  _h('组A日期', '实际发货日期', '实际发运日期')),
        ]
        # 层站门
        floor_cols = [(header_map.get(h), h) for h in ['层数', '站数', '门数']]

        # 快速清空旧数据（只删台账来源，保留报表a/b）
        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(LedgerContract).where(LedgerContract.source == 'ledger'))
        db.session.commit()

        CUTOFF_DATE = date(2013, 1, 1)
        batch = []
        total = 0
        skipped_old = 0

        # iter_rows(values_only=True) 返回 tuple，避免逐个 Cell 对象创建
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 快速日期预检
            sd = _safe_date_openpyxl(row[col_sign_date - 1]) if col_sign_date else None
            scd = _safe_date_openpyxl(row[col_schedule_date - 1]) if col_schedule_date else None
            dd = _safe_date_openpyxl(row[col_delivery_date - 1]) if col_delivery_date else None

            # 三个日期都 < 2025 则跳过
            if not ((sd and sd >= CUTOFF_DATE) or (scd and scd >= CUTOFF_DATE) or (dd and dd >= CUTOFF_DATE)):
                skipped_old += 1
                continue

            # 合同号
            cn = _safe_str(row[col_contract_no - 1]) if col_contract_no else ''
            if not cn:
                continue

            row_data = {'source': 'ledger', 'sign_date': sd, 'schedule_date': scd, 'delivery_date': dd}

            for col_idx, field_name, ftype in col_field_map:
                if col_idx is None:
                    continue
                val = row[col_idx - 1]
                if ftype == 'int':
                    row_data[field_name] = _safe_int(val)
                elif ftype == 'float':
                    row_data[field_name] = _safe_float(val)
                else:
                    row_data[field_name] = _safe_str(val)

            # 层站门
            floor_parts = []
            for ci, _ in floor_cols:
                if ci:
                    fv = _safe_str(row[ci - 1])
                    if fv and fv != '0':
                        floor_parts.append(fv)
            if floor_parts:
                row_data['floors'] = '/'.join(floor_parts)

            batch.append(LedgerContract(**row_data))

            if len(batch) >= 2000:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                total += len(batch)
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)
            db.session.commit()
            total += len(batch)

        wb.close()
        msg = f"台账导入 {total} 条 (跳过 {skipped_old} 条2025年前旧合同)"
        return {"success": True, "message": msg, "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"台账解析失败: {str(e)}", "rows": 0}


# ── Handler 2: 国家映射表 ─────────────────────────────────

def _is_invalid(v):
    """判断单元格值是否为无效映射（#N/A / 0 / None / 空）"""
    if v is None:
        return True
    if v == 0:
        return True
    s = str(v).strip()
    return s in ('#N/A', '')


def parse_country_mapping(file_path: str) -> dict:
    """解析国家-市场-业务员.xlsx Sheet3 → cc_country_mapping（全量替换）

    Sheet3 结构: A=国家, B=2025模块, C=2025大区, D=2026模块, E=2026大区
    使用 D/E 列（2026年映射）作为国家→模块/大区的对应关系。

    健壮性设计：
    - 重复国家：保留第一次出现，跳过后续，记录警告
    - D/E 无效（#N/A/0/None/空）：跳过该行，输出到 unmatched_mapping.txt
    - Sheet 检测：优先匹配 Sheet3，回退到第3个sheet
    - 列检测：先按固定位置，表头不对则自动探测
    """
    uploads_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'uploads'
    )
    os.makedirs(uploads_dir, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # ── 定位 Sheet ──
        ws = None
        for sn in wb.sheetnames:
            if sn == 'Sheet3':
                ws = wb[sn]
                break
        if ws is None:
            for sn in wb.sheetnames:
                if 'Sheet3' in sn:
                    ws = wb[sn]
                    break
        if ws is None and len(wb.sheetnames) >= 3:
            ws = wb[wb.sheetnames[2]]
        if ws is None:
            ws = wb.active

        # ── 定位列 ──
        col_country = 1   # A列
        col_module = 4    # D列（2026年对应模块）
        col_region = 5    # E列（2026九大区）

        header_a = _safe_str(ws.cell(row=1, column=1).value)
        header_d = _safe_str(ws.cell(row=1, column=4).value)
        # 如果表头不符合预期，自动探测
        if header_a and '国家' not in header_a:
            for c in range(1, min(ws.max_column + 1, 10)):
                v = _safe_str(ws.cell(row=1, column=c).value)
                if '国家' in v:
                    col_country = c
                elif '2026' in v and '模块' in v:
                    col_module = c
                elif '2026' in v and '大区' in v:
                    col_region = c

        # ── 清空并重新导入（同一事务）──
        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(CountryMapping))

        total = 0
        skipped_invalid = []   # D/E 无效的
        skipped_dup = []       # 重复国家

        seen_countries = set()

        for r in range(2, ws.max_row + 1):
            raw_country = ws.cell(row=r, column=col_country).value
            country = _safe_str(raw_country)
            if not country:
                continue

            raw_d = ws.cell(row=r, column=col_module).value
            raw_e = ws.cell(row=r, column=col_region).value

            # 检查 D/E 无效 → 记录并跳过
            if _is_invalid(raw_d) or _is_invalid(raw_e):
                skipped_invalid.append((country, raw_d, raw_e))
                continue

            # 检查重复国家 → 保留第一个，跳过后续
            if country in seen_countries:
                skipped_dup.append(country)
                continue
            seen_countries.add(country)

            mapping = CountryMapping(
                country=country,
                module_name=_safe_str(raw_d),
                region=_safe_str(raw_e),
                module_manager='',
                salesperson='',
            )
            db.session.add(mapping)
            total += 1

        db.session.commit()
        wb.close()

        # ── 输出报告文件 ──
        report_lines = []
        report_lines.append(f"国家映射导入报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        report_lines.append(f"源文件Sheet3: A=国家 D=2026模块 E=2026大区")
        report_lines.append("=" * 60)
        report_lines.append(f"成功导入: {total} 条")
        report_lines.append(f"重复跳过: {len(skipped_dup)} 条（同名国家保留首次出现）")
        report_lines.append(f"无效跳过: {len(skipped_invalid)} 条（D/E为#N/A/0/空）")
        report_lines.append("")

        if skipped_dup:
            report_lines.append(f"[重复国家]")
            for c in skipped_dup:
                report_lines.append(f"  {c}")
            report_lines.append("")

        if skipped_invalid:
            report_lines.append(f"[无效映射 D/E为#N/A/0/空]")
            report_lines.append(f"{'国家':<30} {'D(2026模块)':<25} {'E(2026大区)'}")
            report_lines.append("-" * 70)
            for country, d_val, e_val in skipped_invalid:
                report_lines.append(f"{country:<30} {str(d_val):<25} {str(e_val)}")
            report_lines.append("")

        filepath = os.path.join(uploads_dir, 'mapping_report.txt')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report_lines))

        # ── 返回 ──
        parts = [f"导入 {total} 条国家映射"]
        if skipped_dup:
            parts.append(f"{len(skipped_dup)} 条重复跳过")
        if skipped_invalid:
            parts.append(f"{len(skipped_invalid)} 条无效跳过")
        return {
            "success": True,
            "message": "，".join(parts),
            "rows": total,
        }

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"映射表解析失败: {str(e)}", "rows": 0}


# ── Handler 3: 回款明细 ───────────────────────────────────

def parse_payment_collections(file_path: str) -> dict:
    """解析 gm_ht_hthkmx.xls → cc_payment_collection（全量替换）"""
    try:
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # 列位置映射（基于实际文件结构，0-indexed）
        # col0=合同编号, col1=梯号, col2=项目名称, col3=签订日期
        # col4=合同类型, col5=产品状态, col6=大区, col7=小区域
        # col8=节点类型, col9=付款方式, col10=回款额(原币), col11=型号
        # col12=代理商, col13=回款额(人民币), col14=回款日期, col15=币种, col16=备注

        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(PaymentCollection))
        db.session.commit()

        batch = []
        total = 0
        skipped_agent = 0
        skipped_node_type = 0
        for r in range(1, ws.nrows):  # 跳过第0行（标题行）
            contract_no = _safe_str(ws.cell_value(r, 0))
            if not contract_no:
                # 可能是表头行（"合同编号"）
                if '合同' in str(ws.cell_value(r, 0)):
                    continue
                continue

            # 过滤1：合同编号以DHN或YBN开头的跳过
            if contract_no.startswith('DHN') or contract_no.startswith('YBN'):
                continue

            # 过滤2：代理商为"新加坡分公司(关联方)"的行不参与匹配
            agent = _safe_str(ws.cell_value(r, 12))
            if agent == '新加坡分公司(关联方)':
                skipped_agent += 1
                continue

            # 过滤3：款项类型(I列)为"低佣金"或"手续费"的跳过
            node_type = _safe_str(ws.cell_value(r, 8))
            if node_type in ('抵佣金', '手续费'):
                skipped_node_type += 1
                continue

            pc = PaymentCollection(
                contract_no=contract_no,
                ladder_no=_safe_str(ws.cell_value(r, 1)),
                project_name=_safe_str(ws.cell_value(r, 2)),
                sign_date=_safe_date_xlrd(wb, ws.cell_value(r, 3)),
                contract_type=_safe_str(ws.cell_value(r, 4)),
                product_status=_safe_str(ws.cell_value(r, 5)),
                region=_safe_str(ws.cell_value(r, 6)),
                sub_region=_safe_str(ws.cell_value(r, 7)),
                node_type=_safe_str(ws.cell_value(r, 8)),
                payment_method=_safe_str(ws.cell_value(r, 9)),
                payment_amount_orig=_safe_float(ws.cell_value(r, 10)),
                product_model=_safe_str(ws.cell_value(r, 11)),
                agent_name=_safe_str(ws.cell_value(r, 12)),
                payment_amount_rmb=_safe_float(ws.cell_value(r, 13)),
                payment_date=_safe_date_xlrd(wb, ws.cell_value(r, 14)),
                currency=_safe_str(ws.cell_value(r, 15)),
                remark=_safe_str(ws.cell_value(r, 16)),
            )
            batch.append(pc)

            if len(batch) >= 500:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                total += len(batch)
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)
            db.session.commit()
            total += len(batch)

        wb.release_resources()
        msg = f"回款导入 {total} 条记录"
        parts = []
        if skipped_agent:
            parts.append(f"新加坡关联方 {skipped_agent} 条")
        if skipped_node_type:
            parts.append(f"低佣金/手续费 {skipped_node_type} 条")
        if parts:
            msg += "（跳过：" + "，".join(parts) + "）"
        return {"success": True, "message": msg, "rows": total}

    except Exception as e:
        db.session.rollback()
        # 尝试用 openpyxl 回退
        try:
            return _parse_payment_openpyxl(file_path)
        except Exception as e2:
            return {"success": False, "message": f"回款解析失败: {str(e)}", "rows": 0}


def _parse_payment_openpyxl(file_path: str) -> dict:
    """回退：尝试以 .xlsx 格式解析回款文件"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    from sqlalchemy import delete as sa_delete
    db.session.execute(sa_delete(PaymentCollection))
    db.session.commit()

    batch = []
    total = 0
    skipped_agent = 0
    skipped_node_type = 0
    for r in range(2, ws.max_row + 1):
        contract_no = _safe_str(ws.cell(row=r, column=1).value)
        if not contract_no:
            continue

        # 过滤1：合同编号以DHN或YBN开头的跳过
        if contract_no.startswith('DHN') or contract_no.startswith('YBN'):
            continue

        # 过滤2：代理商为"新加坡分公司(关联方)"的行不参与匹配
        agent = _safe_str(ws.cell(row=r, column=13).value)
        if agent == '新加坡分公司(关联方)':
            skipped_agent += 1
            continue

        # 过滤3：款项类型(I列=column9)为"低佣金"或"手续费"的跳过
        node_type = _safe_str(ws.cell(row=r, column=9).value)
        if node_type in ('低佣金', '手续费'):
            skipped_node_type += 1
            continue

        pc = PaymentCollection(
            contract_no=contract_no,
            ladder_no=_safe_str(ws.cell(row=r, column=2).value),
            project_name=_safe_str(ws.cell(row=r, column=3).value),
            sign_date=_safe_date_openpyxl(ws.cell(row=r, column=4).value),
            contract_type=_safe_str(ws.cell(row=r, column=5).value),
            product_status=_safe_str(ws.cell(row=r, column=6).value),
            region=_safe_str(ws.cell(row=r, column=7).value),
            sub_region=_safe_str(ws.cell(row=r, column=8).value),
            node_type=_safe_str(ws.cell(row=r, column=9).value),
            payment_method=_safe_str(ws.cell(row=r, column=10).value),
            payment_amount_orig=_safe_float(ws.cell(row=r, column=11).value),
            product_model=_safe_str(ws.cell(row=r, column=12).value),
            agent_name=_safe_str(ws.cell(row=r, column=13).value),
            payment_amount_rmb=_safe_float(ws.cell(row=r, column=14).value),
            payment_date=_safe_date_openpyxl(ws.cell(row=r, column=15).value),
            currency=_safe_str(ws.cell(row=r, column=16).value),
            remark=_safe_str(ws.cell(row=r, column=17).value),
        )
        batch.append(pc)
        if len(batch) >= 500:
            db.session.bulk_save_objects(batch)
            db.session.commit()
            total += len(batch)
            batch = []

    if batch:
        db.session.bulk_save_objects(batch)
        db.session.commit()
        total += len(batch)

    wb.close()
    msg = f"回款导入 {total} 条记录 (xlsx)"
    parts = []
    if skipped_agent:
        parts.append(f"新加坡关联方 {skipped_agent} 条")
    if skipped_node_type:
        parts.append(f"低佣金/手续费 {skipped_node_type} 条")
    if parts:
        msg += "（跳过：" + "，".join(parts) + "）"
    return {"success": True, "message": msg, "rows": total}


# ── Handler 4: 报表a（商贸配件签单+排产） ──────────────

# 商贸配件6模块（个人业绩跳过集合，含名称变体）
_TRADE_SKIP_MODULES = {
    '商贸-1', '商贸1', '商贸-2', '商贸2', '商贸-3', '商贸3',
    '配件-1', '配件-2', '改造', '更新改造',
}


def _build_module_region_map():
    """从 ANNUAL_TARGETS 建立 模块 -> 大区 映射（判定报表a个人业绩归属哪个大区）"""
    from dashboards.contract_completion.services import ANNUAL_TARGETS
    return {module: region for (region, module) in ANNUAL_TARGETS}


def _load_person_module_map():
    """读取 数据源/模块对应表.xlsx 的 D(姓名)->E(模块) 映射，返回 dict；失败返回空"""
    import openpyxl
    here = os.path.abspath(os.path.dirname(__file__))                  # backend/dashboards/contract_completion
    base = os.path.dirname(os.path.dirname(os.path.dirname(here)))     # 项目根
    path = os.path.join(base, '数据源', '模块对应表.xlsx')
    if not os.path.isfile(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb['整梯模块对应表']
        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = row[3] if len(row) > 3 else None   # D列=姓名
            e = row[4] if len(row) > 4 else None   # E列=模块
            if d and str(d).strip():
                mapping[str(d).strip()] = str(e).strip() if e is not None else ''
        return mapping
    finally:
        wb.close()


def _resolve_personal_target(person2mod, mod2region, person_name):
    """根据报表a col56 人名，返回 (module, region)；商贸配件/无法判定时返回 (None, None)"""
    target_mod = person2mod.get(person_name, '')
    if not target_mod or target_mod in _TRADE_SKIP_MODULES:
        return None, None
    region = mod2region.get(target_mod)
    if not region or region == '商贸配件':
        return None, None
    return target_mod, region


def parse_report_a(file_path: str) -> dict:
    """解析报表a.xls → 提取配件-1/配件-2的签单额、排产额 → 写入ledger表"""
    try:
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # 列映射（0-indexed）:
        # G(col6): 合同额（人民币）
        # K(col10): 签订日期
        # P(col15): 排产日期
        # BC(col54): 模块

        # 个人业绩归属：加载 姓名->模块 与 模块->大区 映射
        person2mod = _load_person_module_map()
        mod2region = _build_module_region_map()

        trade_rows = []
        for r in range(1, ws.nrows):
            module_name = _safe_str(ws.cell_value(r, 54))
            if module_name not in ('配件-1', '配件-2'):
                continue

            sign_date = _safe_date_xlrd(wb, ws.cell_value(r, 10))
            prod_date = _safe_date_xlrd(wb, ws.cell_value(r, 15))
            amount_rmb = _safe_float(ws.cell_value(r, 6))  # 合同额（人民币），单位：元

            project_name = _safe_str(ws.cell_value(r, 2)) if ws.ncols > 2 else ''
            contract_no = _safe_str(ws.cell_value(r, 0))

            # 个人业绩归属：col56 订单备注（人名）→ 目标模块/大区
            person_name = _safe_str(ws.cell_value(r, 56))
            personal_module, personal_region = _resolve_personal_target(
                person2mod, mod2region, person_name)

            trade_rows.append({
                'contract_no': contract_no or f'REPORT_A_{r}',
                'project_name': project_name,
                'module_name': module_name,
                'sign_date': sign_date,
                'schedule_date': prod_date,
                'contract_amount_rmb': amount_rmb,
                'personal_module': personal_module,
                'personal_region': personal_region,
            })

        # 删除之前的报表a来源数据
        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(LedgerContract).where(LedgerContract.source == 'report_a'))

        # 写入ledger表（保留2025+，不做年份过滤，查询时按年筛选）
        total = 0
        batch = []
        for tr in trade_rows:
            # 签单行
            if tr['sign_date'] and tr['sign_date'] >= date(2025, 1, 1):
                lc = LedgerContract(
                    source='report_a',
                    contract_no=tr['contract_no'],
                    project_name=tr['project_name'],
                    mapped_module=tr['module_name'],
                    mapped_region='商贸合计',
                    product_status='',
                    sign_date=tr['sign_date'],
                    contract_amount_rmb=tr['contract_amount_rmb'],
                    unit_count=0,
                    personal_module=tr['personal_module'],
                    personal_region=tr['personal_region'],
                )
                batch.append(lc)
                total += 1

            # 排产行
            if tr['schedule_date'] and tr['schedule_date'] >= date(2025, 1, 1):
                lc = LedgerContract(
                    source='report_a',
                    contract_no=tr['contract_no'],
                    project_name=tr['project_name'],
                    mapped_module=tr['module_name'],
                    mapped_region='商贸合计',
                    product_status='',
                    schedule_date=tr['schedule_date'],
                    contract_amount_rmb=tr['contract_amount_rmb'],
                    unit_count=0,
                    personal_module=tr['personal_module'],
                    personal_region=tr['personal_region'],
                )
                batch.append(lc)
                total += 1

            if len(batch) >= 500:
                db.session.bulk_save_objects(batch)
                db.session.flush()
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)

        db.session.commit()
        wb.release_resources()
        return {"success": True, "message": f"报表a导入 {total} 条配件记录", "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"报表a解析失败: {str(e)}", "rows": 0}


# ── Handler 5: 报表b（商贸配件发货） ──────────────

def parse_report_b(file_path: str) -> dict:
    """解析报表b.xls → 提取配件-1/配件-2的发货额 → 写入ledger表"""
    try:
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # 列映射（0-indexed）:
        # G(col6): 实际发运时间
        # M(col12): 总金额（人民币）
        # P(col15): 模块

        trade_rows = []
        for r in range(1, ws.nrows):
            module_name = _safe_str(ws.cell_value(r, 15))
            if module_name not in ('配件-1', '配件-2'):
                continue

            ship_date = _safe_date_xlrd(wb, ws.cell_value(r, 6))
            amount_rmb = _safe_float(ws.cell_value(r, 12))  # 总金额（人民币），单位：元

            project_name = _safe_str(ws.cell_value(r, 2)) if ws.ncols > 2 else ''
            contract_no = _safe_str(ws.cell_value(r, 0))

            trade_rows.append({
                'contract_no': contract_no or f'REPORT_B_{r}',
                'project_name': project_name,
                'module_name': module_name,
                'delivery_date': ship_date,
                'contract_amount_rmb': amount_rmb,
            })

        # 删除之前的报表b来源数据
        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(LedgerContract).where(LedgerContract.source == 'report_b'))

        total = 0
        batch = []
        for tr in trade_rows:
            if tr['delivery_date'] and tr['delivery_date'] >= date(2025, 1, 1):
                lc = LedgerContract(
                    source='report_b',
                    contract_no=tr['contract_no'],
                    project_name=tr['project_name'],
                    mapped_module=tr['module_name'],
                    mapped_region='商贸合计',
                    product_status='',
                    delivery_date=tr['delivery_date'],
                    contract_amount_rmb=tr['contract_amount_rmb'],
                    unit_count=0,
                )
                batch.append(lc)
                total += 1

            if len(batch) >= 500:
                db.session.bulk_save_objects(batch)
                db.session.flush()
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)

        db.session.commit()
        wb.release_resources()
        return {"success": True, "message": f"报表b导入 {total} 条配件发货记录", "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"报表b解析失败: {str(e)}", "rows": 0}


# ── Handler 6: 海外差值 ─────────────────────────────────

def parse_overseas_diff(file_path: str) -> dict:
    """解析海外差值.xlsx → cc_overseas_diff（按年份全量替换）

    Excel 结构（Sheet1），支持多年份：
      Row 1: [空, 空, 模块1, 模块2, ..., 总]
      后续按年分块，每块4行：
        [2026.7, 签订, val1, val2, ..., total]
        [空,    排产, val1, val2, ..., total]
        [空,    发货, val1, val2, ..., total]
        [空,    回款, val1, val2, ..., total]
        [2025.7, 签订, val1, val2, ..., total]
        ...

    年份从 Col A 提取（如 "2026.7" → 2026），Col B 标识指标类型。
    指标 → 模型字段: 签订→sign_diff, 排产→schedule_diff, 发货→ship_diff, 回款→payment_diff
    单位：万元（Excel值直接使用）。跳过"总"列。
    """
    from dashboards.contract_completion.models import OverseasDiff
    import re

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active  # Sheet1

        # ── Row 1: 模块名列表（Col C 起，跳过"总"） ──
        modules = []
        for c in range(3, ws.max_column + 1):
            v = _safe_str(ws.cell(row=1, column=c).value)
            if v and v != '总':
                modules.append((c, v))

        if not modules:
            wb.close()
            return {"success": False, "message": "未找到模块列（Row1, Col C+）", "rows": 0}

        # 指标标识映射（Col B）
        metric_label_map = {
            '签订': 'sign_diff',
            '排产': 'schedule_diff',
            '发货': 'ship_diff',
            '回款': 'payment_diff',
        }

        # ── 按年分块扫描 ──
        # data_by_year: {2026: {module_name: {sign_diff, schedule_diff, ship_diff, payment_diff}}}
        data_by_year = {}
        current_year = None

        for r in range(2, ws.max_row + 1):
            # 检测 Col A 是否为年份标记（如 "2026.7"）
            raw_a = ws.cell(row=r, column=1).value
            if raw_a is not None and str(raw_a).strip():
                # 提取年份数字
                year_match = re.search(r'(\d{4})', str(raw_a))
                if year_match:
                    current_year = int(year_match.group(1))
                    if current_year not in data_by_year:
                        data_by_year[current_year] = {
                            m[1]: {'sign_diff': None, 'schedule_diff': None,
                                   'ship_diff': None, 'payment_diff': None}
                            for m in modules
                        }

            if current_year is None:
                continue

            # 读取 Col B 确认指标类型
            label = _safe_str(ws.cell(row=r, column=2).value)
            field = metric_label_map.get(label)
            if not field:
                continue

            # 读取各模块的值
            for col, module_name in modules:
                val = ws.cell(row=r, column=col).value
                if val is not None and str(val).strip() not in ('', '-'):
                    try:
                        data_by_year[current_year][module_name][field] = round(float(val), 4)
                    except (ValueError, TypeError):
                        pass

        wb.close()

        # ── 按年份全量替换并写入 ──
        total = 0
        messages = []
        for year, module_data in sorted(data_by_year.items()):
            # 删除该年份旧数据
            from sqlalchemy import delete as sa_delete
            db.session.execute(sa_delete(OverseasDiff).where(OverseasDiff.data_year == year))

            year_total = 0
            for module_name, fields in module_data.items():
                # 如果所有指标都为空则跳过
                if all(v is None for v in fields.values()):
                    continue
                od = OverseasDiff(
                    data_year=year,
                    module_name=module_name,
                    sign_diff=fields['sign_diff'],
                    schedule_diff=fields['schedule_diff'],
                    ship_diff=fields['ship_diff'],
                    payment_diff=fields['payment_diff'],
                )
                db.session.add(od)
                year_total += 1

            total += year_total
            messages.append(f"{year}年{year_total}条")

        db.session.commit()
        return {
            "success": True,
            "message": f"海外差值导入 {total} 条（{', '.join(messages)}）",
            "rows": total,
        }

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"海外差值解析失败: {str(e)}", "rows": 0}


# ── Handler 7: 商贸模块汇总数据 ─────────────────────────

def _parse_trade_data_impl(file_path: str, data_year: int) -> dict:
    """解析商贸数据.xlsx → cc_trade_module_data（全量替换指定年份）

    Excel结构（Sheet1）:
      A=模块, B=签订额(万元), C=回款额(万元), D=排产额(万元), E=发货额(万元)
    """
    from dashboards.contract_completion.models import TradeModuleData

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active  # Sheet1

        # 读取表头（行1）定位列
        col_map = {}
        for c in range(1, ws.max_column + 1):
            v = _safe_str(ws.cell(row=1, column=c).value)
            if '模块' in v:
                col_map['module'] = c
            elif '签订' in v:
                col_map['sign'] = c
            elif '回款' in v:
                col_map['payment'] = c
            elif '排产' in v:
                col_map['schedule'] = c
            elif '发货' in v:
                col_map['ship'] = c

        # 先扫描 Excel 收集模块名，避免误删其他来源的同年数据（如配件回款）
        module_names_in_file = set()
        for r in range(2, ws.max_row + 1):
            mn = _safe_str(ws.cell(row=r, column=col_map.get('module', 1)).value)
            if mn and any(kw in mn for kw in ('商贸', '配件')):
                module_names_in_file.add(mn)

        # 只删除本文件涉及的模块（不影响配件回款等其他来源）
        from sqlalchemy import delete as sa_delete
        for mn in module_names_in_file:
            db.session.execute(
                sa_delete(TradeModuleData).where(
                    TradeModuleData.data_year == data_year,
                    TradeModuleData.module_name == mn
                )
            )
        # 不 commit — 与后续插入在同一事务中

        total = 0
        for r in range(2, ws.max_row + 1):
            module_name = _safe_str(ws.cell(row=r, column=col_map.get('module', 1)).value)
            if not module_name:
                continue
            # 处理商贸/配件模块
            if not any(kw in module_name for kw in ('商贸', '配件')):
                continue

            def _get(col_key):
                c = col_map.get(col_key)
                return _safe_float(ws.cell(row=r, column=c).value) if c else 0.0

            # Excel值(万元) × 10000 → 元（与台账数据单位一致）
            td = TradeModuleData(
                data_year=data_year,
                module_name=module_name,
                sign_amount=_get('sign') * 10000,
                payment_amount=_get('payment') * 10000,
                schedule_amount=_get('schedule') * 10000,
                ship_amount=_get('ship') * 10000,
            )
            db.session.add(td)
            total += 1

        db.session.commit()
        wb.close()
        return {"success": True, "message": f"商贸数据导入 {total} 条（{data_year}年）", "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"商贸数据解析失败: {str(e)}", "rows": 0}


def parse_trade_data(file_path: str) -> dict:
    """解析商贸数据.xlsx → 2026年"""
    return _parse_trade_data_impl(file_path, 2026)


def parse_trade_data_2025(file_path: str) -> dict:
    """解析商贸数据.xlsx → 2025年"""
    return _parse_trade_data_impl(file_path, 2025)


# ── Handler 8: 配件回款 ─────────────────────────────────

def parse_accessories_payment(file_path: str) -> dict:
    """解析配件回款.xls → 筛选回款日期>=2026-01-01的配件-1/配件-2回款
    按模块汇总G列(回款额人民币,元)，写入 TradeModuleData（仅payment字段）
    """
    from dashboards.contract_completion.models import TradeModuleData
    from datetime import date

    try:
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        # 2026-01-01 Excel serial number
        cutoff_serial = 46023

        totals = {'配件-1': 0.0, '配件-2': 0.0}

        for r in range(3, ws.nrows):  # 跳过标题行
            module = _safe_str(ws.cell_value(r, 8))  # I列：模块
            if module not in ('配件-1', '配件-2'):
                continue

            pay_serial = ws.cell_value(r, 7)  # H列：回款日期
            if not pay_serial or pay_serial < cutoff_serial:
                continue

            amount = ws.cell_value(r, 6)  # G列：回款额(人民币) 元
            if amount:
                totals[module] += amount

        wb.release_resources()

        # 仅删除配件-1/配件-2的2026年记录（不影响商贸1/2/3等）
        from sqlalchemy import delete as sa_delete
        db.session.execute(
            sa_delete(TradeModuleData).where(
                TradeModuleData.data_year == 2026,
                TradeModuleData.module_name.in_(['配件-1', '配件-2'])
            )
        )
        # 不 commit — 与后续插入在同一事务中

        total = 0
        for module_name, amount_yuan in totals.items():
            if amount_yuan > 0:
                td = TradeModuleData(
                    data_year=2026,
                    module_name=module_name,
                    sign_amount=0,
                    schedule_amount=0,
                    ship_amount=0,
                    payment_amount=amount_yuan,  # 单位：元
                )
                db.session.add(td)
                total += 1

        db.session.commit()

        wan_1 = round(totals['配件-1'] / 10000)
        wan_2 = round(totals['配件-2'] / 10000)
        msg = f"配件回款导入 {total} 条（2026年）：配件-1={wan_1}万元，配件-2={wan_2}万元"
        return {"success": True, "message": msg, "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"配件回款解析失败: {str(e)}", "rows": 0}


# ── Handler 9: 工期统计 ─────────────────────────────────

# 梯种分类边界: row < BOUNDARY → 直梯/扶梯, row >= BOUNDARY → 改造/外购
_SCHEDULE_CATEGORY_BOUNDARY = 609


def _classify_delay(days):
    """延期分类: early / ontime / minor / major / pending"""
    if days is None:
        return "pending"
    if days < 0:
        return "early"
    if days == 0:
        return "ontime"
    if days <= 5:
        return "minor"
    return "major"


def _is_overdue_warehouse_text(s):
    """判断 K/M 单元格文本是否为「超期未入库」：
    包含「未入库」，且不包含「未到入库时间」（等待入库）、「已暂停」（暂停）与「变更」（变更）。
    """
    if not s:
        return False
    return '未入库' in s and '未到入库时间' not in s and '已暂停' not in s and '变更' not in s


def _determine_stage(i_date, j_date, k_date, k_raw, m_date, m_raw, o_date, is_rejected):
    """判断合同当前所处阶段（移植自 process_data.py）"""
    if is_rejected:
        return "rejected"

    has_k = k_date is not None
    has_m = m_date is not None

    if has_k and has_m:
        if o_date:
            return "completed"
        return "warehoused"
    elif _is_overdue_warehouse_text(k_raw) or _is_overdue_warehouse_text(m_raw):
        return "overdue_warehouse"
    elif has_k or has_m:
        return "partial_warehouse"
    elif k_raw == "未到入库时间" or m_raw == "未到入库时间":
        return "waiting_warehouse"
    else:
        if j_date:
            return "design_done"
        if i_date:
            return "order_done"
        return "pending"


def parse_schedule_tracking(file_path: str) -> dict:
    """解析工期统计表 Excel → cc_schedule_tracking（全量替换）
    移植自 Project_Schedule/scripts/process_data.py
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # 从文件名提取日期（如 20260804）
        date_match = re.search(r'(\d{8})', os.path.basename(file_path))
        file_date = datetime.strptime(date_match.group(1), '%Y%m%d').date() if date_match else None

        # 预加载国家映射表
        mappings = {m.country: m for m in CountryMapping.query.all()}

        # 全量替换
        from sqlalchemy import delete as sa_delete
        db.session.execute(sa_delete(ScheduleTracking))
        db.session.commit()

        total = 0
        skipped = 0

        for r, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            contract_no = _safe_str(row[1])
            if not contract_no:
                skipped += 1
                continue

            # ── 梯种分类 ──
            elevator_type = _safe_str(row[4])
            if r < _SCHEDULE_CATEGORY_BOUNDARY:
                category = elevator_type if elevator_type else "直梯/扶梯"
            else:
                category = "改造/外购"

            # ── 日期列 ──
            i_date = _safe_date_openpyxl(row[8])
            j_date = _safe_date_openpyxl(row[9])
            k_raw = row[10]
            m_raw = row[12]
            o_date = _safe_date_openpyxl(row[14])

            # K/M 日期处理（"未到入库时间"/"未入库" → None）
            k_raw_str = _safe_str(k_raw) if k_raw else None
            m_raw_str = _safe_str(m_raw) if m_raw else None
            _k_is_text = k_raw_str and ('未到入库时间' in k_raw_str or '未入库' in k_raw_str)
            _m_is_text = m_raw_str and ('未到入库时间' in m_raw_str or '未入库' in m_raw_str)
            k_date = _safe_date_openpyxl(k_raw) if k_raw and not _k_is_text else None
            m_date = _safe_date_openpyxl(m_raw) if m_raw and not _m_is_text else None

            # 延期天数
            l_val = row[11]
            n_val = row[13]
            try:
                l_days = int(l_val) if l_val is not None else None
            except (ValueError, TypeError):
                l_days = None
            try:
                n_days = int(n_val) if n_val is not None else None
            except (ValueError, TypeError):
                n_days = None

            # 未入库延期天数计算：L/N为空 + K/M为"超期未入库"文本 → 文件名日期 - J日期
            # 注意："未到入库时间"（等待入库）、"已暂停"（暂停）与"变更"（变更）都不算"未入库"，必须排除
            if l_days is None and j_date and _is_overdue_warehouse_text(k_raw_str) and file_date:
                l_days = (file_date - j_date).days
            if n_days is None and j_date and _is_overdue_warehouse_text(m_raw_str) and file_date:
                n_days = (file_date - j_date).days

            # 驳回
            is_rejected = _safe_str(row[26]) == "是"

            # 项目名 → 映射
            project_name = _safe_str(row[6])
            mapped_region = ""
            mapped_module = ""
            if project_name and project_name in mappings:
                mapped_region = mappings[project_name].region or ""
                mapped_module = mappings[project_name].module_name or ""
            elif project_name:
                mapped_region = "未分类"
                mapped_module = "未分类"

            # 阶段 & 延期分类
            stage = _determine_stage(i_date, j_date, k_date, k_raw_str, m_date, m_raw_str, o_date, is_rejected)
            l_class = _classify_delay(l_days)
            n_class = _classify_delay(n_days)

            st = ScheduleTracking(
                contract_no=contract_no,
                elevator_no=_safe_str(row[2]),
                project_name=project_name,
                elevator_type=elevator_type,
                elevator_model=_safe_str(row[5]),
                quantity=_safe_int(row[3]),
                person=_safe_str(row[7]),
                order_end_date=i_date,
                design_done_date=j_date,
                mech_warehouse_date=k_date,
                mech_warehouse_raw=k_raw_str if _is_overdue_warehouse_text(k_raw_str) else None,
                mech_delay_days=l_days,
                elec_warehouse_date=m_date,
                elec_warehouse_raw=m_raw_str if _is_overdue_warehouse_text(m_raw_str) else None,
                elec_delay_days=n_days,
                audit_done_date=o_date,
                designer_mech=_safe_str(row[19]),
                designer_elec=_safe_str(row[20]),
                design_cycle=_safe_int(row[21]),
                production_cycle=_safe_int(row[23]),
                total_cycle=_safe_int(row[25]),
                is_rejected=is_rejected,
                reject_node=_safe_str(row[27]),
                reject_reason=_safe_str(row[28]),
                scheduled_finish=_safe_date_openpyxl(row[29]),
                delivery_finish=_safe_date_openpyxl(row[30]),
                remark=_safe_str(row[31]),
                mapped_region=mapped_region,
                mapped_module=mapped_module,
                stage=stage,
                category=category,
                l_class=l_class,
                n_class=n_class,
                data_date=file_date,
            )
            db.session.add(st)
            total += 1

        db.session.commit()
        wb.close()
        return {"success": True, "message": f"工期统计导入 {total} 条（跳过 {skipped} 空行）", "rows": total}

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"工期统计解析失败: {str(e)}", "rows": 0}
