"""
合同完成情况表 — 业务逻辑服务层

指标计算、数据聚合、未匹配检测
"""
import os
from datetime import date, datetime, timedelta
from flask import current_app
from sqlalchemy import or_
from models import db
from dashboards.contract_completion.models import (
    LedgerContract, CountryMapping, PaymentCollection, AnnualTarget
)

# ── 缓存 ──────────────────────────────────────────────────

_two_year_cache = None          # 两年对比表缓存结果
_two_year_fingerprint = None    # 缓存时的数据指纹


def _get_data_fingerprint():
    """快速生成数据指纹（轻量 COUNT 查询），用于判断缓存是否失效"""
    from sqlalchemy import or_
    from dashboards.contract_completion.models import TradeModuleData, OverseasDiff
    ledger_count = LedgerContract.query.filter(
        LedgerContract.source.in_(['ledger', 'report_a', 'report_b']),
        or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')
    ).count()
    mapping_count = CountryMapping.query.count()
    payment_count = PaymentCollection.query.count()
    trade_count = TradeModuleData.query.count()
    overseas_count = OverseasDiff.query.count()
    today_str = (date.today() - timedelta(days=1)).isoformat()
    return f"{ledger_count}|{mapping_count}|{payment_count}|{trade_count}|{overseas_count}|{today_str}"


def invalidate_two_year_cache():
    """手动清除两年对比缓存（数据上传后可由外部调用）"""
    global _two_year_cache, _two_year_fingerprint
    _two_year_cache = None
    _two_year_fingerprint = None


# ── 常量 ──────────────────────────────────────────────────

REGION_ORDER = ['俄罗斯', '中亚', '亚洲1', '亚洲2', '美洲', '中东', '非洲', '欧洲', '商贸合计']

METRIC_CONFIG = [
    {"id": "sign_units",      "name": "年签单台",   "unit": "台",   "type": "count"},
    {"id": "sign_amount",     "name": "年签单额",   "unit": "万元", "type": "amount"},
    {"id": "payment_amount",  "name": "回款",       "unit": "万元", "type": "amount"},
    {"id": "ship_units",      "name": "发货台",     "unit": "台",   "type": "count"},
    {"id": "ship_amount",     "name": "发货额",     "unit": "万元", "type": "amount"},
    {"id": "schedule_units",  "name": "排产台",     "unit": "台",   "type": "count"},
    {"id": "schedule_amount", "name": "排产额",     "unit": "万元", "type": "amount"},
]

RATIO_CONFIG = {
    "sign_units":      "sign_units",
    "sign_amount":     "sign_amount",
    "payment_amount":  "payment_amount",
    "ship_units":      "ship_units",
    "ship_amount":     "ship_amount",
    "schedule_units":  "schedule_units",
    "schedule_amount": "schedule_amount",
}

TRADE_MODULES = ['商贸1', '商贸2', '商贸3', '配件-1', '配件-2', '改造']


# ── 工具函数 ──────────────────────────────────────────────

def _to_wan(val):
    """元 → 万元（保留2位小数）"""
    if val is None:
        return 0.0
    return round(float(val) / 10000, 2)


def _safe_ratio(numerator, denominator):
    """安全除法"""
    if not denominator or abs(denominator) < 0.001:
        return 0.0 if (not numerator or abs(numerator) < 0.001) else 1.0
    return round(float(numerator) / float(denominator), 4)


def _load_mapping():
    """加载国家映射表 → {country: {region, module, manager, salesperson}}"""
    mappings = CountryMapping.query.all()
    return {m.country: {
        'region': '商贸合计' if m.region == '商贸配件' else m.region,
        'module': m.module_name,
        'manager': m.module_manager, 'salesperson': m.salesperson,
    } for m in mappings}


def _load_targets(year, region_only=True):
    """加载年度指标 → {(region, module, metric): value}"""
    query = AnnualTarget.query.filter_by(target_year=year)
    if region_only:
        query = query.filter(AnnualTarget.module_name == '')
    targets = query.all()
    result = {}
    for t in targets:
        key = (t.region, t.module_name, t.metric_key)
        result[key] = t.target_value
    return result


def _resolve_contract(c, mapping, unmatched_list):
    """解析合同 → (region, module, manager, salesperson) 或加入unmatched"""
    # 商贸来源的行已经有 mapped_region 和 mapped_module
    if c.source in ('report_a', 'report_b'):
        return (c.mapped_region or '商贸合计', c.mapped_module or '',
                '', '')

    country = c.country
    if not country:
        unmatched_list.append({'contract_no': c.contract_no, 'country': '(空)',
                               'project_name': c.project_name or ''})
        return None

    cm = mapping.get(country)
    if cm is None:
        unmatched_list.append({'contract_no': c.contract_no, 'country': country,
                               'project_name': c.project_name or ''})
        return None

    region = cm['region']
    module = cm['module']
    # "商贸配件"大区归一化到"商贸合计"
    if region == '商贸配件':
        region = '商贸合计'
    # 产品型号含"改造" → 强制模块="改造"
    if c.product_type and '改造' in str(c.product_type):
        module = '改造'
        region = '商贸合计'

    return (region, module, cm['manager'], cm['salesperson'])


# ── 核心计算 ──────────────────────────────────────────────

def _compute_aggregations(year, mapping, unmatched_list):
    """
    遍历台账合同，按大区和模块聚合指标。
    返回: (region_data, module_data, salesperson_data)
    """
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)

    # region_data: {region: {metric_key: value}}
    region_data = {}
    # module_data: {(region, module_name): {metrics...}}
    module_data = {}
    # sp_data: {salesperson: {metrics...}}
    sp_data = {}

    def _init_metrics():
        return {m['id']: 0.0 for m in METRIC_CONFIG}

    contracts = LedgerContract.query.filter(
        or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')
    ).all()

    for c in contracts:
        resolved = _resolve_contract(c, mapping, unmatched_list)
        if resolved is None:
            continue
        region, module, manager, sp = resolved

        # 初始化
        if region not in region_data:
            region_data[region] = _init_metrics()
        mod_key = (region, module)
        if mod_key not in module_data:
            module_data[mod_key] = {
                'region': region, 'module': module, 'manager': manager,
                'prev_schedule_not_shipped': 0, **_init_metrics(),
            }

        # 大区级累加
        rd = region_data[region]
        md = module_data[mod_key]

        # 签单：按签订日期
        if c.sign_date and year_start <= c.sign_date <= year_end:
            rd['sign_units'] += c.unit_count or 0
            rd['sign_amount'] += c.contract_amount_rmb or 0
            md['sign_units'] += c.unit_count or 0
            md['sign_amount'] += c.contract_amount_rmb or 0

        # 排产：按排产日期
        if c.schedule_date and year_start <= c.schedule_date <= year_end:
            rd['schedule_units'] += c.unit_count or 0
            rd['schedule_amount'] += c.contract_amount_rmb or 0
            md['schedule_units'] += c.unit_count or 0
            md['schedule_amount'] += c.contract_amount_rmb or 0

        # 发货：按发货日期
        if c.delivery_date and year_start <= c.delivery_date <= year_end:
            rd['ship_units'] += c.unit_count or 0
            rd['ship_amount'] += c.contract_amount_rmb or 0
            md['ship_units'] += c.unit_count or 0
            md['ship_amount'] += c.contract_amount_rmb or 0

        # 2021年前排产未发货（排产<2022 且 未发货）
        if c.schedule_date and c.schedule_date < date(2022, 1, 1):
            if not c.delivery_date or c.delivery_date > date.today():
                md['prev_schedule_not_shipped'] += c.unit_count or 0

        # 业务员维度
        if sp:
            if sp not in sp_data:
                sp_data[sp] = {
                    'salesperson': sp, 'region': region, 'module': module,
                    **_init_metrics(),
                    'prev_sign_units': 0, 'prev_sign_amount': 0,
                    'prev_ship_units': 0, 'prev_ship_amount': 0,
                    'prev_schedule_units': 0, 'prev_schedule_amount': 0,
                    'prev_payment_amount': 0,
                }
            sd = sp_data[sp]
            if c.sign_date and year_start <= c.sign_date <= year_end:
                sd['sign_units'] += c.unit_count or 0
                sd['sign_amount'] += c.contract_amount_rmb or 0
            if c.schedule_date and year_start <= c.schedule_date <= year_end:
                sd['schedule_units'] += c.unit_count or 0
                sd['schedule_amount'] += c.contract_amount_rmb or 0
            if c.delivery_date and year_start <= c.delivery_date <= year_end:
                sd['ship_units'] += c.unit_count or 0
                sd['ship_amount'] += c.contract_amount_rmb or 0
            # 去年同期：上年数据
            prev_year_start = date(year - 1, 1, 1)
            prev_year_end = date(year - 1, 12, 31)
            if c.sign_date and prev_year_start <= c.sign_date <= prev_year_end:
                sd['prev_sign_units'] += c.unit_count or 0
                sd['prev_sign_amount'] += c.contract_amount_rmb or 0
            if c.schedule_date and prev_year_start <= c.schedule_date <= prev_year_end:
                sd['prev_schedule_units'] += c.unit_count or 0
                sd['prev_schedule_amount'] += c.contract_amount_rmb or 0
            if c.delivery_date and prev_year_start <= c.delivery_date <= prev_year_end:
                sd['prev_ship_units'] += c.unit_count or 0
                sd['prev_ship_amount'] += c.contract_amount_rmb or 0

    # 回款聚合：回款通过合同号匹配台账 → 映射 → 大区/模块
    payments = PaymentCollection.query.filter(
        PaymentCollection.payment_date >= year_start,
        PaymentCollection.payment_date <= year_end,
        ~PaymentCollection.node_type.in_(['抵佣金', '手续费']),
        PaymentCollection.agent_name != '新加坡分公司(关联方)',
        ~PaymentCollection.contract_no.like('DHN%'),
        ~PaymentCollection.contract_no.like('YBN%'),
    ).all()

    # 建立合同号→(region,module,sp)的快速查找
    contract_lookup = {}
    for c in contracts:
        if c.contract_no and c.contract_no not in contract_lookup:
            cm = mapping.get(c.country) if c.country else None
            if cm:
                contract_lookup[c.contract_no] = (cm['region'], cm['module'], cm['salesperson'])
            elif c.mapped_region:
                contract_lookup[c.contract_no] = (c.mapped_region, c.mapped_module or '', '')

    for p in payments:
        info = contract_lookup.get(p.contract_no)
        if not info:
            continue
        pregion, pmodule, psp = info
        amt = p.payment_amount_rmb or 0

        if pregion not in region_data:
            region_data[pregion] = _init_metrics()
        region_data[pregion]['payment_amount'] += amt

        mod_key = (pregion, pmodule)
        if mod_key in module_data:
            module_data[mod_key]['payment_amount'] += amt

        if psp and psp in sp_data:
            sp_data[psp]['payment_amount'] += amt
            # 去年同期回款
            if p.payment_date and date(year-1,1,1) <= p.payment_date <= date(year-1,12,31):
                sp_data[psp]['prev_payment_amount'] += amt

    return region_data, module_data, sp_data


# ── 输出构建函数 ───────────────────────────────────────────

def _build_region_summary(region_data, targets, year):
    """构建大区汇总数据"""
    result = []
    grand = {m['id']: {'target': 0.0, 'actual': 0.0} for m in METRIC_CONFIG}

    for region in REGION_ORDER:
        rd = region_data.get(region, {})
        entry = {'region': region, 'metrics': {}}
        for m in METRIC_CONFIG:
            mk = m['id']
            t_val = targets.get((region, '', mk), 0)
            a_val = rd.get(mk, 0)
            if m['type'] == 'amount':
                t_val = _to_wan(t_val)
                a_val = _to_wan(a_val)
            else:
                t_val = round(t_val, 0)
                a_val = round(a_val, 0)
            ratio = _safe_ratio(a_val, t_val)
            entry['metrics'][mk] = {'target': t_val, 'actual': a_val, 'ratio': ratio}
            grand[mk]['target'] += t_val
            grand[mk]['actual'] += a_val
        result.append(entry)

    grand_row = {'region': '合计', 'metrics': {}, 'is_total': True}
    for m in METRIC_CONFIG:
        mk = m['id']
        g = grand[mk]
        grand_row['metrics'][mk] = {
            'target': round(g['target'], 2),
            'actual': round(g['actual'], 2),
            'ratio': _safe_ratio(g['actual'], g['target']),
        }

    return result, grand_row


def _build_module_detail(module_data, targets, region_filter):
    """构建模块明细数据（按大区分组，含小计行）"""
    result = []
    current_region = None
    subtotal = {m['id']: {'target': 0.0, 'actual': 0.0} for m in METRIC_CONFIG}
    prev_shipped_total = 0

    # 按大区顺序 + 模块名排序
    sorted_modules = sorted(module_data.values(), key=lambda x: (
        REGION_ORDER.index(x['region']) if x['region'] in REGION_ORDER else 99,
        x['module']
    ))

    for md in sorted_modules:
        if region_filter and md['region'] != region_filter:
            continue

        if md['region'] != current_region:
            # 输出上一大区小计
            if current_region:
                st_entry = {
                    'type': 'subtotal', 'row_type': 'subtotal',
                    'region': current_region, 'module_name': f'{current_region}合计',
                    'metrics': {}, 'prev_schedule_not_shipped': prev_shipped_total,
                }
                for m in METRIC_CONFIG:
                    mk = m['id']
                    st = subtotal[mk]
                    st_entry['metrics'][mk] = {
                        'target': round(st['target'], 2),
                        'actual': round(st['actual'], 2),
                        'ratio': _safe_ratio(st['actual'], st['target']),
                    }
                result.append(st_entry)

            current_region = md['region']
            subtotal = {m['id']: {'target': 0.0, 'actual': 0.0} for m in METRIC_CONFIG}
            prev_shipped_total = 0

        tk = {(md['region'], md['module'], mk): targets.get((md['region'], md['module'], mk), 0)
              for mk in [m['id'] for m in METRIC_CONFIG]}

        entry = {
            'type': 'module', 'row_type': 'data',
            'region': md['region'], 'module_name': md['module'],
            'module_manager': md.get('manager', ''),
            'prev_schedule_not_shipped': md.get('prev_schedule_not_shipped', 0),
            'metrics': {},
        }
        for m in METRIC_CONFIG:
            mk = m['id']
            t_val = targets.get((md['region'], md['module'], mk), 0)
            a_val = md.get(mk, 0)
            if m['type'] == 'amount':
                t_val = _to_wan(t_val)
                a_val = _to_wan(a_val)
            else:
                t_val = round(t_val, 0)
                a_val = round(a_val, 0)
            ratio = _safe_ratio(a_val, t_val)
            entry['metrics'][mk] = {'target': t_val, 'actual': a_val, 'ratio': ratio}
            subtotal[mk]['target'] += t_val
            subtotal[mk]['actual'] += a_val
        prev_shipped_total += md.get('prev_schedule_not_shipped', 0)
        result.append(entry)

    # 最后一个大区的小计
    if current_region:
        st_entry = {
            'type': 'subtotal', 'row_type': 'subtotal',
            'region': current_region, 'module_name': f'{current_region}合计',
            'metrics': {}, 'prev_schedule_not_shipped': prev_shipped_total,
        }
        for m in METRIC_CONFIG:
            mk = m['id']
            st = subtotal[mk]
            st_entry['metrics'][mk] = {
                'target': round(st['target'], 2),
                'actual': round(st['actual'], 2),
                'ratio': _safe_ratio(st['actual'], st['target']),
            }
        result.append(st_entry)

    return result


def _build_salesperson_data(sp_data, region_filter):
    """构建业务员表数据"""
    result = []
    for name, sd in sorted(sp_data.items()):
        if region_filter and sd.get('region') != region_filter:
            continue

        entry = {
            'salesperson': name,
            'module': sd.get('module', ''),
            'region': sd.get('region', ''),
            'metrics_prev': {},   # 去年同期
            'metrics_curr': {},   # 今年
            'yoy_change': {},     # 同比增减%
        }

        metric_pairs = [
            ('sign_units', 'sign_units', 'count'),
            ('sign_amount', 'sign_amount', 'amount'),
            ('schedule_units', 'schedule_units', 'count'),
            ('schedule_amount', 'schedule_amount', 'amount'),
            ('ship_units', 'ship_units', 'count'),
            ('ship_amount', 'ship_amount', 'amount'),
            ('payment_amount', 'prev_payment_amount', 'amount'),
        ]

        for mk, prev_key, mtype in metric_pairs:
            v_prev = sd.get(f'prev_{mk}', 0)
            v_curr = sd.get(mk, 0)
            if mtype == 'amount':
                v_prev = _to_wan(v_prev)
                v_curr = _to_wan(v_curr)
            else:
                v_prev = round(v_prev, 0)
                v_curr = round(v_curr, 0)
            change = round((v_curr - v_prev) / v_prev * 100, 1) if v_prev != 0 else (0 if v_curr == 0 else 100)

            entry['metrics_prev'][mk] = v_prev
            entry['metrics_curr'][mk] = v_curr
            entry['yoy_change'][mk] = change

        result.append(entry)
    return result


def _write_unmatched_file(unmatched_list):
    """输出未匹配合同TXT"""
    uploads_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'uploads'
    )
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, 'unmatched_contracts.txt')

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    lines = [f"未匹配合同清单 - 生成时间: {timestamp}"]
    lines.append(f"{'合同号':<25} {'国家':<20} {'项目名称'}")
    lines.append("-" * 80)

    # 去重
    seen = set()
    for item in unmatched_list:
        key = item['contract_no']
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"{item['contract_no']:<25} {item['country']:<20} {item['project_name']}")

    lines.append("-" * 80)
    lines.append(f"共 {len(seen)} 条未匹配记录")

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return filepath, len(seen)


# ── 公开API ───────────────────────────────────────────────

def get_region_summary(year=2026, region_filter=None):
    """大区汇总 (Sheet 1)"""
    mapping = _load_mapping()
    targets = _load_targets(year, region_only=True)
    unmatched = []
    region_data, _, _ = _compute_aggregations(year, mapping, unmatched)
    regions, grand = _build_region_summary(region_data, targets, year)

    if region_filter:
        regions = [r for r in regions if r['region'] == region_filter]

    return {
        'year': year,
        'regions': regions,
        'grand_total': grand,
        'metric_config': METRIC_CONFIG,
        'region_order': REGION_ORDER,
        'unmatched_count': len(unmatched),
    }


def get_module_detail(year=2026, region_filter=None):
    """模块明细 (Sheet 2)"""
    mapping = _load_mapping()
    targets = _load_targets(year, region_only=False)
    unmatched = []
    _, module_data, _ = _compute_aggregations(year, mapping, unmatched)
    modules = _build_module_detail(module_data, targets, region_filter)

    return {
        'year': year,
        'modules': modules,
        'metric_config': METRIC_CONFIG,
        'region_order': REGION_ORDER,
        'unmatched_count': len(unmatched),
    }


def get_salesperson_comparison(year=2026, region_filter=None):
    """业务员表 (Sheet 3)"""
    mapping = _load_mapping()
    unmatched = []
    _, _, sp_data = _compute_aggregations(year, mapping, unmatched)
    salespersons = _build_salesperson_data(sp_data, region_filter)

    return {
        'year': year,
        'salespersons': salespersons,
        'unmatched_count': len(unmatched),
    }


def get_data_status():
    """数据状态"""
    from models import FileUpload, UploadConfig

    counts = {
        'ledger_count': LedgerContract.query.filter(or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')).count(),
        'mapping_count': CountryMapping.query.count(),
        'payment_count': PaymentCollection.query.count(),
    }

    def _last_upload(code):
        cfg = UploadConfig.query.filter_by(code=code).first()
        if not cfg:
            return None
        fu = FileUpload.query.filter_by(upload_config_id=cfg.id)\
            .order_by(FileUpload.uploaded_at.desc()).first()
        return fu.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if fu else None

    # 触发未匹配检测（排除已作废）
    mapping = _load_mapping()
    contracts = LedgerContract.query.filter(
        LedgerContract.source == 'ledger',
        or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')
    ).all()
    unmatched = []
    seen = set()
    for c in contracts:
        if not c.country:
            unmatched.append({'contract_no': c.contract_no, 'country': '(空)',
                              'project_name': c.project_name or ''})
        elif c.country not in mapping:
            key = c.contract_no
            if key not in seen:
                seen.add(key)
                unmatched.append({'contract_no': c.contract_no, 'country': c.country,
                                  'project_name': c.project_name or ''})

    # 写文件
    filepath, ucount = _write_unmatched_file(unmatched) if unmatched else (None, 0)

    return {
        **counts,
        'last_ledger_upload': _last_upload('contract_ledger'),
        'last_mapping_upload': _last_upload('contract_mapping'),
        'last_payment_upload': _last_upload('contract_payment'),
        'last_report_a_upload': _last_upload('contract_report_a'),
        'last_report_b_upload': _last_upload('contract_report_b'),
        'unmatched_count': ucount,
        'unmatched_file': '/uploads/unmatched_contracts.txt' if filepath else None,
    }


def get_unmatched_contracts():
    """返回未匹配合同列表（排除已作废）"""
    mapping = _load_mapping()
    contracts = LedgerContract.query.filter(
        LedgerContract.source == 'ledger',
        or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')
    ).all()
    unmatched = []
    seen = set()
    for c in contracts:
        if not c.country:
            unmatched.append({'contract_no': c.contract_no, 'country': '(空)',
                              'project_name': c.project_name or ''})
        elif c.country not in mapping:
            key = c.contract_no
            if key not in seen:
                seen.add(key)
                unmatched.append({'contract_no': c.contract_no, 'country': c.country,
                                  'project_name': c.project_name or ''})

    filepath, ucount = _write_unmatched_file(unmatched) if unmatched else (None, 0)
    return {
        'count': ucount,
        'file_url': '/uploads/unmatched_contracts.txt' if filepath else None,
        'items': unmatched[:100],  # 最多返回前100条
    }


def get_regions():
    """可用大区列表"""
    return {'regions': REGION_ORDER}


# ── 两年对比表 (核心新功能) ──────────────────────────────

# 指标组定义（严格按照模板列顺序）
TWO_YEAR_GROUPS = [
    {"id": "sign_units",      "name": "签订台数",       "has_growth": True},
    {"id": "sign_amount",     "name": "签订额\n（万元）","has_growth": True},
    {"id": "overseas_diff",   "name": "海外差额",       "has_growth": False},
    {"id": "sign_total",      "name": "签订额合计",     "has_growth": True},
    {"id": "schedule_units",  "name": "排产台数",       "has_growth": True},
    {"id": "schedule_amount",       "name": "排产额",         "has_growth": True},
    {"id": "schedule_overseas_diff","name": "排产海外差额",   "has_growth": False},
    {"id": "schedule_total",        "name": "排产额合计",     "has_growth": True},
    {"id": "ship_units",            "name": "发货台数",       "has_growth": True},
    {"id": "ship_amount",           "name": "发货额",         "has_growth": True},
    {"id": "ship_overseas_diff",    "name": "发货海外差额",   "has_growth": False},
    {"id": "ship_total",            "name": "发货额合计",     "has_growth": True},
    {"id": "payment",               "name": "回款",           "has_growth": False},
    {"id": "overseas_payment","name": "海外回款及其他", "has_growth": False},
    {"id": "payment_total",   "name": "回款额",         "has_growth": True},
]

# 贸易模块顺序
TRADE_MODULES_ORDER = ['商贸1', '商贸2', '商贸3', '配件-1', '配件-2', '改造']

# 模块→市场类别 映射（匹配映射表实际模块名 → 模板类别）
CATEGORY_MAP = {
    # 俄罗斯（映射表用全角括号）
    '俄罗斯（中部）': 'A', '俄罗斯（西部）': 'A', '俄罗斯（东部）': 'A', '罗斯托夫': 'A', '哈巴': 'A',
    '俄罗斯（中南）': 'B', '俄罗斯（西南）': 'B',
    '莫斯科': 'C', '叶卡': 'C', '新西': 'C',
    # 兼容旧名
    '俄罗斯中央部门': 'A', '俄罗斯新阿尔巴特': 'A', '俄罗斯东部': 'A',
    # 中亚
    '哈萨克斯坦-1': 'A', '塔吉克/哈萨克斯坦-2': 'B', '哈萨克斯坦-3': 'C',
    '乌兹别克斯坦': 'D', '吉尔吉斯斯坦': 'D', '阿塞拜疆': 'C', '格鲁吉亚': 'D', '蒙古': 'C',
    # 亚洲1
    '朝鲜-韩国': 'D', '新加坡市政会': 'A', '新加坡HDB': 'A',
    '泰国': 'D', '泰国2': 'D', '金三角': 'D',
    '马来西亚': 'B', '马尔代夫': 'D', '香港-台湾': 'D',
    # 亚洲2
    '越南-2': 'C', '印度公建': 'B', '巴基斯坦': 'D',
    '孟加拉-1': 'B', '澳大利亚': 'C', '菲律宾': 'C', '菲律宾-2': 'C',
    '印尼-1': 'C', '印尼-2': 'C', '越南-1（工厂）': 'B', '印度私营': 'A',
    # 美洲
    '墨西哥-1': 'A', '墨西哥-2': 'A', '秘鲁': 'C', '智利': 'C',
    '加勒比海': 'A', '多米尼加': 'C', '哥伦比亚': 'A',
    '秘鲁2': 'D', '巴西': 'D',
    # 中东
    '阿联酋-2': 'B', '沙特工厂': 'A', '沙特-1': 'A',
    '科威特': 'C', '阿联酋-1': 'A', '卡塔尔': 'C',
    '伊拉克': 'D', '巴勒斯坦': 'C', '伊朗+阿曼': 'D', '阿联酋3': 'C',
    # 非洲
    '埃及-1': 'A', '埃及-2': 'D', '肯尼亚/坦桑尼亚': 'C',
    '尼日利亚/埃塞俄比亚': 'C', '非洲法语区': 'D', '南非/安格拉': 'D',
    # 欧洲
    '德国西班牙': 'D', '东欧': 'D', '英国意大利': 'D',
    # 商贸
    '商贸1': '', '商贸2': '', '商贸3': '', '配件-1': '', '配件-2': '', '改造': '',
}

# AZT合同 → 模块映射（台账中不存在的合同，回款通过此映射归入对应模块）
AZT_MODULE_MAP = {
    'AZT-230002T': '加勒比海',
    'AZT-240002T': '俄罗斯（东部）',
    'AZT-240003T': '科威特',
    'AZT-240008T': '俄罗斯（东部）',
    'AZT-250009T': '澳大利亚',
    'AZT-250010T': '俄罗斯（西南）',
    'AZT-250013T': '俄罗斯（东部）',
    'AZT-250014T': '俄罗斯（东部）',
    'AZT-250015T': '俄罗斯（西南）',
}


def _get_category(module_name):
    """获取模块的市场类别，优先精确匹配，其次模板映射"""
    return CATEGORY_MAP.get(module_name, '')


def get_two_year_comparison():
    """
    两年对比表 — 严格按照模板 Sheet 3 格式
    日期动态计算：当年1月1日~今日，去年1月1日~去年同日
    行结构：以国家映射表为准，按大区→模块分组，含小计行、商贸行、国际总计

    缓存策略：基于数据指纹（表行数+日期），数据未变时直接返回缓存。
    """
    global _two_year_cache, _two_year_fingerprint

    # 检查缓存
    fp = _get_data_fingerprint()
    if _two_year_fingerprint == fp and _two_year_cache is not None:
        return _two_year_cache

    # 数据截止日期 = 昨天（当天数据通常尚未入库）
    cutoff = date.today() - timedelta(days=1)
    year_curr = cutoff.year
    year_prev = cutoff.year - 1

    # 今年范围：1月1日 ~ 昨天
    curr_start = date(year_curr, 1, 1)
    curr_end = cutoff
    # 去年范围：1月1日 ~ 去年同日
    prev_start = date(year_prev, 1, 1)
    prev_end = date(year_prev, cutoff.month, cutoff.day)

    # 加载国家映射表
    mapping = _load_mapping()
    contract_map = {}  # contract_no → (region, module)
    for c in LedgerContract.query.filter(
        LedgerContract.source == 'ledger'
    ).all():
        if c.contract_no and c.country:
            # 改造梯回款归入商贸合计/改造（与签约/排产/发货逻辑一致）
            if c.product_type and '改造' in str(c.product_type):
                contract_map[c.contract_no] = ('商贸合计', '改造')
            else:
                cm = mapping.get(c.country)
                if cm:
                    contract_map[c.contract_no] = (cm['region'], cm['module'])

    # 补充 AZT 合同映射（这些合同台账中不存在，通过固定映射归入模块）
    _azt_region_cache = {}
    for azt_cn, azt_module in AZT_MODULE_MAP.items():
        if azt_module not in _azt_region_cache:
            for cm in mapping.values():
                if cm['module'] == azt_module:
                    _azt_region_cache[azt_module] = cm['region']
                    break
            else:
                _azt_region_cache[azt_module] = None
        region = _azt_region_cache[azt_module]
        if region:
            contract_map[azt_cn] = (region, azt_module)

    # ── 聚合：按 (region, module) 分组 ──
    agg = {}
    # 单独追踪真正改造梯（product_type含"改造"）的台数，用于国际总计排除
    gaizao_true_units = {
        'sign_units_prev': 0, 'sign_units_curr': 0,
        'schedule_units_prev': 0, 'schedule_units_curr': 0,
        'ship_units_prev': 0, 'ship_units_curr': 0,
    }

    def _new_entry(region, module):
        return {
            'region': region, 'module': module,
            'sign_units_prev': 0, 'sign_units_curr': 0,
            'sign_amount_prev': 0.0, 'sign_amount_curr': 0.0,
            'schedule_units_prev': 0, 'schedule_units_curr': 0,
            'schedule_amount_prev': 0.0, 'schedule_amount_curr': 0.0,
            'schedule_overseas_diff_prev': 0.0, 'schedule_overseas_diff_curr': 0.0,
            'ship_units_prev': 0, 'ship_units_curr': 0,
            'ship_amount_prev': 0.0, 'ship_amount_curr': 0.0,
            'ship_overseas_diff_prev': 0.0, 'ship_overseas_diff_curr': 0.0,
            'payment_prev': 0.0, 'payment_curr': 0.0,
        }

    # 预填充：映射表中的所有模块（即使无数据也展示）
    seen_modules = set()
    for cm in mapping.values():
        r, m = cm['region'], cm['module']
        if not r or not m:
            continue
        key = (r, m)
        if key not in seen_modules:
            seen_modules.add(key)
            agg[key] = _new_entry(r, m)
    # 商贸模块也预填充
    for tm in TRADE_MODULES_ORDER:
        key = ('商贸合计', tm)
        if key not in seen_modules:
            seen_modules.add(key)
            agg[key] = _new_entry('商贸合计', tm)
    # 年度指标中有但映射表中没有的模块也预填充（如德国西班牙），确保即使无合同数据也展示
    for (region, module), targets in ANNUAL_TARGETS.items():
        if region in ('商贸配件',):
            continue  # 商贸配件的指标映射到商贸合计
        key = (region, module)
        if key not in seen_modules:
            seen_modules.add(key)
            agg[key] = _new_entry(region, module)

    def _ensure(region, module):
        key = (region, module)
        if key not in agg:
            agg[key] = _new_entry(region, module)
        return agg[key]

    # 遍历台账（排除已作废）
    for c in LedgerContract.query.filter(
        LedgerContract.source.in_(['ledger', 'report_a', 'report_b']),
        or_(LedgerContract.product_status == None, LedgerContract.product_status != '已作废')
    ).all():
        # 确定 region / module
        if c.source in ('report_a', 'report_b'):
            region = c.mapped_region or '商贸合计'
            module = c.mapped_module or ''
        elif c.country and c.country in mapping:
            cm = mapping[c.country]
            region = cm['region']
            module = cm['module']
            # 改造强制归入商贸合计
            if c.product_type and '改造' in str(c.product_type):
                region = '商贸合计'
                module = '改造'
        else:
            continue  # 未匹配的跳过

        if not region or not module:
            continue

        d = _ensure(region, module)
        # 判断是否为真正改造梯（product_type含"改造"），其台数需从国际总计中排除
        is_true_gaizao = bool(c.product_type and '改造' in str(c.product_type))

        # 签订
        if c.sign_date:
            if prev_start <= c.sign_date <= prev_end:
                d['sign_units_prev'] += c.unit_count or 0
                d['sign_amount_prev'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['sign_units_prev'] += c.unit_count or 0
            if curr_start <= c.sign_date <= curr_end:
                d['sign_units_curr'] += c.unit_count or 0
                d['sign_amount_curr'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['sign_units_curr'] += c.unit_count or 0

        # 排产
        if c.schedule_date:
            if prev_start <= c.schedule_date <= prev_end:
                d['schedule_units_prev'] += c.unit_count or 0
                d['schedule_amount_prev'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['schedule_units_prev'] += c.unit_count or 0
            if curr_start <= c.schedule_date <= curr_end:
                d['schedule_units_curr'] += c.unit_count or 0
                d['schedule_amount_curr'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['schedule_units_curr'] += c.unit_count or 0

        # 发货
        if c.delivery_date:
            if prev_start <= c.delivery_date <= prev_end:
                d['ship_units_prev'] += c.unit_count or 0
                d['ship_amount_prev'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['ship_units_prev'] += c.unit_count or 0
            if curr_start <= c.delivery_date <= curr_end:
                d['ship_units_curr'] += c.unit_count or 0
                d['ship_amount_curr'] += c.contract_amount_rmb or 0
                if is_true_gaizao:
                    gaizao_true_units['ship_units_curr'] += c.unit_count or 0

    # 回款聚合（排除：抵佣金/手续费、新加坡分公司关联方、DHN/YBN合同）
    payments = PaymentCollection.query.filter(
        ~PaymentCollection.node_type.in_(['抵佣金', '手续费']),
        PaymentCollection.agent_name != '新加坡分公司(关联方)',
        ~PaymentCollection.contract_no.like('DHN%'),
        ~PaymentCollection.contract_no.like('YBN%'),
    ).all()
    for p in payments:
        info = contract_map.get(p.contract_no)
        if not info:
            continue
        pregion, pmodule = info
        amt = p.payment_amount_rmb or 0
        d = _ensure(pregion, pmodule)
        if p.payment_date:
            if prev_start <= p.payment_date <= prev_end:
                d['payment_prev'] += amt
            if curr_start <= p.payment_date <= curr_end:
                d['payment_curr'] += amt

    # ── 构建行列表 ──
    rows = []

    # 金额转万元（取整）
    def _wan(v):
        return round(v / 10000)

    def _growth(prev, curr):
        if prev == 0:
            return None  # 显示 '-'
        return round((curr - prev) / prev * 100)  # 整数百分比

    def _make_row(typ, region, module, d):
        cat = _get_category(module) if typ == 'data' else ''

        # 海外差值：万元单位，直接从 d 取（None → 显示为 '-'）
        od_sign_c = d.get('overseas_diff_curr') or 0
        od_schedule_c = d.get('schedule_overseas_diff_curr') or 0
        od_ship_c = d.get('ship_overseas_diff_curr') or 0
        od_payment_c = d.get('overseas_payment_curr') or 0
        # 2025年海外差值均为 None
        od_sign_p = d.get('overseas_diff_prev')
        od_schedule_p = d.get('schedule_overseas_diff_prev')
        od_ship_p = d.get('ship_overseas_diff_prev')
        od_payment_p = d.get('overseas_payment_prev')

        # 合计 = 国内额（元→万元）+ 海外差额（万元）
        st_p = _wan(d['sign_amount_prev']) + (od_sign_p or 0)
        st_c = _wan(d['sign_amount_curr']) + od_sign_c
        sc_p = _wan(d['schedule_amount_prev']) + (od_schedule_p or 0)
        sc_c = _wan(d['schedule_amount_curr']) + od_schedule_c
        sh_p = _wan(d['ship_amount_prev']) + (od_ship_p or 0)
        sh_c = _wan(d['ship_amount_curr']) + od_ship_c
        pt_p = _wan(d['payment_prev']) + (od_payment_p or 0)
        pt_c = _wan(d['payment_curr']) + od_payment_c

        # 合计增长率基于合计值计算
        def _total_growth(tp, tc):
            if tp == 0:
                return None
            return round((tc - tp) / tp * 100)

        return {
            'type': typ,
            'region': region,
            'module': module,
            'category': cat,
            'sign_units_prev': d['sign_units_prev'],
            'sign_units_curr': d['sign_units_curr'],
            'sign_units_growth': _growth(d['sign_units_prev'], d['sign_units_curr']),
            'sign_amount_prev': _wan(d['sign_amount_prev']),
            'sign_amount_curr': _wan(d['sign_amount_curr']),
            'sign_amount_growth': _growth(d['sign_amount_prev'], d['sign_amount_curr']),
            'overseas_diff_prev': od_sign_p, 'overseas_diff_curr': d.get('overseas_diff_curr'),
            'sign_total_prev': st_p,
            'sign_total_curr': st_c,
            'sign_total_growth': _total_growth(st_p, st_c),
            'schedule_units_prev': d['schedule_units_prev'],
            'schedule_units_curr': d['schedule_units_curr'],
            'schedule_units_growth': _growth(d['schedule_units_prev'], d['schedule_units_curr']),
            'schedule_amount_prev': _wan(d['schedule_amount_prev']),
            'schedule_amount_curr': _wan(d['schedule_amount_curr']),
            'schedule_amount_growth': _growth(d['schedule_amount_prev'], d['schedule_amount_curr']),
            'schedule_overseas_diff_prev': od_schedule_p, 'schedule_overseas_diff_curr': d.get('schedule_overseas_diff_curr'),
            'schedule_total_prev': sc_p,
            'schedule_total_curr': sc_c,
            'schedule_total_growth': _total_growth(sc_p, sc_c),
            'ship_units_prev': d['ship_units_prev'],
            'ship_units_curr': d['ship_units_curr'],
            'ship_units_growth': _growth(d['ship_units_prev'], d['ship_units_curr']),
            'ship_amount_prev': _wan(d['ship_amount_prev']),
            'ship_amount_curr': _wan(d['ship_amount_curr']),
            'ship_amount_growth': _growth(d['ship_amount_prev'], d['ship_amount_curr']),
            'ship_overseas_diff_prev': od_ship_p, 'ship_overseas_diff_curr': d.get('ship_overseas_diff_curr'),
            'ship_total_prev': sh_p,
            'ship_total_curr': sh_c,
            'ship_total_growth': _total_growth(sh_p, sh_c),
            'payment_prev': _wan(d['payment_prev']),
            'payment_curr': _wan(d['payment_curr']),
            'overseas_payment_prev': od_payment_p, 'overseas_payment_curr': d.get('overseas_payment_curr'),
            'payment_total_prev': pt_p,
            'payment_total_curr': pt_c,
            'payment_total_growth': _total_growth(pt_p, pt_c),
        }

    # ── 加载海外差值数据（两年）并注入聚合 ──
    from dashboards.contract_completion.models import OverseasDiff

    def _load_overseas_lookup(year):
        """加载指定年份的海外差值 {module_name: {sign_diff, schedule_diff, ship_diff, payment_diff}}"""
        lookup = {}
        for od in OverseasDiff.query.filter_by(data_year=year).all():
            lookup[od.module_name] = {
                'sign_diff': od.sign_diff,
                'schedule_diff': od.schedule_diff,
                'ship_diff': od.ship_diff,
                'payment_diff': od.payment_diff,
            }
        if '印度' in lookup:
            lookup['印度公建'] = lookup.pop('印度')
        return lookup

    overseas_prev = _load_overseas_lookup(year_prev)
    overseas_curr = _load_overseas_lookup(year_curr)

    def _round_wan(v):
        """万元值取整（海外差值已是万元，直接 round）"""
        if v is None:
            return None
        return round(v)

    # 注入到聚合条目中（万元单位，不经过 _wan 转换；取整存储）
    for key, d in agg.items():
        region, module = key
        od_p = overseas_prev.get(module, {})
        od_c = overseas_curr.get(module, {})
        d['overseas_diff_prev'] = _round_wan(od_p.get('sign_diff'))
        d['overseas_diff_curr'] = _round_wan(od_c.get('sign_diff'))
        d['schedule_overseas_diff_prev'] = _round_wan(od_p.get('schedule_diff'))
        d['schedule_overseas_diff_curr'] = _round_wan(od_c.get('schedule_diff'))
        d['ship_overseas_diff_prev'] = _round_wan(od_p.get('ship_diff'))
        d['ship_overseas_diff_curr'] = _round_wan(od_c.get('ship_diff'))
        d['overseas_payment_prev'] = _round_wan(od_p.get('payment_diff'))
        d['overseas_payment_curr'] = _round_wan(od_c.get('payment_diff'))

    # 大区排序
    region_order = ['俄罗斯', '中亚', '亚洲1', '亚洲2', '美洲', '中东', '非洲', '欧洲']

    # 收集非商贸模块并按大区排序
    normal_entries = [(k, v) for k, v in agg.items() if v['region'] not in ('商贸合计', '商贸配件')]
    normal_entries.sort(key=lambda x: (
        region_order.index(x[1]['region']) if x[1]['region'] in region_order else 99,
        x[1]['module']
    ))

    # 按大区分组构建行
    current_region = None
    subtotals = {}

    for key, d in normal_entries:
        region, module = key

        if region != current_region:
            # 输出上一大区小计
            if current_region and current_region in subtotals:
                rows.append(_make_row('subtotal', current_region,
                                      f'{current_region}合计', subtotals[current_region]))
            current_region = region
            subtotals[region] = {k: 0 for k in d}

        # 累加到小计
        for k in d:
            if isinstance(d[k], (int, float)):
                subtotals[region][k] += d[k]

        rows.append(_make_row('data', region, module, d))

    # 最后一个大区小计
    if current_region and current_region in subtotals:
        rows.append(_make_row('subtotal', current_region,
                              f'{current_region}合计', subtotals[current_region]))

    # ── 注入商贸模块汇总数据（从商贸数据.xlsx上传，两年） ──
    from dashboards.contract_completion.models import TradeModuleData

    def _inject_trade_data(year, suffix):
        """注入指定年份的商贸模块数据到聚合表
        仅覆盖非零字段，避免配件-1/2的payment-only记录覆盖合同级签单/排产/发货数据
        """
        for td in TradeModuleData.query.filter_by(data_year=year).all():
            tmod = td.module_name
            if tmod in TRADE_MODULES_ORDER:
                d = _ensure('商贸合计', tmod)
                if td.sign_amount:
                    d[f'sign_amount_{suffix}'] = td.sign_amount
                if td.schedule_amount:
                    d[f'schedule_amount_{suffix}'] = td.schedule_amount
                if td.ship_amount:
                    d[f'ship_amount_{suffix}'] = td.ship_amount
                if td.payment_amount:
                    d[f'payment_{suffix}'] = td.payment_amount

    _inject_trade_data(year_prev, 'prev')
    _inject_trade_data(year_curr, 'curr')

    # ── 商贸行 ──
    trade_total = {}
    for tmod in TRADE_MODULES_ORDER:
        d = agg.get(('商贸合计', tmod))
        if d:
            if not trade_total:
                trade_total = {k: 0 for k in d}
            for k in d:
                if isinstance(d[k], (int, float)):
                    trade_total[k] += d[k]
            row = _make_row('trade', '商贸合计', tmod, d)
            # 商贸/配件行不显示金额增长率（2025年数据不完整，增长比例无意义）
            for gf in ['sign_amount_growth', 'sign_total_growth',
                       'schedule_amount_growth', 'schedule_total_growth',
                       'ship_amount_growth', 'ship_total_growth',
                       'payment_total_growth']:
                row[gf] = None
            rows.append(row)

    # 商贸配件合计
    if trade_total:
        rows.append(_make_row('subtotal', '商贸合计', '商贸配件合计', trade_total))

    # ── 国际总计（聚合原始数据，避免万元重复转换） ──
    grand_raw = {}
    for key, d in agg.items():
        if not grand_raw:
            grand_raw = {k: 0 for k in d}
        for k in d:
            if isinstance(d[k], (int, float)):
                grand_raw[k] = grand_raw.get(k, 0) + d[k]
    # 国际总计的台数不包含真正改造梯（product_type含"改造"的合同台数）
    for field in ['sign_units_prev', 'sign_units_curr',
                  'schedule_units_prev', 'schedule_units_curr',
                  'ship_units_prev', 'ship_units_curr']:
        grand_raw[field] = grand_raw.get(field, 0) - gaizao_true_units.get(field, 0)
    if grand_raw:
        rows.append(_make_row('grand_total', '', '国际总计', grand_raw))

    result = {
        'title': '国贸签订、排产、发货两年对比',
        'date_prev_end': prev_end.strftime('%Y-%m-%d'),
        'date_curr_end': curr_end.strftime('%Y-%m-%d'),
        'year_prev': year_prev,
        'year_curr': year_curr,
        'metric_groups': TWO_YEAR_GROUPS,
        'region_order': region_order,
        'rows': rows,
    }

    # 存入缓存
    _two_year_cache = result
    _two_year_fingerprint = fp

    return result



# ── 年度指标 ──────────────────────────────────────────

# 年度指标 — 从 2026年合同完成情况表 抄录
# key: (region, module)
ANNUAL_TARGETS = {
    ('俄罗斯', '俄罗斯（中部）'): {
        "sign_units": 1683, "sign_amount": 35680.0,
        "payment": 13775.0,
        "ship_units": 750, "ship_amount": 14500.0,
        "schedule_units": 750, "schedule_amount": 14500.0,
        "person": '丛峻', "backlog_units": 102,
    },
    ('俄罗斯', '俄罗斯（西部）'): {
        "sign_units": 1683, "sign_amount": 35680.0,
        "payment": 13775.0,
        "ship_units": 750, "ship_amount": 14500.0,
        "schedule_units": 750, "schedule_amount": 14500.0,
        "person": '单一', "backlog_units": 107,
    },
    ('俄罗斯', '俄罗斯（东部）'): {
        "sign_units": 400, "sign_amount": 6544.0,
        "payment": 2524.15,
        "ship_units": 160, "ship_amount": 2657.0,
        "schedule_units": 160, "schedule_amount": 2657.0,
        "person": '宋艾亭', "backlog_units": 30,
    },
    ('俄罗斯', '罗斯托夫'): {
        "sign_units": 270, "sign_amount": 4428.0,
        "payment": 1710.0,
        "ship_units": 120, "ship_amount": 1800.0,
        "schedule_units": 120, "schedule_amount": 1800.0,
        "person": '张执玮', "backlog_units": 61,
    },
    ('俄罗斯', '哈巴'): {
        "sign_units": 270, "sign_amount": 4428.0,
        "payment": 1710.0,
        "ship_units": 120, "ship_amount": 1800.0,
        "schedule_units": 120, "schedule_amount": 1800.0,
        "person": '张小帆', "backlog_units": 28,
    },
    ('俄罗斯', '俄罗斯（中南）'): {
        "sign_units": 245, "sign_amount": 4200.0,
        "payment": 1689.1,
        "ship_units": 100, "ship_amount": 1778.0,
        "schedule_units": 100, "schedule_amount": 1778.0,
        "person": '张沁媛', "backlog_units": 2,
    },
    ('俄罗斯', '俄罗斯（西南）'): {
        "sign_units": 489, "sign_amount": 7525.12,
        "payment": 2823.4,
        "ship_units": 210, "ship_amount": 2972.0,
        "schedule_units": 210, "schedule_amount": 2972.0,
        "person": '刘景伟', "backlog_units": 27,
    },
    ('俄罗斯', '莫斯科'): {
        "sign_units": 72, "sign_amount": 1085.0,
        "payment": 427.5,
        "ship_units": 30, "ship_amount": 450.0,
        "schedule_units": 30, "schedule_amount": 450.0,
        "person": '张东辉', "backlog_units": 0,
    },
    ('俄罗斯', '叶卡'): {
        "sign_units": 72, "sign_amount": 1085.0,
        "payment": 427.5,
        "ship_units": 30, "ship_amount": 450.0,
        "schedule_units": 30, "schedule_amount": 450.0,
        "person": '孙继伟', "backlog_units": 0,
    },
    ('俄罗斯', '新西'): {
        "sign_units": 72, "sign_amount": 1085.0,
        "payment": 427.5,
        "ship_units": 30, "ship_amount": 450.0,
        "schedule_units": 30, "schedule_amount": 450.0,
        "person": '（空）', "backlog_units": 0,
    },
    ('中亚', '蒙古'): {
        "sign_units": 81, "sign_amount": 1260.0,
        "payment": 475.0,
        "ship_units": 32, "ship_amount": 500.0,
        "schedule_units": 32, "schedule_amount": 500.0,
        "person": '逄顺福', "backlog_units": 7,
    },
    ('中亚', '哈萨克斯坦-1'): {
        "sign_units": 581, "sign_amount": 5575.68,
        "payment": 2188.8,
        "ship_units": 239, "ship_amount": 2304.0,
        "schedule_units": 239, "schedule_amount": 2304.0,
        "person": '彭凤琴', "backlog_units": 65,
    },
    ('中亚', '塔吉克/哈萨克斯坦-2'): {
        "sign_units": 207, "sign_amount": 2082.5,
        "payment": 807.5,
        "ship_units": 85, "ship_amount": 850.0,
        "schedule_units": 85, "schedule_amount": 850.0,
        "person": '于洋', "backlog_units": 31,
    },
    ('中亚', '哈萨克斯坦-3'): {
        "sign_units": 145, "sign_amount": 1500.0,
        "payment": 570.0,
        "ship_units": 60, "ship_amount": 600.0,
        "schedule_units": 60, "schedule_amount": 600.0,
        "person": '秦力超', "backlog_units": 9,
    },
    ('中亚', '乌兹别克斯坦'): {
        "sign_units": 81, "sign_amount": 735.0,
        "payment": 285.0,
        "ship_units": 33, "ship_amount": 300.0,
        "schedule_units": 33, "schedule_amount": 300.0,
        "person": '孙博伦', "backlog_units": 15,
    },
    ('中亚', '吉尔吉斯斯坦'): {
        "sign_units": 127, "sign_amount": 980.0,
        "payment": 380.0,
        "ship_units": 52, "ship_amount": 400.0,
        "schedule_units": 52, "schedule_amount": 400.0,
        "person": '马月', "backlog_units": 0,
    },
    ('中亚', '阿塞拜疆'): {
        "sign_units": 73, "sign_amount": 816.66666655,
        "payment": 332.5,
        "ship_units": 32, "ship_amount": 350.0,
        "schedule_units": 32, "schedule_amount": 350.0,
        "person": '韩宇', "backlog_units": 52,
    },
    ('中亚', '格鲁吉亚'): {
        "sign_units": 44, "sign_amount": 490.0,
        "payment": 190.0,
        "ship_units": 20, "ship_amount": 200.0,
        "schedule_units": 20, "schedule_amount": 200.0,
        "person": '孙博伦', "backlog_units": 0,
    },
    ('亚洲1', '朝鲜-韩国'): {
        "sign_units": 54, "sign_amount": 540.0,
        "payment": 250.955281626998,
        "ship_units": 32, "ship_amount": 264.163454344208,
        "schedule_units": 32, "schedule_amount": 264.163454344208,
        "person": '安鹏霖', "backlog_units": 10,
    },
    ('亚洲1', '新加坡市政会'): {
        "sign_units": 277, "sign_amount": 6100.0,
        "payment": 2045.35,
        "ship_units": 101, "ship_amount": 2153.0,
        "schedule_units": 101, "schedule_amount": 2153.0,
        "person": '祁阳', "backlog_units": 2,
    },
    ('亚洲1', '新加坡HDB'): {
        "sign_units": 566, "sign_amount": 20400.0,
        "payment": 8075.0,
        "ship_units": 260, "ship_amount": 8500.0,
        "schedule_units": 260, "schedule_amount": 8500.0,
        "person": '王珊珊', "backlog_units": 63,
    },
    ('亚洲1', '泰国'): {
        "sign_units": 90, "sign_amount": 840.0,
        "payment": 190.0,
        "ship_units": 25, "ship_amount": 200.0,
        "schedule_units": 25, "schedule_amount": 200.0,
        "person": '张琬怡', "backlog_units": 5,
    },
    ('亚洲1', '泰国2'): {
        "sign_units": 136, "sign_amount": 1200.0,
        "payment": 285.0,
        "ship_units": 36, "ship_amount": 300.0,
        "schedule_units": 36, "schedule_amount": 300.0,
        "person": '刘瑞', "backlog_units": 11,
    },
    ('亚洲1', '金三角'): {
        "sign_units": 40, "sign_amount": 200.0,
        "payment": 190.0,
        "ship_units": 36, "ship_amount": 200.0,
        "schedule_units": 36, "schedule_amount": 200.0,
        "person": '张琬怡', "backlog_units": 0,
    },
    ('亚洲1', '马来西亚'): {
        "sign_units": 109, "sign_amount": 900.0,
        "payment": 494.95,
        "ship_units": 61, "ship_amount": 521.0,
        "schedule_units": 61, "schedule_amount": 521.0,
        "person": '鲁鸿飞', "backlog_units": 9,
    },
    ('亚洲1', '马尔代夫'): {
        "sign_units": 36, "sign_amount": 630.0,
        "payment": 250.955281626998,
        "ship_units": 17, "ship_amount": 264.163454344208,
        "schedule_units": 17, "schedule_amount": 264.163454344208,
        "person": '叶哲铭', "backlog_units": 4,
    },
    ('亚洲1', '香港-台湾'): {
        "sign_units": 20, "sign_amount": 200.0,
        "payment": 190.0,
        "ship_units": 15, "ship_amount": 200.0,
        "schedule_units": 15, "schedule_amount": 200.0,
        "person": '王海娇', "backlog_units": 0,
    },
    ('亚洲2', '越南-2'): {
        "sign_units": 235, "sign_amount": 1673.35,
        "payment": 648.85,
        "ship_units": 63, "ship_amount": 683.0,
        "schedule_units": 63, "schedule_amount": 683.0,
        "person": '韩月华', "backlog_units": 1,
    },
    ('亚洲2', '印度公建'): {
        "sign_units": 66, "sign_amount": 930.0,
        "payment": 414.432922440495,
        "ship_units": 28, "ship_amount": 436.245181516311,
        "schedule_units": 28, "schedule_amount": 436.245181516311,
        "person": '张鹏', "backlog_units": 20,
    },
    ('亚洲2', '巴基斯坦'): {
        "sign_units": 87, "sign_amount": 688.0,
        "payment": 251.75,
        "ship_units": 32, "ship_amount": 265.0,
        "schedule_units": 32, "schedule_amount": 265.0,
        "person": '耿建伟', "backlog_units": 0,
    },
    ('亚洲2', '孟加拉-1'): {
        "sign_units": 238, "sign_amount": 1768.9,
        "payment": 685.9,
        "ship_units": 73, "ship_amount": 722.0,
        "schedule_units": 73, "schedule_amount": 722.0,
        "person": '刘晓虎', "backlog_units": 3,
    },
    ('亚洲2', '澳大利亚'): {
        "sign_units": 136, "sign_amount": 964.0,
        "payment": 361.0,
        "ship_units": 32, "ship_amount": 380.0,
        "schedule_units": 32, "schedule_amount": 380.0,
        "person": '董永生', "backlog_units": 8,
    },
    ('亚洲2', '菲律宾'): {
        "sign_units": 69, "sign_amount": 710.0,
        "payment": 332.5,
        "ship_units": 27, "ship_amount": 350.0,
        "schedule_units": 27, "schedule_amount": 350.0,
        "person": '吴旭东', "backlog_units": 0,
    },
    ('亚洲2', '印尼-1'): {
        "sign_units": 263, "sign_amount": 1640.0,
        "payment": 380.0,
        "ship_units": 42, "ship_amount": 400.0,
        "schedule_units": 42, "schedule_amount": 400.0,
        "person": '张欢', "backlog_units": 1,
    },
    ('亚洲2', '越南-1（工厂）'): {
        "sign_units": 148, "sign_amount": 1470.0,
        "payment": 570.0,
        "ship_units": 55, "ship_amount": 600.0,
        "schedule_units": 55, "schedule_amount": 600.0,
        "person": '于春光', "backlog_units": 106,
    },
    ('亚洲2', '印度私营'): {
        "sign_units": 246, "sign_amount": 2934.0,
        "payment": 1577.0,
        "ship_units": 82, "ship_amount": 1660.0,
        "schedule_units": 82, "schedule_amount": 1660.0,
        "person": '任嘉庆', "backlog_units": 48,
    },
    ('亚洲2', '菲律宾-2'): {
        "sign_units": 85, "sign_amount": 650.0,
        "payment": 332.5,
        "ship_units": 27, "ship_amount": 350.0,
        "schedule_units": 27, "schedule_amount": 350.0,
        "person": '吴永顺', "backlog_units": 0,
    },
    ('亚洲2', '印尼-2'): {
        "sign_units": 250, "sign_amount": 1700.0,
        "payment": 380.0,
        "ship_units": 42, "ship_amount": 400.0,
        "schedule_units": 42, "schedule_amount": 400.0,
        "person": '岳亮', "backlog_units": 0,
    },
    ('美洲', '墨西哥-1'): {
        "sign_units": 450, "sign_amount": 6270.0,
        "payment": 2897.5,
        "ship_units": 210, "ship_amount": 3050.0,
        "schedule_units": 210, "schedule_amount": 3050.0,
        "person": '尤健宇', "backlog_units": 36,
    },
    ('美洲', '墨西哥-2'): {
        "sign_units": 450, "sign_amount": 6270.0,
        "payment": 2897.5,
        "ship_units": 210, "ship_amount": 3050.0,
        "schedule_units": 210, "schedule_amount": 3050.0,
        "person": '肖宇晗', "backlog_units": 35,
    },
    ('美洲', '秘鲁'): {
        "sign_units": 100, "sign_amount": 1390.0,
        "payment": 389.5,
        "ship_units": 35, "ship_amount": 410.0,
        "schedule_units": 35, "schedule_amount": 410.0,
        "person": '李知源', "backlog_units": 21,
    },
    ('美洲', '智利'): {
        "sign_units": 120, "sign_amount": 1670.0,
        "payment": 456.0,
        "ship_units": 40, "ship_amount": 480.0,
        "schedule_units": 40, "schedule_amount": 480.0,
        "person": '翟文龙', "backlog_units": 7,
    },
    ('美洲', '加勒比海'): {
        "sign_units": 520, "sign_amount": 7248.8,
        "payment": 2502.3,
        "ship_units": 200, "ship_amount": 2634.0,
        "schedule_units": 200, "schedule_amount": 2634.0,
        "person": '范皓铭', "backlog_units": 61,
    },
    ('美洲', '多米尼加'): {
        "sign_units": 72, "sign_amount": 1000.0,
        "payment": 380.0,
        "ship_units": 40, "ship_amount": 400.0,
        "schedule_units": 40, "schedule_amount": 400.0,
        "person": '梁国裕', "backlog_units": 1,
    },
    ('美洲', '哥伦比亚'): {
        "sign_units": 800, "sign_amount": 11151.0,
        "payment": 4132.5,
        "ship_units": 360, "ship_amount": 4350.0,
        "schedule_units": 360, "schedule_amount": 4350.0,
        "person": '田媛媛', "backlog_units": 76,
    },
    ('美洲', '秘鲁2'): {
        "sign_units": 50, "sign_amount": 700.0,
        "payment": 285.0,
        "ship_units": 30, "ship_amount": 300.0,
        "schedule_units": 30, "schedule_amount": 300.0,
        "person": '魏珍荣', "backlog_units": 0,
    },
    ('美洲', '巴西'): {
        "sign_units": 60, "sign_amount": 840.0,
        "payment": 190.0,
        "ship_units": 20, "ship_amount": 200.0,
        "schedule_units": 20, "schedule_amount": 200.0,
        "person": '杨春', "backlog_units": 0,
    },
    ('中东', '阿联酋-2'): {
        "sign_units": 122, "sign_amount": 3360.0,
        "payment": 1330.0,
        "ship_units": 62, "ship_amount": 1400.0,
        "schedule_units": 62, "schedule_amount": 1400.0,
        "person": '李军辉', "backlog_units": 5,
    },
    ('中东', '沙特工厂'): {
        "sign_units": 733, "sign_amount": 6860.0,
        "payment": 2242.0,
        "ship_units": 200, "ship_amount": 2360.0,
        "schedule_units": 200, "schedule_amount": 2360.0,
        "person": '武永佳', "backlog_units": 0,
    },
    ('中东', '沙特-1'): {
        "sign_units": 299, "sign_amount": 3710.0,
        "payment": 1425.0,
        "ship_units": 150, "ship_amount": 1500.0,
        "schedule_units": 150, "schedule_amount": 1500.0,
        "person": '陈科锦', "backlog_units": 108,
    },
    ('中东', '科威特'): {
        "sign_units": 136, "sign_amount": 1820.0,
        "payment": 695.4,
        "ship_units": 64, "ship_amount": 732.0,
        "schedule_units": 64, "schedule_amount": 732.0,
        "person": '赵俊峰', "backlog_units": 0,
    },
    ('中东', '阿联酋-1'): {
        "sign_units": 416, "sign_amount": 3710.0,
        "payment": 1425.0,
        "ship_units": 150, "ship_amount": 1500.0,
        "schedule_units": 150, "schedule_amount": 1500.0,
        "person": '陈科锦', "backlog_units": 8,
    },
    ('中东', '卡塔尔'): {
        "sign_units": 50, "sign_amount": 700.0,
        "payment": 475.0,
        "ship_units": 42, "ship_amount": 500.0,
        "schedule_units": 42, "schedule_amount": 500.0,
        "person": '李军辉', "backlog_units": 0,
    },
    ('中东', '伊拉克'): {
        "sign_units": 81, "sign_amount": 700.0,
        "payment": 285.0,
        "ship_units": 42, "ship_amount": 300.0,
        "schedule_units": 42, "schedule_amount": 300.0,
        "person": '吕超', "backlog_units": 1,
    },
    ('中东', '巴勒斯坦'): {
        "sign_units": 181, "sign_amount": 1680.0,
        "payment": 665.0,
        "ship_units": 80, "ship_amount": 700.0,
        "schedule_units": 80, "schedule_amount": 700.0,
        "person": '吕超', "backlog_units": 5,
    },
    ('中东', '伊朗+阿曼'): {
        "sign_units": 45, "sign_amount": 700.0,
        "payment": 285.0,
        "ship_units": 25, "ship_amount": 300.0,
        "schedule_units": 25, "schedule_amount": 300.0,
        "person": '张朕铭', "backlog_units": 17,
    },
    ('非洲', '埃及-1'): {
        "sign_units": 900, "sign_amount": 6000.0,
        "payment": 2470.0,
        "ship_units": 400, "ship_amount": 2600.0,
        "schedule_units": 400, "schedule_amount": 2600.0,
        "person": '石云龙', "backlog_units": 494,
    },
    ('非洲', '埃及-2'): {
        "sign_units": 100, "sign_amount": 800.0,
        "payment": 285.0,
        "ship_units": 30, "ship_amount": 300.0,
        "schedule_units": 30, "schedule_amount": 300.0,
        "person": '丁绎澎', "backlog_units": 0,
    },
    ('非洲', '肯尼亚/坦桑尼亚'): {
        "sign_units": 101, "sign_amount": 980.0,
        "payment": 381.749525320429,
        "ship_units": 33, "ship_amount": 401.841605600452,
        "schedule_units": 33, "schedule_amount": 401.841605600452,
        "person": '史宇哲', "backlog_units": 5,
    },
    ('非洲', '尼日利亚/埃塞俄比亚'): {
        "sign_units": 100, "sign_amount": 980.0,
        "payment": 381.9,
        "ship_units": 33, "ship_amount": 402.0,
        "schedule_units": 33, "schedule_amount": 402.0,
        "person": '袁帅', "backlog_units": 4,
    },
    ('非洲', '非洲法语区'): {
        "sign_units": 90, "sign_amount": 800.0,
        "payment": 285.0,
        "ship_units": 30, "ship_amount": 300.0,
        "schedule_units": 30, "schedule_amount": 300.0,
        "person": '孙小婷', "backlog_units": 7,
    },
    ('非洲', '南非/安格拉'): {
        "sign_units": 90, "sign_amount": 800.0,
        "payment": 285.0,
        "ship_units": 30, "ship_amount": 300.0,
        "schedule_units": 30, "schedule_amount": 300.0,
        "person": '刘纪龙', "backlog_units": 1,
    },
    ('欧洲', '德国西班牙'): {
        "sign_units": 34, "sign_amount": 520.0,
        "payment": 199.5,
        "ship_units": 15, "ship_amount": 210.0,
        "schedule_units": 15, "schedule_amount": 210.0,
        "person": '（空）', "backlog_units": 0,
    },
    ('欧洲', '东欧'): {
        "sign_units": 50, "sign_amount": 500.0,
        "payment": 190.95,
        "ship_units": 15, "ship_amount": 201.0,
        "schedule_units": 15, "schedule_amount": 201.0,
        "person": '吴雪明', "backlog_units": 0,
    },
    ('欧洲', '英国意大利'): {
        "sign_units": 34, "sign_amount": 520.0,
        "payment": 199.5,
        "ship_units": 15, "ship_amount": 210.0,
        "schedule_units": 15, "schedule_amount": 210.0,
        "person": '王俊豪', "backlog_units": 0,
    },
    ('商贸配件', '商贸1'): {
        "sign_units": 0, "sign_amount": 1402.0,
        "payment": 570.0,
        "ship_units": 0, "ship_amount": 600.0,
        "schedule_units": 0, "schedule_amount": 600.0,
        "person": '张涛', "backlog_units": 0,
    },
    ('商贸配件', '商贸2'): {
        "sign_units": 0, "sign_amount": 720.0,
        "payment": 285.0,
        "ship_units": 0, "ship_amount": 300.0,
        "schedule_units": 0, "schedule_amount": 300.0,
        "person": '李成富', "backlog_units": 0,
    },
    ('商贸配件', '商贸3'): {
        "sign_units": 0, "sign_amount": 720.0,
        "payment": 285.0,
        "ship_units": 0, "ship_amount": 300.0,
        "schedule_units": 0, "schedule_amount": 300.0,
        "person": '李美珊', "backlog_units": 0,
    },
    ('商贸配件', '配件-1'): {
        "sign_units": 0, "sign_amount": 4800.0,
        "payment": 1900.0,
        "ship_units": 0, "ship_amount": 2000.0,
        "schedule_units": 0, "schedule_amount": 2000.0,
        "person": '赵莹', "backlog_units": 0,
    },
    ('商贸配件', '配件-2'): {
        "sign_units": 0, "sign_amount": 2400.0,
        "payment": 950.0,
        "ship_units": 0, "ship_amount": 1000.0,
        "schedule_units": 0, "schedule_amount": 1000.0,
        "person": '吴航', "backlog_units": 0,
    },
    ('商贸配件', '改造'): {
        "sign_units": 0, "sign_amount": 1960.0,
        "payment": 950.0,
        "ship_units": 0, "ship_amount": 1000.0,
        "schedule_units": 0, "schedule_amount": 1000.0,
        "person": '苏利', "backlog_units": 16,
    },
}


# ── 年度完成情况 ──────────────────────────────────────

# 指标ID → 实际值字段映射 (金额用_total, 台数用原始)
_ACTUAL_FIELD = {
    'sign_units': 'sign_units_curr',
    'sign_amount': 'sign_total_curr',
    'payment': 'payment_total_curr',
    'ship_units': 'ship_units_curr',
    'ship_amount': 'ship_total_curr',
    'schedule_units': 'schedule_units_curr',
    'schedule_amount': 'schedule_total_curr',
}

_METRIC_KEYS = ['sign_units', 'sign_amount', 'schedule_units', 'schedule_amount', 'ship_units', 'ship_amount', 'payment']


def get_annual_completion():
    """
    2026年合同完成情况表
    复用两年对比的聚合结果，合并硬编码年度指标，计算完成全年比。
    金额指标使用 _total 字段（含海外差额）。
    """
    base = get_two_year_comparison()
    rows = base['rows']
    today = date.today()
    data_date = today - timedelta(days=1)

    def _ratio(actual, target):
        if target == 0 or target is None:
            return None
        return round(actual / target, 4)

    # 预计算各区域的子模块指标合计，用于 subtotal 行
    region_targets = {}
    for (r, m), t in ANNUAL_TARGETS.items():
        if r not in region_targets:
            region_targets[r] = {k: 0 for k in _METRIC_KEYS}
            region_targets[r]['_person'] = ''
            region_targets[r]['_backlog'] = 0
        for k in _METRIC_KEYS:
            region_targets[r][k] += t.get(k, 0)
        region_targets[r]['_backlog'] += t.get('backlog_units', 0)

    # 商贸合计 subtotal 使用 商贸配件 的指标汇总
    if '商贸配件' in region_targets:
        region_targets['商贸合计'] = region_targets['商贸配件']

    # 大区排序
    region_order = ['俄罗斯', '中亚', '亚洲1', '亚洲2', '美洲', '中东', '非洲', '欧洲']

    result_rows = []
    seq = 0

    for row in rows:
        typ = row['type']
        region = row.get('region', '')
        module = row.get('module', '')

        entry = {
            'type': typ,
            'region': region,
            'module': module if typ in ('data', 'trade') else (module or region),
            'category': row.get('category', ''),
        }

        # Look up target (商贸合计 rows use 商贸配件 targets)
        target = ANNUAL_TARGETS.get((region, module), {})
        if not target and region == '商贸合计':
            target = ANNUAL_TARGETS.get(('商贸配件', module), {})

        if typ in ('data', 'trade'):
            seq += 1
            entry['seq'] = seq
            for mk in _METRIC_KEYS:
                af = _ACTUAL_FIELD[mk]
                actual = row.get(af, 0) or 0
                tgt = target.get(mk, 0) if target else 0
                entry[mk + '_target'] = tgt
                entry[mk + '_actual'] = actual
                entry[mk + '_ratio'] = _ratio(actual, tgt)
            entry['person'] = target.get('person', '') if target else ''
            entry['backlog_units'] = target.get('backlog_units', 0) if target else 0

        elif typ == 'subtotal':
            # Subtota: actual from row, target from sum of children
            for mk in _METRIC_KEYS:
                af = _ACTUAL_FIELD[mk]
                actual = row.get(af, 0) or 0
                tgt = region_targets.get(region, {}).get(mk, 0)
                entry[mk + '_target'] = tgt
                entry[mk + '_actual'] = actual
                entry[mk + '_ratio'] = _ratio(actual, tgt)
            entry['person'] = ''
            entry['backlog_units'] = region_targets.get(region, {}).get('_backlog', 0)

        elif typ == 'grand_total':
            # Grand total: actual from row, target from sum of all
            all_target = {k: 0 for k in _METRIC_KEYS}
            all_backlog = 0
            for t in ANNUAL_TARGETS.values():
                for k in _METRIC_KEYS:
                    all_target[k] += t.get(k, 0)
                all_backlog += t.get('backlog_units', 0)
            for mk in _METRIC_KEYS:
                af = _ACTUAL_FIELD[mk]
                actual = row.get(af, 0) or 0
                tgt = all_target.get(mk, 0)
                entry[mk + '_target'] = tgt
                entry[mk + '_actual'] = actual
                entry[mk + '_ratio'] = _ratio(actual, tgt)
            entry['person'] = ''
            entry['backlog_units'] = all_backlog

        # 改造 columns (暂空)
        for gk in ['gaizao_sign_units', 'gaizao_sign_amount',
                    'gaizao_schedule_units', 'gaizao_schedule_amount',
                    'gaizao_ship_units', 'gaizao_ship_amount']:
            entry[gk] = 0

        result_rows.append(entry)

    return {
        'title': '2026年海外市场经营系统合同完成情况',
        'data_date': data_date.strftime('%Y-%m-%d'),
        'year': today.year,
        'metric_keys': _METRIC_KEYS,
        'region_order': region_order,
        'rows': result_rows,
    }



# ── Excel 导出 ──────────────────────────────────────────

def export_two_year_comparison_xlsx(hidden_metric_ids=None):
    """生成两年对比表 Excel 文件（所有数值为硬编码，与页面显示一致），返回文件路径
    hidden_metric_ids: 要隐藏的指标ID列表，如 ['sign_amount', 'overseas_diff']
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    data = get_two_year_comparison()
    rows = data['rows']
    groups = data['metric_groups']
    year_prev = data['year_prev']
    year_curr = data['year_curr']

    # Filter out hidden metric columns
    if hidden_metric_ids:
        hidden_set = set(hidden_metric_ids)
        groups = [g for g in groups if g['id'] not in hidden_set]

    wb = Workbook()
    ws = wb.active
    ws.title = '两年对比'

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=10)
    header_fill = PatternFill('solid', fgColor='EEF2F7')
    data_font = Font(name='微软雅黑', size=10)
    subtotal_fill = PatternFill('solid', fgColor='D6E4F0')
    grand_fill = PatternFill('solid', fgColor='B4C6E7')
    bold_font = Font(name='微软雅黑', bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 列宽：A=18, B=8, 每个指标组3列×12
    col_widths = [20, 8]
    for g in groups:
        n = 3 if g['has_growth'] else 2
        for _ in range(n):
            col_widths.append(12)

    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # Row 1: 标题
    total_cols = 2 + sum(3 if g['has_growth'] else 2 for g in groups)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=data['title'])
    c.font = Font(name='微软雅黑', bold=True, size=14)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Row 2: 表头第一层（组名）
    col = 1
    ws.cell(row=2, column=1, value='模块').font = header_font
    ws.cell(row=2, column=2, value='市场类别').font = header_font
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    col = 3
    for g in groups:
        n = 3 if g['has_growth'] else 2
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+n-1)
        c = ws.cell(row=2, column=col, value=g['name'].replace('\n', ''))
        c.font = header_font
        c.alignment = center_align
        col += n

    # Row 3: 表头第二层（年份子列）
    col = 3
    for g in groups:
        ws.cell(row=3, column=col, value=str(year_prev)).font = header_font
        ws.cell(row=3, column=col+1, value=str(year_curr)).font = header_font
        if g['has_growth']:
            ws.cell(row=3, column=col+2, value='增长比例').font = header_font
            col += 3
        else:
            col += 2
    ws.row_dimensions[3].height = 20

    # 数据行（从 row 4 开始）
    excel_row = 4

    # 预先计算每个 group 对应的列号
    group_cols = {}  # gid -> (col_prev, col_curr, col_growth)
    col = 3
    for g in groups:
        if g['has_growth']:
            group_cols[g['id']] = (col, col+1, col+2)
            col += 3
        else:
            group_cols[g['id']] = (col, col+1, None)
            col += 2

    for row_data in rows:
        r = excel_row
        typ = row_data['type']

        # 模块名和类别
        cell_a = ws.cell(row=r, column=1, value=row_data['module'])
        if typ in ('data', 'trade'):
            ws.cell(row=r, column=2, value=row_data.get('category', ''))
        elif typ in ('subtotal', 'grand_total'):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

        # 写各组数据 — 全部使用硬编码值（数据源已预计算所有合计/增长率）
        for g in groups:
            gid = g['id']
            cp, cc, cg = group_cols[gid]
            prev_key = f'{gid}_prev'
            curr_key = f'{gid}_curr'
            growth_key = f'{gid}_growth'

            pv = row_data.get(prev_key)
            cv = row_data.get(curr_key)
            if pv is not None:
                ws.cell(row=r, column=cp, value=pv)
            if cv is not None:
                ws.cell(row=r, column=cc, value=cv)
            if cg:
                gv = row_data.get(growth_key)
                if gv is not None:
                    ws.cell(row=r, column=cg, value=gv / 100)
                    ws.cell(row=r, column=cg).number_format = '0%'

        # 行样式
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

        if typ == 'subtotal':
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).fill = subtotal_fill
                ws.cell(row=r, column=c).font = bold_font
        elif typ == 'grand_total':
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).fill = grand_fill
                ws.cell(row=r, column=c).font = bold_font

        excel_row += 1

    # 表头样式
    for r in range(2, 4):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # 冻结表头
    ws.freeze_panes = 'A4'

    # 保存
    import os, tempfile
    tmp = os.path.join(tempfile.gettempdir(), 'two_year_comparison.xlsx')
    wb.save(tmp)
    return tmp



def export_annual_completion_xlsx(hide_extra=False):
    """导出年度完成情况表 Excel（双行表头、居中、整数、硬编码值）"""
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = get_annual_completion()
    rows = data['rows']
    metric_keys = data['metric_keys']

    wb = Workbook()
    ws = wb.active
    ws.title = '完成表'

    # Styles
    hdr_font = Font(name='微软雅黑', bold=True, size=10)
    hdr_fill = PatternFill('solid', fgColor='EEF2F7')
    data_font = Font(name='微软雅黑', size=10)
    subtotal_fill = PatternFill('solid', fgColor='D6E4F0')
    grand_fill = PatternFill('solid', fgColor='B4C6E7')
    bold_font = Font(name='微软雅黑', bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pct_fmt = '0.0%'
    int_fmt = '#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # ── Row 1: Group headers ──
    # Columns: 1:序号 2:类别 3:模块 | 4-24: 7 metrics x 3 | 25-30:改造 x 6 | 31:积压台数 | 32:负责人
    metric_labels = ['签订台数', '签订额', '排产台数', '排产额', '发货台数', '发货额', '回款']

    # Fixed headers
    for c, label in [(1, '序号'), (2, '类别'), (3, '模块')]:
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    # Metric group headers (row 1, colspan=3)
    for i, ml in enumerate(metric_labels):
        c = 4 + i * 3
        cell = ws.cell(row=1, column=c, value=ml)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c+2)

    # Gaizao group header (row 1, colspan=6) — only if not hiding extra
    gaizao_start = 4 + 7 * 3  # = 25
    if not hide_extra:
        cell = ws.cell(row=1, column=gaizao_start, value='改造')
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=gaizao_start, end_row=1, end_column=gaizao_start+5)

        # Backlog + Person headers
        cell = ws.cell(row=1, column=gaizao_start+6, value='积压台数')
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=gaizao_start+6, end_row=2, end_column=gaizao_start+6)
        cell = ws.cell(row=1, column=gaizao_start+7, value='负责人')
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=gaizao_start+7, end_row=2, end_column=gaizao_start+7)

    # ── Row 2: Sub-headers ──
    for c in [1, 2, 3]:
        cell = ws.cell(row=2, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin_border
    for i in range(7):
        c = 4 + i * 3
        for j, label in enumerate(['指标', '实际完成', '完成全年比']):
            cell = ws.cell(row=2, column=c+j, value=label)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
    if not hide_extra:
        gaizao_subs = ['签单台数','签单额','排产台数','排产额','发货台数','发货额']
        for j, label in enumerate(gaizao_subs):
            cell = ws.cell(row=2, column=gaizao_start+j, value=label)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border

    # Freeze header rows
    ws.freeze_panes = 'A3'

    # ── Data rows ──
    excel_row = 3  # data starts at row 3 (rows 1-2 are headers)

    for row in rows:
        typ = row['type']
        region = row.get('region', '')
        module = row.get('module', '')
        is_merged = typ in ('subtotal', 'grand_total')

        # Columns 1-3: merge for subtotal/grand_total
        if is_merged:
            cell = ws.cell(row=excel_row, column=1, value=module)
            cell.font = bold_font; cell.alignment = center; cell.border = thin_border
            ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=3)
            # Fill merged cells
            for c in range(1, 4):
                ws.cell(row=excel_row, column=c).border = thin_border
        else:
            ws.cell(row=excel_row, column=1, value=row.get('seq', ''))
            ws.cell(row=excel_row, column=2, value=row.get('category', ''))

        if not is_merged:
            ws.cell(row=excel_row, column=3, value=module)

        # Metric values — 全部使用硬编码值（数据源已预计算所有合计/比率）
        for i, mk in enumerate(metric_keys):
            col = 4 + i * 3
            tgt = row.get(mk + '_target')
            act = row.get(mk + '_actual')
            rat = row.get(mk + '_ratio')

            if tgt is not None:
                ws.cell(row=excel_row, column=col, value=tgt).number_format = int_fmt
            if act is not None:
                ws.cell(row=excel_row, column=col+1, value=act).number_format = int_fmt
            if rat is not None:
                ws.cell(row=excel_row, column=col+2, value=rat).number_format = pct_fmt

        # Gaizao columns + backlog + person (only if not hiding extra)
        col = gaizao_start
        if not hide_extra:
            for gk in ['gaizao_sign_units','gaizao_sign_amount','gaizao_schedule_units',
                        'gaizao_schedule_amount','gaizao_ship_units','gaizao_ship_amount']:
                ws.cell(row=excel_row, column=col, value=row.get(gk, 0) or 0)
                col += 1
            # Backlog
            c = ws.cell(row=excel_row, column=col, value=row.get('backlog_units', 0))
            c.number_format = int_fmt; col += 1
            # Person
            ws.cell(row=excel_row, column=col, value=row.get('person', ''))
        else:
            col = gaizao_start  # no extra columns

        # Style all cells in this row
        last_col = col
        for c in range(1, last_col + 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = bold_font if is_merged else data_font
            cell.border = thin_border
            cell.alignment = center
            if typ == 'subtotal':
                cell.fill = subtotal_fill
            elif typ == 'grand_total':
                cell.fill = grand_fill

        excel_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 14
    for i in range(7):
        c_base = 4 + i * 3
        ws.column_dimensions[get_column_letter(c_base)].width = 11
        ws.column_dimensions[get_column_letter(c_base+1)].width = 11
        ws.column_dimensions[get_column_letter(c_base+2)].width = 9

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    tmpdir = os.path.join(os.path.dirname(__file__), '..', '..', 'uploads')
    os.makedirs(tmpdir, exist_ok=True)
    filepath = os.path.join(tmpdir, 'annual_completion_export.xlsx')
    wb.save(filepath)
    return filepath