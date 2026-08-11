"""
import_budget.py — 月度预算表 Excel → JSON 转换工具
============================================================
用法:
  py -3.11 import_budget.py "8月预计完成表.xlsx"
  py -3.11 import_budget.py "8月预计完成表.xlsx" --month 2026-08

输出:
  budget_data/预算预测_YYYY-MM.json

自动适配双行表头（Row1=标题/Row2=表头）、跳过合计行、关键词匹配列。
"""

import os
import re
import sys
import json
import openpyxl

# ── 预算 Excel 列名关键词（扩展版，兼容多种命名） ──
BUDGET_COL_KEYWORDS = {
    "module": ["模块", "区域/模块"],
    "sign_units": ["签订台数", "签单台数", "新签梯量", "签单台"],
    "sign_amount": ["签订额", "签单额", "签订金额", "签单金额", "新签金额"],
    "prod_units": ["排产台数", "排产台", "生产台数", "排产"],
    "prod_amount": ["排产额", "排产金额"],
    "ship_units": ["发货台数", "发货台"],
    "ship_amount": ["发货额", "发货金额"],
}

# JSON 输出目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BUDGET_DATA_DIR = os.path.join(SCRIPT_DIR, "budget_data")


def _safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return 0.0


def _find_col_by_keywords(headers, keywords):
    """在表头列表中按关键词查找列索引（0-based），返回第一个匹配"""
    for kw in keywords:
        for idx, h in enumerate(headers):
            if h and kw in str(h):
                return idx
    return None


def _is_subtotal_or_total(module_name, region_name=""):
    """判断是否为小计行或合计行"""
    mn = str(module_name or "").strip()
    rn = str(region_name or "").strip()
    combined = mn + rn
    # 跳过包含合计/总计/小计的行
    for kw in ["合计", "总计", "小计"]:
        if kw in combined or kw in mn:
            return True
    return False


def _extract_month(excel_path):
    """从文件名或 Sheet 名提取月份: '8月预计完成表.xlsx' → '2026-08'"""
    import datetime
    basename = os.path.basename(excel_path)
    # 尝试从文件名提取月份数字
    m = re.search(r"(\d+)\s*月", basename)
    if m:
        month_num = int(m.group(1))
        year = datetime.datetime.now().year
        return f"{year}-{month_num:02d}"
    # 尝试从 Sheet 名提取
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for sn in wb.sheetnames:
            m = re.search(r"(\d+)\s*月", sn)
            if m:
                month_num = int(m.group(1))
                year = datetime.datetime.now().year
                wb.close()
                return f"{year}-{month_num:02d}"
        wb.close()
    except Exception:
        pass
    # 默认当月
    now = datetime.datetime.now()
    return f"{now.year}-{now.month:02d}"


def import_budget(excel_path, target_month=None):
    """
    读取预算 Excel → 生成 JSON 文件。

    Args:
        excel_path: 预算 Excel 文件路径
        target_month: 目标月份，如 "2026-08"。为 None 时自动从文件名提取。

    Returns:
        {"success": True/False, "message": "...", "json_path": "...", "module_count": int}
    """
    if target_month is None:
        target_month = _extract_month(excel_path)

    if not os.path.isfile(excel_path):
        return {"success": False, "message": f"文件不存在: {excel_path}"}

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── 定位数据 Sheet ──
    ws = _find_data_sheet(wb)
    if ws is None:
        wb.close()
        return {"success": False, "message": f"未找到数据Sheet。可用Sheet: {wb.sheetnames}"}

    print(f"[导入] Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

    # ── 检测表头行 ──
    header_row, headers = _detect_headers(ws)
    if not headers:
        wb.close()
        return {"success": False, "message": "无法识别表头行，请确认Excel格式"}

    print(f"[导入] 表头行: Row {header_row + 1}")
    print(f"[导入] 表头: {headers}")

    # ── 关键词匹配列 ──
    col_module = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["module"])
    col_sign_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["sign_units"])
    col_sign_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["sign_amount"])
    col_prod_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["prod_units"])
    col_prod_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["prod_amount"])
    col_ship_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["ship_units"])
    col_ship_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["ship_amount"])

    # col_region 可选（用于辅助跳过合计行）
    col_region = _find_col_by_keywords(headers, ["大区", "区域"])

    if col_module is None:
        wb.close()
        return {"success": False, "message": "未找到模块名列，请检查表头是否包含'模块'关键词"}

    missing = []
    if col_sign_units is None:
        missing.append("签单台数")
    if col_sign_amount is None:
        missing.append("签单额")
    if col_prod_units is None:
        missing.append("排产台数")
    if col_prod_amount is None:
        missing.append("排产额")
    if col_ship_units is None:
        missing.append("发货台数")
    if col_ship_amount is None:
        missing.append("发货额")
    if missing:
        print(f"[警告] 未找到列: {', '.join(missing)}，将使用默认值 0")

    # ── 逐行读取 ──
    forecast = {}
    skipped_subtotal = 0
    skipped_empty = 0

    for r in range(header_row + 1, ws.max_row):  # 0-based
        # 序号列不是纯数字 → 跳过（合计行/项目明细行/空行）
        seq_val = str(ws.cell(row=r + 1, column=1).value or "").strip()
        if seq_val and not seq_val.isdigit():
            continue

        module_name = str(ws.cell(row=r + 1, column=col_module + 1).value or "").strip()
        if not module_name:
            skipped_empty += 1
            continue

        # 获取大区名（如存在）
        region_name = ""
        if col_region is not None:
            region_name = str(ws.cell(row=r + 1, column=col_region + 1).value or "").strip()

        # 跳过合计/小计/总计行
        if _is_subtotal_or_total(module_name, region_name):
            skipped_subtotal += 1
            continue

        forecast[module_name] = {
            "sign_units": _safe_float(ws.cell(row=r + 1, column=col_sign_units + 1).value) if col_sign_units is not None else 0,
            "sign_amount": _safe_float(ws.cell(row=r + 1, column=col_sign_amount + 1).value) if col_sign_amount is not None else 0,
            "prod_units": _safe_float(ws.cell(row=r + 1, column=col_prod_units + 1).value) if col_prod_units is not None else 0,
            "prod_amount": _safe_float(ws.cell(row=r + 1, column=col_prod_amount + 1).value) if col_prod_amount is not None else 0,
            "ship_units": _safe_float(ws.cell(row=r + 1, column=col_ship_units + 1).value) if col_ship_units is not None else 0,
            "ship_amount": _safe_float(ws.cell(row=r + 1, column=col_ship_amount + 1).value) if col_ship_amount is not None else 0,
        }

    wb.close()

    # ── 保存 JSON ──
    os.makedirs(BUDGET_DATA_DIR, exist_ok=True)
    json_filename = f"budget_forecast_{target_month}.json"
    json_path = os.path.join(BUDGET_DATA_DIR, json_filename)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(forecast, f, ensure_ascii=False, indent=2)

    msg = (f"导入完成: {len(forecast)} 个模块 → {json_filename}"
           f"（跳过 {skipped_subtotal} 合计行, {skipped_empty} 空行）")
    print(f"[导入] {msg}")

    # 预览前几个
    for mod_name, data in list(forecast.items())[:5]:
        if any(v != 0 for v in data.values()):
            print(f"  {mod_name}: 签{data['sign_amount']:.0f}万/{data['sign_units']:.0f}台, "
                  f"排{data['prod_amount']:.0f}万/{data['prod_units']:.0f}台, "
                  f"发{data['ship_amount']:.0f}万/{data['ship_units']:.0f}台")

    return {"success": True, "message": msg, "json_path": json_path, "module_count": len(forecast)}


def _find_data_sheet(wb):
    """定位数据 Sheet：优先非空前有数据的（排除 Sheet1 这种空 sheet）"""
    candidates = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row > 1 and ws.max_column > 1:
            # check if there's actual data
            if ws.cell(row=2, column=1).value is not None or ws.cell(row=1, column=1).value is not None:
                candidates.append(sn)
    if candidates:
        return wb[candidates[0]]
    return wb.active if wb.active and wb.active.max_row > 1 else None


def _detect_headers(ws):
    """
    检测表头所在行（0-based）。
    - 先读 Row 1（index 0）：如果大部分列有值 → 这就是表头
    - 如果 Row 1 有效值 ≤ 2 → 可能是标题行，读 Row 2 作为表头
    - 返回 (row_index, [header_strings])
    """
    # 读 Row 1
    r1_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        r1_vals.append(str(v).strip() if v is not None else "")
    r1_non_empty = sum(1 for v in r1_vals if v)

    # 如果 Row 1 有足够的列标题（≥4个有效值），用它
    if r1_non_empty >= 4:
        return 0, r1_vals

    # 否则读 Row 2
    r2_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        r2_vals.append(str(v).strip() if v is not None else "")
    r2_non_empty = sum(1 for v in r2_vals if v)

    if r2_non_empty >= 4:
        return 1, r2_vals

    # 最后尝试 Row 1
    return 0, r1_vals


# ── 命令行入口 ──
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="月度预算表 Excel → JSON")
    parser.add_argument("excel", help="预算 Excel 文件路径")
    parser.add_argument("--month", "-m", help="目标月份，如 2026-08（默认从文件名提取）", default=None)
    args = parser.parse_args()

    result = import_budget(args.excel, args.month)
    print(f"\n{'=' * 50}")
    if result["success"]:
        print(f"[OK] {result['message']}")
        print(f"  文件: {result['json_path']}")
    else:
        print(f"[FAIL] {result['message']}")
        sys.exit(1)
