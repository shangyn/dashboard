"""
generate_report_v3.py - 国贸台账统计报表生成 (V3)
============================================================
改造点：台账/映射/报表A/B 数据从数据库查询，预算从上传文件读取。
替代 generate_wrapper.py + generate_report.py + generate_report_june.py 的子进程调用链。
可直接被 Flask import 调用，无需 subprocess。
"""
import os
import re
import datetime
from collections import OrderedDict

import openpyxl

# 复用 generate_report_june 的 fill_template_dynamic（模板填充逻辑不变）
from generate_report_june import fill_template_dynamic

# 复用 generate_dashboard 的 HTML 生成
from generate_dashboard import read_dashboard_data, read_table_rows, generate_html


# ============================================================
# 配置
# ============================================================
TEMPLATE_NAME = "签单排产发货_模板.xlsx"
REGION_ORDER = ["俄罗斯", "亚洲1", "亚洲2", "中东", "中亚", "美洲", "非洲", "欧洲"]
TRADE_PARTS_MODULES = ["商贸1", "商贸2", "商贸3", "配件-1", "配件-2", "改造"]

# 预算 Excel 列名关键词（按列名搜索，兼容"签单"和"签订"两种命名）
BUDGET_COL_KEYWORDS = {
    "module": ["模块", "区域/模块"],
    "sign_units": ["签订台数", "签单台数", "新签梯量", "签单台"],
    "sign_amount": ["签订额", "签单额", "签订金额", "签单金额", "新签金额"],
    "prod_units": ["排产台数", "排产台", "生产台数", "排产"],
    "prod_amount": ["排产额", "排产金额"],
    "ship_units": ["发货台数", "发货台"],
    "ship_amount": ["发货额", "发货金额"],
}


def _safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('', '-', '#DIV/0!', '#N/A', '#REF!', '#VALUE!'):
        return 0.0
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return 0.0


def _find_col_by_keywords(headers, keywords):
    """在表头行中按关键词查找列索引（1-based）"""
    for kw in keywords:
        for idx, h in enumerate(headers, 1):
            if h and kw in str(h):
                return idx
    return None


def build_forecast_from_budget(budget_file_path, target_month):
    """从预算 Excel 读取预测数据。
    返回: {module_name: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount}}
    """
    wb = openpyxl.load_workbook(budget_file_path, data_only=True)
    ws = wb.active

    # 读表头行（第1行）
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=c).value or "").strip())

    # 按关键词找列
    col_module = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["module"])
    col_sign_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["sign_units"])
    col_sign_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["sign_amount"])
    col_prod_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["prod_units"])
    col_prod_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["prod_amount"])
    col_ship_units = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["ship_units"])
    col_ship_amount = _find_col_by_keywords(headers, BUDGET_COL_KEYWORDS["ship_amount"])

    if not col_module:
        wb.close()
        raise ValueError("预算文件中未找到模块名列，请检查表头是否包含'模块'关键词")

    forecast = {}
    for r in range(2, ws.max_row + 1):
        module_name = str(ws.cell(row=r, column=col_module).value or "").strip()
        if not module_name:
            continue

        forecast[module_name] = {
            "sign_units": _safe_float(ws.cell(row=r, column=col_sign_units).value) if col_sign_units else 0,
            "sign_amount": _safe_float(ws.cell(row=r, column=col_sign_amount).value) if col_sign_amount else 0,
            "prod_units": _safe_float(ws.cell(row=r, column=col_prod_units).value) if col_prod_units else 0,
            "prod_amount": _safe_float(ws.cell(row=r, column=col_prod_amount).value) if col_prod_amount else 0,
            "ship_units": _safe_float(ws.cell(row=r, column=col_ship_units).value) if col_ship_units else 0,
            "ship_amount": _safe_float(ws.cell(row=r, column=col_ship_amount).value) if col_ship_amount else 0,
        }

    wb.close()
    print(f"[V3] 预算文件读取完成: {len(forecast)} 个模块")
    return forecast


def build_forecast_from_json():
    """从 budget_data/ 目录读取最新的预算 JSON 文件。

    自动选择月份最新的 JSON（文件名格式：预算预测_YYYY-MM.json），
    无需手动选择月份。

    Returns:
        {module_name: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount}}

    Raises:
        FileNotFoundError: budget_data/ 目录下没有预算 JSON 文件
    """
    import json
    import glob

    script_dir = os.path.dirname(os.path.abspath(__file__))
    budget_dir = os.path.join(script_dir, "budget_data")

    if not os.path.isdir(budget_dir):
        raise FileNotFoundError(
            f"预算数据目录不存在: {budget_dir}。请先运行 import_budget.py 导入预算表。"
        )

    json_files = glob.glob(os.path.join(budget_dir, "budget_forecast_*.json"))
    if not json_files:
        raise FileNotFoundError(
            f"未找到预算JSON文件。请先将月度预算表放入 {budget_dir} 目录，"
            "或运行 import_budget.py 导入预算表。"
        )

    # 按文件名中月份排序，取最新
    json_files.sort(reverse=True)
    latest_file = json_files[0]

    with open(latest_file, "r", encoding="utf-8") as f:
        forecast = json.load(f)

    month_str = os.path.basename(latest_file).replace("budget_forecast_", "").replace(".json", "")
    print(f"[V3] 预算JSON读取完成: {len(forecast)} 个模块 ← {month_str}")
    return forecast


def build_stats_from_db(target_month):
    """从数据库查询台账/报表A/B数据，按模块汇总签单/排产/发货。
    返回: {module_name: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount}}
    """
    from dashboards.contract_completion.models import LedgerContract

    # 加载映射表
    mapping = {}
    from dashboards.contract_completion.models import CountryMapping
    for m in CountryMapping.query.all():
        mapping[m.country] = m

    # 加载所有台账数据（不限 source）
    contracts = LedgerContract.query.all()

    # 汇总
    from collections import defaultdict
    stats = defaultdict(lambda: {
        "sign_units": 0, "sign_amount": 0,
        "prod_units": 0, "prod_amount": 0,
        "ship_units": 0, "ship_amount": 0,
    })

    month_prefix = target_month  # "2026-08"

    for c in contracts:
        # 排除已作废合同
        if c.product_status and str(c.product_status).strip() == '已作废':
            continue

        # 排除改造梯（产品型号含"改造"，与外购梯不同，改造梯不参与签排发统计）
        if c.product_type and '改造' in str(c.product_type):
            continue

        # 确定模块名
        module_name = None
        country = c.country or ""

        # 先看映射表
        if country in mapping:
            module_name = mapping[country].module_name
        # 再看 mapped_module（已在导入时填充的）
        if not module_name and c.mapped_module:
            module_name = c.mapped_module
        # 商贸配件来源
        if c.source in ('report_a', 'report_b') and c.mapped_module:
            module_name = c.mapped_module
        if not module_name:
            continue

        # 签单
        if c.sign_date and str(c.sign_date).startswith(month_prefix):
            stats[module_name]["sign_units"] += c.unit_count or 0
            # 金额单位转换：元 → 万元
            stats[module_name]["sign_amount"] += (c.contract_amount_rmb or 0) / 10000

        # 排产
        if c.schedule_date and str(c.schedule_date).startswith(month_prefix):
            stats[module_name]["prod_units"] += c.unit_count or 0
            stats[module_name]["prod_amount"] += (c.contract_amount_rmb or 0) / 10000

        # 发货
        if c.delivery_date and str(c.delivery_date).startswith(month_prefix):
            stats[module_name]["ship_units"] += c.unit_count or 0
            stats[module_name]["ship_amount"] += (c.contract_amount_rmb or 0) / 10000

    print(f"[V3] DB 统计完成: {len(stats)} 个模块, {len(contracts)} 条合同")
    return dict(stats)


def build_region_modules_from_db():
    """从数据库映射表构建 {大区: [模块名列表]}。
    返回: OrderedDict
    """
    from dashboards.contract_completion.models import CountryMapping

    region_modules = OrderedDict()
    for region in REGION_ORDER:
        region_modules[region] = []

    # 从 CountryMapping 收集每个大区下的模块
    seen = set()
    for m in CountryMapping.query.order_by(CountryMapping.region, CountryMapping.module_name).all():
        region = m.region or ""
        module = m.module_name or ""
        if not region or not module:
            continue
        if region not in region_modules:
            region_modules[region] = []
        key = (region, module)
        if key not in seen:
            region_modules[region].append(module)
            seen.add(key)

    # 去除空大区
    region_modules = OrderedDict((k, v) for k, v in region_modules.items() if v)

    print(f"[V3] 映射表加载完成: {len(region_modules)} 个大区, {sum(len(v) for v in region_modules.values())} 个模块")
    return region_modules


def generate_report(target_month, budget_file_path=None, output_excel_path=None):
    """
    主入口：生成签单排产发货 Excel 报表。

    Args:
        target_month: 目标月份，如 "2026-08"
        budget_file_path: 预算 Excel 文件路径（可选，为 None 时从 JSON 读取）
        output_excel_path: 输出 Excel 路径

    Returns:
        {"success": True/False, "message": "...", "excel_path": "..."}
    """
    try:
        # 模板路径（相对于 generate_dashboard 目录）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, TEMPLATE_NAME)
        if not os.path.isfile(template_path):
            # 尝试数据源目录
            alt = os.path.join(script_dir, "数据源excel", TEMPLATE_NAME)
            if os.path.isfile(alt):
                template_path = alt
            else:
                return {"success": False, "message": f"模板文件不存在: {TEMPLATE_NAME}", "excel_path": None}

        # 1. 预算数据（优先从 JSON 读取，否则从上传文件）
        if budget_file_path and os.path.isfile(budget_file_path):
            print(f"[V3] 读取预算文件: {budget_file_path}")
            forecast = build_forecast_from_budget(budget_file_path, target_month)
        else:
            print(f"[V3] 从 JSON 读取预算数据...")
            forecast = build_forecast_from_json()

        # 2. 台账/报表A/B 数据（从DB）
        print(f"[V3] 从数据库查询台账数据...")
        stats = build_stats_from_db(target_month)

        # 3. 映射表（从DB）
        print(f"[V3] 从数据库加载映射表...")
        region_modules = build_region_modules_from_db()

        # 4. 填充模板 → Excel
        print(f"[V3] 填充模板 → {output_excel_path}")
        fill_template_dynamic(
            template_path=template_path,
            forecast=forecast,
            stats=stats,
            region_modules=region_modules,
            target_month=target_month,
            output_path=output_excel_path,
        )

        print(f"[V3] 报表生成成功: {output_excel_path}")
        return {"success": True, "message": "报表生成成功", "excel_path": output_excel_path}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"报表生成失败: {str(e)}", "excel_path": None}


def generate_dashboard_html(target_month, excel_path, output_html_path):
    """
    从已生成的 Excel 生成 HTML 看板。

    Args:
        target_month: 目标月份
        excel_path: 签单排产发货 Excel 路径
        output_html_path: 输出 HTML 路径

    Returns:
        {"success": True/False, "message": "...", "html_path": "..."}
    """
    try:
        print(f"[V3] 读取报表数据: {excel_path}")
        region_data, grand = read_dashboard_data(excel_path)
        table_rows = read_table_rows(excel_path)

        print(f"[V3] 生成 HTML 看板: {output_html_path}")
        generate_html(
            region_data=region_data,
            grand=grand,
            target_month=target_month,
            output_path=output_html_path,
            table_rows=table_rows,
            excel_path=excel_path,
        )

        print(f"[V3] HTML 看板生成成功: {output_html_path}")
        return {"success": True, "message": "看板生成成功", "html_path": output_html_path}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"看板生成失败: {str(e)}", "html_path": None}
