#!py -3.11
"""
generate_report_june.py - 6月国贸台账统计报表生成
============================================================
用法:
  py -3.11 generate_report_june.py --extract          # 一次性：从预算表抽取6月预测数据→JSON
  py -3.11 generate_report_june.py --month 2026-06    # 生成6月报表（含配件填充）

与5月 generate_report.py 的区别：
  预测数据来源 = 国际市场运营系统合同预算情况表（而非同期对比表）
  预测值从预先生成的JSON文件读取，整个6月复用同一份预测数据。
  此脚本同时包含了报表生成 + 配件数据填充，是6月的一站式入口。
============================================================
"""
import argparse
import sys
import os
import json
import datetime
import openpyxl
import re
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

# ---- 复用 generate_report 的核心函数（不修改原文件）----
from generate_report import (
    find_file, safe_float, find_columns_by_header,
    build_mapping, process_ledger,
    recalc_with_excel,
    LEDGER_KW, MAPPING_KW, TEMPLATE_NAME, MAPPING_SHEET,
    # 列映射辅助（复用，不重写）
    _normal_col_rules, _trade_col_rules, _ratio_pair, _col_to_field,
    _AMOUNT_COLS, FMT_UNITS, FMT_AMOUNT, FMT_RATIO,
    TRADE_PARTS_KW,
)

# ============================================================
# 配置
# ============================================================
BUDGET_JSON = "预算预测_2026-07.json"
BUDGET_KW = ["预算", "国际市场"]

# 大区排列顺序（与模板一致）
REGION_ORDER = ["俄罗斯", "亚洲1", "亚洲2", "中东", "中亚", "美洲", "非洲", "欧洲"]

# 商贸配件模块（不在映射表中，固定输出）
TRADE_PARTS_MODULES = ["商贸1", "商贸2", "商贸3", "配件-1", "配件-2", "改造"]

# 合计行颜色
SUBTOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # 大区合计=浅蓝
GRAND_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")     # 全国贸合计=深蓝


def _ratio_formula(fc_col, ac_col, row_num):
    """完成比公式: =IF(预测=0,IF(实际=0,0,1),实际/预测)"""
    return f"=IF({fc_col}{row_num}=0,IF({ac_col}{row_num}=0,0,1),{ac_col}{row_num}/{fc_col}{row_num})"


def _apply_style(cell, col_idx, ref_styles):
    """将模板参考样式（数字格式/字体/对齐/边框）应用到单元格"""
    if col_idx in ref_styles:
        ref = ref_styles[col_idx]
        cell.number_format = ref["number_format"]
        if ref["font"]:
            cell.font = ref["font"]
        if ref["alignment"]:
            cell.alignment = ref["alignment"]
        if ref["border"]:
            cell.border = ref["border"]


def fill_template_dynamic(template_path, forecast, stats, region_modules,
                          target_month, output_path):
    """动态生成报表 — 映射表驱动模块布局，模板仅提供表头/格式/商贸配件

    Args:
        template_path: 模板Excel路径
        forecast: {模块名: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount}}
        stats: {模块名: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount}}
        region_modules: OrderedDict {大区名: [模块名列表]}
        target_month: 目标月份 (如 "2026-06")
        output_path: 输出路径
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    month_num = int(target_month.split("-")[1])

    # ---- 1. 标题月份替换 + 日期更新 ----
    title_text = str(ws.cell(row=1, column=1).value or "")
    tm = re.search(r"(\d+)月", title_text)
    template_month = int(tm.group(1)) if tm else month_num

    if template_month != month_num:
        for rn in [1, 3, 4]:
            for c in range(1, 21):
                cell = ws.cell(row=rn, column=c)
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace(f"{template_month}月", f"{month_num}月")

    ws.cell(row=2, column=1).value = f"报表日期：{datetime.date.today()}"

    # ---- 2. 保存模板第5行的样式作为模块行参考 ----
    ref_styles = {}
    for c in range(1, 21):
        ref_cell = ws.cell(row=5, column=c)
        ref_styles[c] = {
            "number_format": ref_cell.number_format,
            "font": ref_cell.font.copy() if ref_cell.font else None,
            "alignment": ref_cell.alignment.copy() if ref_cell.alignment else None,
            "border": ref_cell.border.copy() if ref_cell.border else None,
        }

    # ---- 3. 删除模板行5及之后所有行 ----
    total_rows = ws.max_row
    if total_rows >= 5:
        ws.delete_rows(5, total_rows - 4)

    # ---- 4. 动态生成8大区模块行 + 子合计 ----
    current_row = 5
    seq = 1
    subtotal_rows = []  # 记录子合计行号，用于全国贸合计
    normal_col_rules = _normal_col_rules()
    trade_col_rules = _trade_col_rules()
    ratio_cols = ["E", "H", "K", "N", "Q", "T"]  # 完成比列
    value_cols = "CDEFGHIJKLMNOPQRST"
    grand_totals = {}  # 全国贸合计：累计各大区/商贸配件合计

    all_filled_modules = set()  # 追踪已输出的模块，用于日志

    for region in REGION_ORDER:
        modules = region_modules.get(region, [])
        if not modules:
            print(f"  [跳过] 大区 [{region}] 无模块数据")
            continue

        region_start_row = current_row
        region_totals = {}  # 当前大区各列合计（完成比列除外）

        for mod_name in modules:
            fc = forecast.get(mod_name, {})
            st = stats.get(mod_name, {})
            all_filled_modules.add(mod_name)

            # A列: 序号
            ws.cell(row=current_row, column=1).value = seq
            # B列: 模块名
            ws.cell(row=current_row, column=2).value = mod_name

            # C-T列: 按普通模块规则填充
            for col_letter, (col_type, group) in normal_col_rules.items():
                col_idx = column_index_from_string(col_letter)
                cell = ws.cell(row=current_row, column=col_idx)

                # 应用模板样式（数字格式/字体/对齐/边框）
                _apply_style(cell, col_idx, ref_styles)

                if col_type == "forecast":
                    field = f"{group}_{'amount' if col_letter in _AMOUNT_COLS else 'units'}"
                    cell.value = fc.get(field, 0)
                elif col_type == "actual":
                    field = f"{group}_{'amount' if col_letter in _AMOUNT_COLS else 'units'}"
                    cell.value = st.get(field, 0)
                elif col_type == "ratio":
                    # 直接写入完成比数值（不再写公式，避免手机端不重算显示0）
                    suffix = "amount" if col_letter in _AMOUNT_COLS else "units"
                    field = f"{group}_{suffix}"
                    fc_val = float(fc.get(field, 0) or 0)
                    ac_val = float(st.get(field, 0) or 0)
                    cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
                    cell.number_format = FMT_RATIO

            # A/B列应用模板样式
            _apply_style(ws.cell(row=current_row, column=1), 1, ref_styles)
            _apply_style(ws.cell(row=current_row, column=2), 2, ref_styles)

            seq += 1
            current_row += 1

        # 累计该大区各列合计（读取已写入的数值）
        for r in range(region_start_row, current_row):
            for col_letter in value_cols:
                if col_letter in ratio_cols:
                    continue
                v = ws.cell(row=r, column=column_index_from_string(col_letter)).value
                if isinstance(v, (int, float)):
                    region_totals[col_letter] = region_totals.get(col_letter, 0) + v

        # ---- 写子合计行 ----
        region_end_row = current_row - 1
        subtotal_rows.append(current_row)

        # 子合计A列：标签 + 浅蓝背景 + 样式
        subtotal_a = ws.cell(row=current_row, column=1)
        subtotal_a.value = f"{region}合计"
        subtotal_a.fill = SUBTOTAL_FILL
        _apply_style(subtotal_a, 1, ref_styles)
        # A+B列合并（与模板一致）
        try:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=2)
        except Exception:
            pass

        CL = get_column_letter
        for col_letter in "CDEFGHIJKLMNOPQRST":
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=current_row, column=col_idx)

            # 浅蓝背景
            cell.fill = SUBTOTAL_FILL

            # 应用模板样式（含边框）+ 粗体
            _apply_style(cell, col_idx, ref_styles)
            bold_font = cell.font.copy() if cell.font else openpyxl.styles.Font()
            bold_font.bold = True
            cell.font = bold_font

            if col_letter in ratio_cols:  # 完成比列
                fc_col, ac_col = _ratio_pair(col_letter)
                fc_val = region_totals.get(fc_col, 0)
                ac_val = region_totals.get(ac_col, 0)
                cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
                cell.number_format = FMT_RATIO
            else:
                # 直接用累计值写入合计（不再写SUM公式，兼容手机端）
                cell.value = region_totals.get(col_letter, 0)

        # 累计到全国贸合计
        for col_letter in value_cols:
            grand_totals[col_letter] = grand_totals.get(col_letter, 0) + region_totals.get(col_letter, 0)

        current_row += 1

    # ---- 5. 写商贸配件 ----
    trade_start_row = current_row
    trade_totals = {}  # 商贸配件各列合计（完成比列除外）
    for mod_name in TRADE_PARTS_MODULES:
        fc = forecast.get(mod_name, {})
        st = stats.get(mod_name, {})
        all_filled_modules.add(mod_name)

        ws.cell(row=current_row, column=1).value = seq
        ws.cell(row=current_row, column=2).value = mod_name
        # A/B列应用模板样式
        _apply_style(ws.cell(row=current_row, column=1), 1, ref_styles)
        _apply_style(ws.cell(row=current_row, column=2), 2, ref_styles)

        for col_letter, rule in trade_col_rules.items():
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=current_row, column=col_idx)

            # 应用模板样式（数字格式/字体/对齐/边框）
            _apply_style(cell, col_idx, ref_styles)

            if rule == "dash":
                cell.value = "-"
                cell.number_format = "@"
            elif rule == "zero":
                if mod_name == "改造":
                    field = _col_to_field(col_letter)
                    cell.value = st.get(field, 0) if field else 0
                else:
                    cell.value = 0
            elif rule == "import":
                field = _col_to_field(col_letter)
                cell.value = fc.get(field, 0) if field else 0
            elif rule == "formula":
                # 直接写入完成比数值（不再写公式，兼容手机端）
                fc_col, ac_col = _ratio_pair(col_letter)
                fc_val = ws.cell(row=current_row, column=column_index_from_string(fc_col)).value
                ac_val = ws.cell(row=current_row, column=column_index_from_string(ac_col)).value
                fc_val = float(fc_val) if isinstance(fc_val, (int, float)) else 0
                ac_val = float(ac_val) if isinstance(ac_val, (int, float)) else 0
                cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
                cell.number_format = FMT_RATIO

        seq += 1
        current_row += 1

    # 累计商贸配件各列合计（读取已写入的数值）
    for r in range(trade_start_row, current_row):
        for col_letter in value_cols:
            if col_letter in ratio_cols:
                continue
            v = ws.cell(row=r, column=column_index_from_string(col_letter)).value
            if isinstance(v, (int, float)):
                trade_totals[col_letter] = trade_totals.get(col_letter, 0) + v

    # ---- 6. 写商贸配件合计 ----
    trade_end_row = current_row - 1
    subtotal_rows.append(current_row)

    # 商贸配件合计A列：标签 + 浅蓝背景 + 样式
    trade_sub_a = ws.cell(row=current_row, column=1)
    trade_sub_a.value = "商贸配件合计"
    trade_sub_a.fill = SUBTOTAL_FILL
    _apply_style(trade_sub_a, 1, ref_styles)
    try:
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2)
    except Exception:
        pass

    CL = get_column_letter
    for col_letter in "CDEFGHIJKLMNOPQRST":
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = SUBTOTAL_FILL

        # 应用模板样式（含边框）+ 粗体
        _apply_style(cell, col_idx, ref_styles)
        bold_font = cell.font.copy() if cell.font else openpyxl.styles.Font()
        bold_font.bold = True
        cell.font = bold_font

        if col_letter in ratio_cols:
            fc_col, ac_col = _ratio_pair(col_letter)
            fc_val = trade_totals.get(fc_col, 0)
            ac_val = trade_totals.get(ac_col, 0)
            cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
            cell.number_format = FMT_RATIO
        else:
            cell.value = trade_totals.get(col_letter, 0)

    # 商贸配件合计并入全国贸合计
    for col_letter in value_cols:
        grand_totals[col_letter] = grand_totals.get(col_letter, 0) + trade_totals.get(col_letter, 0)

    current_row += 1

    # ---- 7. 写全国贸合计 ----
    grand_row = current_row
    # 全国贸合计A列：标签 + 深蓝背景 + 样式
    grand_a = ws.cell(row=grand_row, column=1)
    grand_a.value = "全国贸合计"
    grand_a.fill = GRAND_FILL
    _apply_style(grand_a, 1, ref_styles)
    try:
        ws.merge_cells(start_row=grand_row, start_column=1,
                       end_row=grand_row, end_column=2)
    except Exception:
        pass

    for col_letter in "CDEFGHIJKLMNOPQRST":
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=grand_row, column=col_idx)
        cell.fill = GRAND_FILL

        # 应用模板样式（含边框）+ 粗体
        _apply_style(cell, col_idx, ref_styles)
        bold_font = cell.font.copy() if cell.font else openpyxl.styles.Font()
        bold_font.bold = True
        cell.font = bold_font

        if col_letter in ratio_cols:
            fc_col, ac_col = _ratio_pair(col_letter)
            fc_val = grand_totals.get(fc_col, 0)
            ac_val = grand_totals.get(ac_col, 0)
            cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
            cell.number_format = FMT_RATIO
        else:
            # 直接用累计值写入（不再写公式，兼容手机端）
            cell.value = grand_totals.get(col_letter, 0)

    # ---- 8. 日志：预算JSON中有但映射表中无的模块 ----
    extra_in_forecast = set(forecast.keys()) - all_filled_modules
    if extra_in_forecast:
        print(f"  [信息] 预算预测中有但未输出的模块 ({len(extra_in_forecast)}):")
        for m in sorted(extra_in_forecast):
            print(f"    - {m}")

    # ---- 9. 保存 ----
    wb.save(output_path)
    wb.close()
    normal_count = sum(
        len(region_modules.get(r, [])) for r in REGION_ORDER
    )
    print(f"[填充] 普通模块: {normal_count}, 商贸配件: {len(TRADE_PARTS_MODULES)}, "
          f"子合计: {len(subtotal_rows)}")


def find_budget_file(directory):
    """在目录中找预算表Excel（必须同时包含所有关键词）"""
    files = [f for f in os.listdir(directory)
             if (f.endswith(".xlsx") or f.endswith(".xls")) and not f.startswith("~$")]
    for f in files:
        if all(kw in f for kw in BUDGET_KW):
            return os.path.join(directory, f)
    return None


def find_file_all(directory, keywords):
    """在目录中找包含所有关键词的xlsx文件（与 generate_report.find_file 不同，要求全部匹配）

    因为预算表含"市场"关键词，会干扰 generate_report.find_file 的"任意关键词匹配"逻辑，
    所以映射文件搜索需要此严格匹配。
    """
    files = [f for f in os.listdir(directory)
             if f.endswith(".xlsx") and not f.startswith("~$")]
    for f in files:
        if all(kw in f for kw in keywords):
            return os.path.join(directory, f)
    return None


# ============================================================
# 预算表 → JSON 抽取
# ============================================================
def extract_budget_to_json(budget_path, output_path):
    """一次性抽取：读取预算表"明细"sheet → JSON

    预算表"明细"sheet 结构（固定列位置）：
      col3  = 模块名
      col7  = 新签梯量 (6月)
      col14 = 新签金额 (6月)    ← 已是万元
      col21 = 排产台 (6月)
      col28 = 排产额 (6月)      ← 已是万元
      col35 = 发货台 (6月)
      col42 = 发货额 (6月)      ← 已是万元
    跳过模块名为空的小计行。
    """
    wb = openpyxl.load_workbook(budget_path, data_only=True)
    ws = None

    for name in wb.sheetnames:
        if "明细" in name:
            ws = wb[name]
            break

    if ws is None:
        print("[错误] 未找到'明细'sheet")
        wb.close()
        return False

    print(f"[抽取] Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

    forecast = {}
    skipped = 0

    for r in range(5, ws.max_row + 1):
        mod_name = ws.cell(row=r, column=3).value
        if mod_name is None:
            skipped += 1
            continue
        mod_name = str(mod_name).strip()
        if not mod_name:
            skipped += 1
            continue

        forecast[mod_name] = {
            "sign_units":  safe_float(ws.cell(row=r, column=7).value),
            "sign_amount": safe_float(ws.cell(row=r, column=14).value),
            "prod_units":  safe_float(ws.cell(row=r, column=21).value),
            "prod_amount": safe_float(ws.cell(row=r, column=28).value),
            "ship_units":  safe_float(ws.cell(row=r, column=35).value),
            "ship_amount": safe_float(ws.cell(row=r, column=42).value),
        }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(forecast, f, ensure_ascii=False, indent=2)

    print(f"[抽取] 模块: {len(forecast)}, 跳过(小计行): {skipped}")
    print(f"[抽取] → {output_path}")

    # 简要预览
    for mod_name, data in forecast.items():
        if any(v != 0 for v in data.values()):
            print(f"  {mod_name}: "
                  f"签单{data['sign_amount']:.0f}万/{data['sign_units']:.0f}台, "
                  f"排产{data['prod_amount']:.0f}万/{data['prod_units']:.0f}台, "
                  f"发货{data['ship_amount']:.0f}万/{data['ship_units']:.0f}台")

    wb.close()
    return True


# ============================================================
# JSON → forecast dict
# ============================================================
def build_budget_forecast(json_path):
    """从JSON读取6月预测数据

    返回格式与 generate_report.build_forecast_data() 完全一致：
      { 模块名: {sign_units, sign_amount, prod_units, prod_amount, ship_units, ship_amount} }
    确保 fill_template() 可以直接消费。
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    forecast = {mod_name: dict(vals) for mod_name, vals in data.items()}
    print(f"[预算预测] 加载 {len(forecast)} 个模块 ← {os.path.basename(json_path)}")
    return forecast


def generate_june_report(target_month, work_dir, source_dir):
    """生成6月报表（核心逻辑，供 fill_trade_parts_june.py 调用）

    返回: (report_path, unmatched_set)
    """
    json_path = os.path.join(work_dir, BUDGET_JSON)

    # 检查JSON
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"未找到预算预测JSON: {json_path}\n"
                                f"  请先运行: py -3.11 generate_report_june.py --extract")

    # 匹配数据源文件
    # 注意：映射文件用 find_file_all，避免预算表（含"市场"）干扰 find_file 的任意关键词匹配
    ledger_path = find_file(source_dir, LEDGER_KW)
    mapping_path = find_file_all(source_dir, MAPPING_KW)
    template_path = os.path.join(source_dir, TEMPLATE_NAME)

    for name, path in [("台账", ledger_path), ("映射", mapping_path), ("模板", template_path)]:
        if not path or not os.path.exists(path):
            raise FileNotFoundError(f"未找到{name}文件: {path}")

    today_str = datetime.date.today().strftime("%m-%d")
    report_path = os.path.join(work_dir, f"签单排产发货_{today_str}.xlsx")

    print(f"[文件] 台账: {os.path.basename(ledger_path)}")
    print(f"[文件] 映射: {os.path.basename(mapping_path)}")
    print(f"[文件] 模板: {os.path.basename(template_path)}")
    print(f"[文件] 预算预测: {os.path.basename(json_path)}")
    print(f"[文件] 输出: {os.path.basename(report_path)}\n")

    # Step 1: 映射表
    print("[Step 1] 构建国家→模块映射...")
    wb_map = openpyxl.load_workbook(mapping_path, data_only=True)
    wb_ledger = None
    try:
        country_to_module, country_to_region, region_modules = build_mapping(wb_map)

        # Step 2: 预算预测
        print("\n[Step 2] 加载预算预测数据...")
        forecast = build_budget_forecast(json_path)

        # Step 3: 台账统计
        print("\n[Step 3] 处理台账数据...")
        wb_ledger = openpyxl.load_workbook(ledger_path, data_only=True)
        stats, unmatched = process_ledger(wb_ledger, country_to_module, target_month)

        # Step 4: 填充模板
        print("\n[Step 4] 填充模板...")
        fill_template_dynamic(template_path, forecast, stats, region_modules,
                              target_month, report_path)
        print(f"  [完成] {os.path.basename(report_path)}")

        return report_path, unmatched
    finally:
        wb_map.close()
        if wb_ledger is not None:
            wb_ledger.close()


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="6月台账统计报表生成")
    parser.add_argument("--month", default=None, help="统计月份, 如 2026-06")
    parser.add_argument("--extract", action="store_true",
                        help="一次性从预算表抽取6月预测数据到JSON")
    parser.add_argument("--dir", default=os.getcwd(), help="输出目录")
    parser.add_argument("--source", default="数据源excel", help="数据源子目录（相对于--dir）")
    args = parser.parse_args()

    work_dir = args.dir
    source_dir = os.path.join(work_dir, args.source)
    json_path = os.path.join(work_dir, BUDGET_JSON)

    # ================================================================
    # --extract 模式：一次性抽取预算数据 → JSON
    # ================================================================
    if args.extract:
        print(f"{'=' * 60}")
        print(f"  预算表预测数据抽取 → {BUDGET_JSON}")
        print(f"{'=' * 60}\n")

        budget_path = find_budget_file(source_dir)
        if not budget_path or not os.path.exists(budget_path):
            print("[错误] 未找到预算表文件")
            print(f"  关键词: {BUDGET_KW}")
            print(f"  搜索目录: {source_dir}")
            sys.exit(1)

        print(f"[文件] 预算表: {os.path.basename(budget_path)}")
        ok = extract_budget_to_json(budget_path, json_path)
        if ok:
            print(f"\n[完成] JSON已生成，后续可直接用 --month 2026-06 生成报表")
        sys.exit(0 if ok else 1)

    # ================================================================
    # 正常报表生成模式
    # ================================================================
    if args.month:
        target_month = args.month
    else:
        today = datetime.date.today()
        target_month = f"{today.year}-{today.month:02d}"

    print(f"{'=' * 60}")
    print(f"  6月台账统计报表生成")
    print(f"  月份: {target_month}")
    print(f"{'=' * 60}\n")

    # 核心：生成报表
    try:
        report_path, unmatched = generate_june_report(target_month, work_dir, source_dir)
    except FileNotFoundError as e:
        print(f"[错误] {e}")
        sys.exit(1)

    # 未匹配国家
    if unmatched:
        unmatched_path = os.path.join(work_dir, f"未匹配国家_{target_month}.txt")
        with open(unmatched_path, "w", encoding="utf-8") as f:
            for c in sorted(unmatched):
                f.write(f"{c}\n")
        print(f"\n[信息] 未匹配国家: {unmatched_path}")

    # Excel 重算
    print()
    recalc_with_excel(report_path)

    print(f"\n{'=' * 60}")
    print(f"  完成! {os.path.basename(report_path)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
