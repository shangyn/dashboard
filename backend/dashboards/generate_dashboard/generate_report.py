#!py -3.11
"""
generate_report.py - 台账统计报表生成 (模板填充版)
============================================================
用法: py -3.11 generate_report.py --month 2026-05

读取"签单排产发货_模板.xlsx"作为输出模板，
将台账/映射/同期对比数据填充进去，生成最终报表。

Python 3.11 + openpyxl
============================================================
"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import argparse
import os
import sys
import re
import datetime
from collections import defaultdict, OrderedDict

# ============================================================
# 配置 (可按需修改)
# ============================================================
# ---- 文件匹配关键词 ----
LEDGER_KW = ["台账", "合同标的"]
MAPPING_KW = ["国家", "市场", "业务员"]
FORECAST_KW = ["同期对比", "预计签排发", "签排发"]
TEMPLATE_NAME = "签单排产发货_模板.xlsx"

# ---- Sheet名 ----
MAPPING_SHEET = "Sheet3"
FORECAST_SHEET = "上半年单月明细"

# ---- 台账排除规则 ----
EXCLUDE_PRODUCT_STATUS = "已作废"
EXCLUDE_ELEVATOR_TYPE = "改造梯"

# ---- 商贸配件识别 ----
TRADE_PARTS_KW = ["商贸", "配件", "重锅", "改造"]

# ---- 数字格式 ----
FMT_UNITS = "0"                          # 台数: 0→0, 5→5
FMT_AMOUNT = "#,##0.##;-#,##0.##;0"      # 合同额: 0→0, 5230→5,230, 5.5→5.5
FMT_RATIO = "0.00%;-0.00%;0"             # 完成比: 0→0, 0.05→5.00%


# ============================================================
# 工具函数
# ============================================================
def find_file(directory, keywords):
    """在目录中找包含任意关键词的xlsx文件（排除临时文件）"""
    files = [f for f in os.listdir(directory)
             if f.endswith(".xlsx") and not f.startswith("~$")]
    for f in files:
        for kw in keywords:
            if kw in f:
                return os.path.join(directory, f)
    return None


def find_columns_by_header(ws, header_map, search_rows=5):
    """按表头关键词搜索列位置，返回 {别名: 列号(1-based)}"""
    candidates = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=search_rows, values_only=True), start=1):
        for c, cell_val in enumerate(row, start=1):
            val = str(cell_val or "").strip()
            if not val:
                continue
            for alias, keywords in header_map.items():
                for kw in keywords:
                    if kw in val:
                        is_exact = (val == kw)
                        score = (is_exact, len(kw))
                        candidates.setdefault(alias, []).append((score, r, c))
                        break
    best = {}
    for alias in header_map:
        if alias not in candidates:
            continue
        top = max(candidates[alias], key=lambda x: (x[0][0], x[0][1], -x[1], -x[2]))
        best[alias] = top[2]
    # 去重：同列冲突选高分
    col_assign = {}
    for alias, col in best.items():
        if col not in col_assign or candidates[alias][0] > candidates[col_assign[col]][0]:
            col_assign[col] = alias
    return {alias: col for col, alias in col_assign.items()}


def safe_float(v):
    """安全转换为float，处理 #DIV/0! 等错误值"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ============================================================
# 数据处理
# ============================================================
def build_mapping(wb):
    """从映射文件读取 国家→模块→大区"""
    ws = wb[MAPPING_SHEET]
    print(f"[映射] Sheet: {MAPPING_SHEET}")

    header_map = {
        "country": ["国家"],
        "module": ["2026年对应模块", "对应模块"],
        "region": ["2026九大区"],
    }
    cols = find_columns_by_header(ws, header_map)
    print(f"[映射] 列定位: country={cols.get('country')}, module={cols.get('module')}, region={cols.get('region')}")

    country_to_module = {}
    country_to_region = {}
    region_modules = OrderedDict()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        country = str(row[cols["country"] - 1]).strip() if row[cols["country"] - 1] else ""
        module = str(row[cols["module"] - 1]).strip() if row[cols["module"] - 1] else ""
        region = str(row[cols["region"] - 1]).strip() if row[cols["region"] - 1] else ""

        if not country or not module or not region:
            continue
        if region in ("#N/A", "0", "None", ""):
            continue
        if module in ("#N/A", "0", "None", ""):
            continue

        country_to_module[country] = module
        country_to_region[country] = region
        region_modules.setdefault(region, [])
        if module not in region_modules[region]:
            region_modules[region].append(module)

    print(f"[映射] 有效国家: {len(country_to_module)}, 大区: {len(region_modules)}")
    return country_to_module, country_to_region, region_modules


def build_forecast_data(wb, target_month):
    """从同期对比文件提取各模块预测值"""
    month_num = int(target_month.split("-")[1])
    chinese = ["", "一", "二", "三", "四", "五", "六", "七", "八", "九", "十", "十一", "十二"]
    month_cn = chinese[month_num]

    ws = wb[FORECAST_SHEET]
    print(f"[同期对比] Sheet: {FORECAST_SHEET}, 目标: {month_num}月")

    header_map = {
        "seq": ["序号"],
        "module": ["模块"],
        "sign_fc": [f"预计{month_num}月签单", f"预计{month_cn}月签单"],
        "prod_fc": [f"预计{month_num}月排产", f"预计{month_cn}月排产"],
        "ship_fc": [f"预计{month_num}月发货", f"预计{month_cn}月发货"],
    }
    cols = find_columns_by_header(ws, header_map, search_rows=8)
    print(f"[同期对比] 列定位: sign={cols.get('sign_fc')}, prod={cols.get('prod_fc')}, ship={cols.get('ship_fc')}")

    forecast = {}
    trade_parts = []
    in_trade = False

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        mod_val = row[cols["module"] - 1] if "module" in cols else None
        if not mod_val or "合计" in str(mod_val):
            continue
        mod_name = str(mod_val).strip()

        sc = cols.get("sign_fc", 0)
        pc = cols.get("prod_fc", 0)
        shc = cols.get("ship_fc", 0)

        forecast[mod_name] = {
            "sign_units": safe_float(row[sc - 1]) if sc else 0,
            "sign_amount": safe_float(row[sc + 4]) if sc and len(row) > sc + 4 else 0,
            "prod_units": safe_float(row[pc - 1]) if pc else 0,
            "prod_amount": safe_float(row[pc + 4]) if pc and len(row) > pc + 4 else 0,
            "ship_units": safe_float(row[shc - 1]) if shc else 0,
            "ship_amount": safe_float(row[shc + 4]) if shc and len(row) > shc + 4 else 0,
        }

        if any(kw in mod_name for kw in TRADE_PARTS_KW):
            in_trade = True
        if in_trade:
            trade_parts.append(mod_name)

    print(f"[同期对比] 模块: {len(forecast)}, 商贸配件: {len(trade_parts)}")
    return forecast


def process_ledger(wb, country_to_module, target_month):
    """处理台账数据"""
    ws = wb.active
    print(f"[台账] Sheet: {ws.title}, 行数: {ws.max_row}")

    header_map = {
        "country": ["国家"],
        "units": ["台数"],
        "contract_amt": ["合同额（人民币）", "合同额"],
        "elevator_type": ["产品型号"],
        "product_status": ["产品状态"],
        "sign_date": ["签订日期"],
        "prod_date": ["排产日期"],
        "ship_date": ["组A日期"],
    }
    cols = find_columns_by_header(ws, header_map, search_rows=3)
    print(f"[台账] 列定位: sign_date={cols.get('sign_date')}, prod_date={cols.get('prod_date')}, ship_date={cols.get('ship_date')}")

    stats = defaultdict(lambda: {
        "sign_units": 0.0, "sign_amount": 0.0,
        "prod_units": 0.0, "prod_amount": 0.0,
        "ship_units": 0.0, "ship_amount": 0.0,
    })
    unmatched = set()
    excl_ba = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ps = str(row[cols["product_status"] - 1]).strip() if row[cols["product_status"] - 1] else ""
        et = str(row[cols["elevator_type"] - 1]).strip() if row[cols["elevator_type"] - 1] else ""

        if ps == EXCLUDE_PRODUCT_STATUS:
            excl_ba += 1
            continue

        country = str(row[cols["country"] - 1]).strip() if row[cols["country"] - 1] else ""
        units = safe_float(row[cols["units"] - 1])
        contract_amt = safe_float(row[cols["contract_amt"] - 1]) / 10000  # 元→万元

        sign_date = str(row[cols["sign_date"] - 1]).strip() if row[cols["sign_date"] - 1] else ""
        prod_date = str(row[cols["prod_date"] - 1]).strip() if row[cols["prod_date"] - 1] else ""
        ship_date = str(row[cols["ship_date"] - 1]).strip() if row[cols["ship_date"] - 1] else ""

        # M列含"改造" → 强制归属"改造"模块
        if "改造" in et:
            module = "改造"
        else:
            module = country_to_module.get(country)
            if not module:
                unmatched.add(country)
                continue

        if sign_date.startswith(target_month):
            stats[module]["sign_units"] += units
            stats[module]["sign_amount"] += contract_amt
        if prod_date.startswith(target_month):
            stats[module]["prod_units"] += units
            stats[module]["prod_amount"] += contract_amt
        if ship_date.startswith(target_month):
            stats[module]["ship_units"] += units
            stats[module]["ship_amount"] += contract_amt

    print(f"[台账] 排除(已作废:{excl_ba}), 未匹配:{len(unmatched)}")
    return stats, unmatched


# ============================================================
# 模板填充
# ============================================================
def fill_template(template_path, forecast, stats, target_month, output_path):
    """填充模板并保存"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    month_num = int(target_month.split("-")[1])
    chinese = ["", "一", "二", "三", "四", "五", "六", "七", "八", "九", "十", "十一", "十二"]
    month_cn = chinese[month_num]

    # 1. 检测模板原始月份并替换
    title_text = str(ws.cell(row=1, column=1).value or "")
    tm = re.search(r"(\d+)月", title_text)
    template_month = int(tm.group(1)) if tm else month_num

    if template_month != month_num:
        for rn in [1, 3, 4]:
            for c in range(1, 21):
                cell = ws.cell(row=rn, column=c)
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace(f"{template_month}月", f"{month_num}月")

    # 2. 更新日期
    ws.cell(row=2, column=1).value = f"报表日期：{datetime.date.today()}"

    # 3. 完成比公式模板
    def ratio_f(fc, ac, r):
        return f"=IF({fc}{r}=0,IF({ac}{r}=0,0,1),{ac}{r}/{fc}{r})"

    # 4. 逐行填充
    subtotal_rows = []
    grand_row = None
    filled = 0
    trade_filled = 0

    for row_num in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row=row_num, column=1).value or "").strip()
        b_val = str(ws.cell(row=row_num, column=2).value or "").strip()

        # 空行跳过
        if not a_val and not b_val:
            continue

        # 全国贸合计 → 记录位置
        if "全国贸" in a_val:
            grand_row = row_num
            continue

        # 子合计行 → 记录位置，保留公式
        if "合计" in a_val:
            subtotal_rows.append(row_num)
            continue

        # 判断商贸配件
        is_trade = any(kw in b_val for kw in TRADE_PARTS_KW)
        mod_name = b_val
        fc = forecast.get(mod_name, {})
        st = stats.get(mod_name, {})

        if is_trade:
            # 商贸配件列规则
            for col_letter, rule in _trade_col_rules().items():
                col_idx = column_index_from_string(col_letter)
                cell = ws.cell(row=row_num, column=col_idx)
                if rule == "dash":
                    cell.value = "-"
                    cell.number_format = "@"
                elif rule == "zero":
                    if mod_name == "改造":
                        field = _col_to_field(col_letter)
                        cell.value = st.get(field, 0) if field else 0
                    else:
                        cell.value = 0
                elif rule == "import":
                    field = _col_to_field(col_letter)
                    cell.value = fc.get(field, 0) if field else 0
                elif rule == "formula":
                    fc_col, ac_col = _ratio_pair(col_letter)
                    cell.value = ratio_f(fc_col, ac_col, row_num)
                    cell.number_format = FMT_RATIO
            trade_filled += 1
        else:
            # 普通模块：按列填充
            for col_letter, (col_type, group) in _normal_col_rules().items():
                col_idx = column_index_from_string(col_letter)
                cell = ws.cell(row=row_num, column=col_idx)

                if col_type == "forecast":
                    field = f"{group}_{'amount' if col_letter in _AMOUNT_COLS else 'units'}"
                    cell.value = fc.get(field, 0)
                elif col_type == "actual":
                    field = f"{group}_{'amount' if col_letter in _AMOUNT_COLS else 'units'}"
                    cell.value = st.get(field, 0)
                elif col_type == "ratio":
                    fc_col, ac_col = _ratio_pair(col_letter)
                    cell.value = ratio_f(fc_col, ac_col, row_num)
                    cell.number_format = FMT_RATIO
            filled += 1

    # 5. 重写全国贸合计公式
    if grand_row and subtotal_rows:
        # 合并A+B列
        try:
            ws.unmerge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=2)
        except Exception:
            pass
        ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=2)
        ws.cell(row=grand_row, column=1).value = "全国贸合计"

        CL = get_column_letter
        for ratio_col in ["E", "H", "K", "N", "Q", "T"]:
            fc_col, ac_col = _ratio_pair(ratio_col)
            ratio_idx = column_index_from_string(ratio_col)
            fc_idx = column_index_from_string(fc_col)
            ac_idx = column_index_from_string(ac_col)

            for vc_idx in (fc_idx, ac_idx):
                parts = [f"{CL(vc_idx)}{sr}" for sr in subtotal_rows]
                cell = ws.cell(row=grand_row, column=vc_idx)
                cell.value = "=" + "+".join(parts)

            cell = ws.cell(row=grand_row, column=ratio_idx)
            cell.value = ratio_f(fc_col, ac_col, grand_row)
            cell.number_format = FMT_RATIO

    wb.save(output_path)
    wb.close()
    print(f"[填充] 普通模块: {filled}, 商贸配件: {trade_filled}, 子合计: {len(subtotal_rows)}")


# ---- 列映射辅助 ----
# 金额列（使用金额格式）
_AMOUNT_COLS = {"F", "G", "L", "M", "R", "S"}

# 列 → 数据分组
def _col_group(col):
    if col in "CDEFGH":
        return "sign"
    if col in "IJKLMN":
        return "prod"
    return "ship"


def _normal_col_rules():
    """普通模块：每列的数据来源类型"""
    rules = {}
    for col in "CDEFGHIJKLMNOPQRST":
        if col in "CDFGIJLMOPRS":  # 非完成比列
            if col in "CFILOR":    # 预测列
                rules[col] = ("forecast", _col_group(col))
            else:                  # 实际列 DGJMPS
                rules[col] = ("actual", _col_group(col))
        else:                      # 完成比列 EHK NQT
            rules[col] = ("ratio", _col_group(col))
    return rules


def _ratio_pair(col):
    """完成比列 → (预测列, 实际列)"""
    pairs = {
        "E": ("C", "D"), "H": ("F", "G"),
        "K": ("I", "J"), "N": ("L", "M"),
        "Q": ("O", "P"), "T": ("R", "S"),
    }
    return pairs.get(col, ("C", "D"))


def _trade_col_rules():
    """商贸配件：每列的填充规则 (dash/zero/import/formula)"""
    return {
        "C": "dash", "D": "dash", "E": "dash",
        "F": "import",
        "G": "zero", "H": "formula",
        "I": "dash", "J": "dash", "K": "dash",
        "L": "import",
        "M": "zero", "N": "formula",
        "O": "dash", "P": "dash", "Q": "dash",
        "R": "import",
        "S": "zero", "T": "formula",
    }


def _col_to_field(col):
    """列 → forecast/stats 数据字段名"""
    mapping = {
        "F": "sign_amount", "L": "prod_amount", "R": "ship_amount",
        "G": "sign_amount", "M": "prod_amount", "S": "ship_amount",
    }
    return mapping.get(col)


# ============================================================
# Excel COM 重算（确保公式缓存值正确，兼容手机端）
# ============================================================
def recalc_with_excel(filepath):
    """用Excel打开文件重算并保存，写入公式缓存值"""
    try:
        import win32com.client
    except ImportError:
        print("[警告] 未安装 pywin32，跳过Excel重算。安装: pip install pywin32")
        return

    abs_path = os.path.abspath(filepath)
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        wb = excel.Workbooks.Open(abs_path)
        wb.Save()
        wb.Close()
        wb = None
        print(f"[重算] Excel公式缓存已更新")
    except Exception as e:
        print(f"[警告] Excel重算失败: {e}")
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="台账统计报表生成 (模板填充版)")
    parser.add_argument("--month", default=None, help="统计月份, 如 2026-05")
    parser.add_argument("--dir", default=os.getcwd(), help="文件所在目录")
    parser.add_argument("--source", default="数据源excel", help="数据源Excel子目录（相对于--dir）")
    parser.add_argument("--ledger", default=None, help="台账Excel路径")
    parser.add_argument("--mapping", default=None, help="映射Excel路径")
    parser.add_argument("--forecast", default=None, help="同期对比Excel路径")
    parser.add_argument("--template", default=None, help="输出模板Excel路径")
    parser.add_argument("--output", default=None, help="输出文件路径")
    args = parser.parse_args()

    work_dir = args.dir
    source_dir = os.path.join(work_dir, args.source)

    # 月份
    if args.month:
        target_month = args.month
    else:
        today = datetime.date.today()
        target_month = f"{today.year}-{today.month:02d}"

    print(f"{'=' * 60}")
    print(f"  台账统计报表生成")
    print(f"  月份: {target_month}")
    print(f"{'=' * 60}\n")

    # 匹配文件
    ledger_path = args.ledger or find_file(source_dir, LEDGER_KW)
    mapping_path = args.mapping or find_file(source_dir, MAPPING_KW)
    forecast_path = args.forecast or find_file(source_dir, FORECAST_KW)
    template_path = args.template or os.path.join(source_dir, TEMPLATE_NAME)

    # 输出文件名
    today_str = datetime.date.today().strftime("%m-%d")
    output_path = args.output or os.path.join(work_dir, f"签单排产发货_{today_str}.xlsx")

    # 检查文件
    for name, path in [("台账", ledger_path), ("映射", mapping_path),
                        ("同期对比", forecast_path), ("模板", template_path)]:
        if not path or not os.path.exists(path):
            print(f"[错误] 未找到{name}文件: {path}")
            sys.exit(1)

    print(f"[文件] 台账: {os.path.basename(ledger_path)}")
    print(f"[文件] 映射: {os.path.basename(mapping_path)}")
    print(f"[文件] 同期对比: {os.path.basename(forecast_path)}")
    print(f"[文件] 模板: {os.path.basename(template_path)}")
    print(f"[文件] 输出: {os.path.basename(output_path)}\n")

    # 加载数据文件 (data_only=True 读取计算值)
    wb_map = openpyxl.load_workbook(mapping_path, data_only=True)
    wb_fc = openpyxl.load_workbook(forecast_path, data_only=True)
    wb_ledger = openpyxl.load_workbook(ledger_path, data_only=True)

    try:
        # 1. 映射
        country_to_module, _, _ = build_mapping(wb_map)

        # 2. 同期对比
        forecast = build_forecast_data(wb_fc, target_month)

        # 3. 台账
        stats, unmatched = process_ledger(wb_ledger, country_to_module, target_month)

        # 4. 填充
        fill_template(template_path, forecast, stats, target_month, output_path)
    finally:
        for wb in [wb_map, wb_fc, wb_ledger]:
            try:
                wb.close()
            except Exception:
                pass

    # 5. 未匹配国家
    if unmatched:
        unmatched_path = os.path.join(work_dir, f"未匹配国家_{target_month}.txt")
        with open(unmatched_path, "w", encoding="utf-8") as f:
            for c in sorted(unmatched):
                f.write(f"{c}\n")
        print(f"[信息] 未匹配国家: {unmatched_path}")

    # 6. Excel重算（确保手机端公式缓存值正确）
    recalc_with_excel(output_path)

    print(f"\n{'=' * 60}")
    print(f"  完成! {os.path.basename(output_path)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
