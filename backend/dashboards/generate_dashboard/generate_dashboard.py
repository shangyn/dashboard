#!py -3.11
"""
generate_dashboard.py - 签单排产发货看板生成
============================================================
用法: py -3.11 generate_dashboard.py --month 2026-05

读取已生成的"签单排产发货_MM-DD.xlsx"，
提取9个子合计 + 全国贸合计数据，
生成独立HTML看板（概览卡片 + 6个双环图表）。
============================================================
"""
import openpyxl
from openpyxl.utils import column_index_from_string
import argparse
import base64
import os
import sys
import json
import datetime
import socket
import http.server
import socketserver
import threading
import time


# ============================================================
# 配置
# ============================================================
OUTPUT_TEMPLATE = "签单排产发货_{date}.xlsx"

# 6个指标定义
METRICS = [
    {"id": "sign_units",   "name": "签单台数", "unit": "台",   "fc_col": "C", "ac_col": "D", "ratio_col": "E"},
    {"id": "sign_amount",  "name": "签单金额", "unit": "万元", "fc_col": "F", "ac_col": "G", "ratio_col": "H"},
    {"id": "prod_units",   "name": "排产台数", "unit": "台",   "fc_col": "I", "ac_col": "J", "ratio_col": "K"},
    {"id": "prod_amount",  "name": "排产金额", "unit": "万元", "fc_col": "L", "ac_col": "M", "ratio_col": "N"},
    {"id": "ship_units",   "name": "发货台数", "unit": "台",   "fc_col": "O", "ac_col": "P", "ratio_col": "Q"},
    {"id": "ship_amount",  "name": "发货金额", "unit": "万元", "fc_col": "R", "ac_col": "S", "ratio_col": "T"},
]

# 9区域颜色
REGION_COLORS = [
    "#4C78A8", "#54A867", "#C8963C", "#D4645C", "#5AA3AE",
    "#7BA868", "#E07B42", "#8C6DB8", "#C86C8A",
]

# 内环颜色
INNER_COLORS = {
    "sign_units":  "#3b82f6",
    "sign_amount": "#3b82f6",
    "prod_units":  "#3b82f6",
    "prod_amount": "#3b82f6",
    "ship_units":  "#3b82f6",
    "ship_amount": "#3b82f6",
}


# ============================================================
# 数据提取
# ============================================================
def safe_num(v):
    """安全转为数值，处理 '-' 文本"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("-", "", "#DIV/0!", "#N/A", "#REF!"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_dashboard_data(excel_path, workbook=None):
    """从输出Excel提取看板数据"""
    if workbook is not None:
        wb = workbook
    else:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    print(f"[数据] 读取: {os.path.basename(excel_path)}, Sheet: {ws.title}")

    subtotal_rows = []
    grand_row = None

    for r in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row=r, column=1).value or "").strip()
        if "全国贸" in a_val:
            grand_row = r
        elif "合计" in a_val:
            subtotal_rows.append(r)

    region_data = []
    prev_subtotal = 4

    for sr in subtotal_rows:
        region_name = str(ws.cell(row=sr, column=1).value or "").strip()
        module_values = {m["id"]: {"fc": 0.0, "ac": 0.0} for m in METRICS}
        has_modules = False
        for mr in range(prev_subtotal + 1, sr):
            a_val = str(ws.cell(row=mr, column=1).value or "").strip()
            b_val = str(ws.cell(row=mr, column=2).value or "").strip()
            if not a_val or not b_val:
                continue
            has_modules = True
            for m in METRICS:
                fc_val = safe_num(ws.cell(row=mr, column=column_index_from_string(m["fc_col"])).value)
                ac_val = safe_num(ws.cell(row=mr, column=column_index_from_string(m["ac_col"])).value)
                module_values[m["id"]]["fc"] += fc_val
                module_values[m["id"]]["ac"] += ac_val

        if has_modules:
            entry = {"name": region_name}
            for m in METRICS:
                fc = module_values[m["id"]]["fc"]
                ac = module_values[m["id"]]["ac"]
                ratio = ac / fc if fc != 0 else (0 if ac == 0 else 1)
                entry[m["id"]] = {
                    "fc": round(fc, 0),
                    "ac": round(ac, 0),
                    "ratio": round(ratio, 4),
                }
            region_data.append(entry)

        prev_subtotal = sr

    # 按完成比高低交错排列，避免同侧堆积
    if len(region_data) > 1:
        avg_ratios = []
        for i, r in enumerate(region_data):
            avg_r = sum(r[m["id"]]["ratio"] for m in METRICS) / len(METRICS)
            avg_ratios.append((avg_r, r))
        avg_ratios.sort(key=lambda x: x[0], reverse=True)
        n = len(avg_ratios)
        half = (n + 1) // 2
        interleaved = []
        for i in range(half):
            interleaved.append(avg_ratios[i][1])
            if i + half < n:
                interleaved.append(avg_ratios[i + half][1])
        region_data = interleaved

    # 全国贸合计
    grand = {"name": "全国贸合计"}
    for m in METRICS:
        total_fc = sum(r[m["id"]]["fc"] for r in region_data)
        total_ac = sum(r[m["id"]]["ac"] for r in region_data)
        ratio = total_ac / total_fc if total_fc != 0 else (0 if total_ac == 0 else 1)
        grand[m["id"]] = {
            "fc": round(total_fc, 0),
            "ac": round(total_ac, 0),
            "ratio": round(ratio, 4),
        }

    print(f"[数据] 区域: {len(region_data)}, 全国贸合计: 已计算")
    return region_data, grand


# ============================================================
# 表格数据提取
# ============================================================
def read_table_rows(excel_path, workbook=None):
    """读取Excel全部数据行，返回列表供表格展示"""
    if workbook is not None:
        wb = workbook
    else:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    def rv(row, col_letter):
        """读取单元格原始值"""
        v = ws.cell(row=row, column=column_index_from_string(col_letter)).value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return v
        s = str(v).strip()
        if s in ("", "-"):
            return s
        try:
            return float(s)
        except ValueError:
            return s

    def ratio_val(fc, ac):
        """完成比公式"""
        if fc is None or fc == 0:
            return 0 if (ac is None or ac == 0) else 1
        return (ac or 0) / fc

    ratio_pairs = [("C","D","E"), ("F","G","H"), ("I","J","K"), ("L","M","N"), ("O","P","Q"), ("R","S","T")]

    all_rows = []
    subtotal_positions = []
    idx = 0
    for r in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row=r, column=1).value or "").strip()
        b_val = str(ws.cell(row=r, column=2).value or "").strip()
        if not a_val and not b_val:
            continue

        is_grand = "全国贸" in a_val
        is_subtotal = "合计" in a_val and not is_grand

        cols = {}
        for col_letter in "CDEFGHIJKLMNOPQRST":
            cols[col_letter] = rv(r, col_letter)

        # 计算模块行的完成比
        if not is_subtotal and not is_grand:
            for fc_col, ac_col, ratio_col in ratio_pairs:
                if cols[ratio_col] is None or cols[ratio_col] == "":
                    fc = cols[fc_col]
                    ac = cols[ac_col]
                    if fc is not None and fc != "" and fc != "-":
                        cols[ratio_col] = ratio_val(fc, ac)
                    if ac is not None and isinstance(ac, str) and ac == "-":
                        cols[ratio_col] = "-"

        all_rows.append({
            "a": a_val,
            "b": b_val,
            "isSubtotal": is_subtotal,
            "isGrand": is_grand,
            "cols": cols,
        })

        if is_subtotal or is_grand:
            subtotal_positions.append(len(all_rows) - 1)

    # 分离子合计和全国贸合计
    sub_positions = [p for p in subtotal_positions if all_rows[p]["isSubtotal"]]
    grand_positions = [p for p in subtotal_positions if all_rows[p]["isGrand"]]

    value_cols = ["C","D","F","G","I","J","L","M","O","P","R","S"]

    # 子合计 = 上方模块行求和
    prev = -1
    for st_pos in sub_positions:
        for col in value_cols:
            total = 0
            for i in range(prev + 1, st_pos):
                v = all_rows[i]["cols"].get(col)
                if isinstance(v, (int, float)):
                    total += v
            all_rows[st_pos]["cols"][col] = total
        for fc_col, ac_col, ratio_col in ratio_pairs:
            fc = all_rows[st_pos]["cols"][fc_col]
            ac = all_rows[st_pos]["cols"][ac_col]
            if isinstance(fc, (int, float)) and isinstance(ac, (int, float)):
                all_rows[st_pos]["cols"][ratio_col] = ratio_val(fc, ac)
        prev = st_pos

    # 全国贸合计 = 所有子合计之和
    for g_pos in grand_positions:
        for col in value_cols:
            total = 0
            for sp in sub_positions:
                v = all_rows[sp]["cols"].get(col)
                if isinstance(v, (int, float)):
                    total += v
            all_rows[g_pos]["cols"][col] = total
        for fc_col, ac_col, ratio_col in ratio_pairs:
            fc = all_rows[g_pos]["cols"][fc_col]
            ac = all_rows[g_pos]["cols"][ac_col]
            if isinstance(fc, (int, float)) and isinstance(ac, (int, float)):
                all_rows[g_pos]["cols"][ratio_col] = ratio_val(fc, ac)

    return all_rows


# ============================================================
# HTML生成
# ============================================================
def generate_html(region_data, grand, target_month, output_path, table_rows=None, excel_path=None):
    month_num = int(target_month.split("-")[1])
    title = f"{target_month[:4]}年{month_num}月签单排产发货看板"

    # 原始Excel文件 Base64 编码，供导出下载
    excel_b64 = ""
    excel_filename = ""
    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            excel_b64 = base64.b64encode(f.read()).decode("ascii")
        excel_filename = os.path.basename(excel_path)

    chart_data = json.dumps({
        "regions": region_data,
        "grand": grand,
        "metrics": METRICS,
        "regionColors": REGION_COLORS,
        "innerColors": INNER_COLORS,
        "tableRows": table_rows or [],
        "excelB64": excel_b64,
        "excelFilename": excel_filename,
    }, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<script src="echarts.min.js"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
    background: linear-gradient(160deg, #f0f2f5 0%, #e3e8ee 50%, #edf0f5 100%);
    min-height: 100vh;
    color: #1a2332;
}}
.container {{ width: 100%; max-width: 1920px; margin: 0 auto; padding: 36px 24px; }}

.hero {{ text-align: center; margin-bottom: 28px; }}
.hero h1 {{ font-size: 34px; font-weight: 800; letter-spacing: 3px; color: #0d1b2a; }}
.hero .date {{ font-size: 18px; color: #8899aa; margin-top: 8px; }}

.summary {{
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 14px;
    margin-bottom: 36px;
}}
.summary-card {{
    position: relative;
    background: #fff;
    border-radius: 16px;
    padding: 22px 10px 16px;
    text-align: center;
    box-shadow: 0 2px 16px rgba(0,0,0,0.05);
    transition: all 0.35s cubic-bezier(0.25, 0.46, 0.45, 0.94);
    cursor: default;
    overflow: hidden;
    min-width: 0;
}}
.summary-card::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 4px;
    border-radius: 16px 16px 0 0;
}}
.summary-card:nth-child(1)::before {{ background: #4C78A8; }}
.summary-card:nth-child(2)::before {{ background: #54A867; }}
.summary-card:nth-child(3)::before {{ background: #C8963C; }}
.summary-card:nth-child(4)::before {{ background: #5AA3AE; }}
.summary-card:nth-child(5)::before {{ background: #7BA868; }}
.summary-card:nth-child(6)::before {{ background: #8C6DB8; }}
.summary-card:hover {{
    transform: translateY(-4px) scale(1.02);
    box-shadow: 0 12px 32px rgba(0,0,0,0.10);
}}
.summary-card .sc-label {{
    font-size: 13px; color: #1a2332; margin-bottom: 6px;
    letter-spacing: 2px; font-weight: 600;
}}
.summary-card .sc-value {{
    font-size: 36px; font-weight: 800; color: #2563eb; line-height: 1.15;
}}
.summary-card .sc-unit {{
    font-size: 14px; font-weight: 500; color: #2563eb; margin-left: 1px;
}}
.summary-card .sc-ratio {{
    display: inline-block;
    margin-top: 8px;
    padding: 3px 12px;
    border-radius: 16px;
    font-size: 13px; font-weight: 700;
}}
.ratio-badge-high {{ background: #ecfdf5; color: #059669; }}
.ratio-badge-mid {{ background: #eff6ff; color: #2563eb; }}
.ratio-badge-low {{ background: #ecfdf5; color: #059669; }}

.section-title {{
    text-align: center;
    font-size: 20px;
    font-weight: 700;
    color: #0d1b2a;
    margin-bottom: 16px;
    letter-spacing: 2px;
}}

.region-legend {{
    display: flex;
    flex-wrap: wrap;
    justify-content: center;
    gap: 10px 22px;
    margin-bottom: 30px;
    padding: 14px 20px;
    background: #fff;
    border-radius: 12px;
    box-shadow: 0 1px 8px rgba(0,0,0,0.04);
}}
.legend-item {{
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 13px;
    color: #333;
    font-weight: 500;
}}
.legend-dot {{
    width: 10px; height: 10px;
    border-radius: 3px;
    flex-shrink: 0;
}}

.charts-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 22px;
    margin-bottom: 36px;
}}
.chart-card {{
    background: #fff;
    border-radius: 16px;
    padding: 18px 14px 14px;
    box-shadow: 0 2px 16px rgba(0,0,0,0.05);
    overflow: hidden;
    min-width: 0;
}}
.chart-title {{
    font-size: 16px;
    font-weight: 700;
    text-align: left;
    color: #0d1b2a;
    margin-bottom: 12px;
    padding-left: 6px;
    letter-spacing: 1px;
}}
.chart-wrap {{
    width: 100%;
    height: 540px;
    overflow: hidden;
}}

.toggle-bar {{
    display: flex;
    justify-content: center;
    gap: 12px;
    margin-bottom: 28px;
}}
.toggle-btn {{
    padding: 10px 36px;
    border: 2px solid #3b82f6;
    border-radius: 8px;
    background: #fff;
    color: #3b82f6;
    font-size: 15px;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.25s;
    letter-spacing: 2px;
}}
.toggle-btn.active {{
    background: #3b82f6;
    color: #fff;
}}
.toggle-btn:hover:not(.active) {{
    background: #eff6ff;
}}
.toggle-btn:focus {{ outline: none; }}

#view-charts.hide {{ display: none; }}
#view-table {{ display: none; }}
#view-table.show {{ display: block; }}

.table-wrap {{
    overflow-x: auto;
    background: #fff;
    border-radius: 12px;
    box-shadow: 0 2px 16px rgba(0,0,0,0.05);
    margin-bottom: 36px;
}}
.data-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
}}
.data-table thead th {{
    background: #3b82f6;
    color: #fff;
    padding: 5px 4px;
    font-weight: 600;
    font-size: 14px;
    white-space: nowrap;
    position: sticky;
    top: 0;
    z-index: 1;
    border: 1px solid #dbeafe;
}}
.data-table thead th.sub-header {{
    background: #60a5fa;
    padding: 3px 4px;
}}
.data-table tbody td {{
    padding: 3px 5px;
    border: 1px solid #e5e7eb;
    text-align: center;
    white-space: nowrap;
    font-size: 14px;
}}
.data-table .col-a {{ text-align: center; width: 36px; color: #9ca3af; }}
.data-table .col-b {{ text-align: left; padding-left: 6px; }}
.data-table .row-subtotal td {{ background: #D6E4F0; font-weight: 700; font-size: 16px; padding: 6px 5px; }}
.data-table .row-grand td {{ background: #B4C6E7; font-weight: 800; font-size: 16px; padding: 8px 5px; }}
.data-table .row-normal:nth-child(even) td {{ background: #f8fafc; }}
.data-table .row-normal:nth-child(even).row-subtotal td {{ background: #D6E4F0; }}
.data-table .row-region-header td {{
    background: #4c8ef9; color: #fff; font-weight: 600; font-size: 13px;
    padding: 4px 4px; border: 1px solid #dbeafe;
}}
.data-table .row-region-header.row-region-subheader td {{
    background: #4c8ef9; font-size: 12px; font-weight: 500;
}}

.ratio-low  {{ color: #ff00008c !important; font-weight: 600; }}
.ratio-mid  {{ color: #ff0000 !important; font-weight: 600; }}
.ratio-high {{ color: #ff0000 !important; font-weight: 600; }}

.footer {{ text-align: center; color: #aab5c0; font-size: 12px; margin-top: 12px; letter-spacing: 1px; }}

@media (max-width: 1100px) {{
    .summary {{ grid-template-columns: repeat(3, 1fr); }}
    .charts-grid {{ grid-template-columns: 1fr; }}
    .chart-wrap {{ height: 480px; }}
    .chart-card {{ padding: 14px 10px 10px; }}
    .container {{ padding: 20px 16px; }}
}}
@media (max-width: 640px) {{
    .summary {{ grid-template-columns: repeat(2, 1fr); }}
    .hero h1 {{ font-size: 22px; }}
    .summary-card .sc-value {{ font-size: 28px; }}
    .chart-wrap {{ height: 400px; }}
    .chart-card {{ padding: 10px 6px 6px; }}
    .chart-title {{ font-size: 14px; }}
}}
@media (max-width: 480px) {{
    .summary {{ grid-template-columns: 1fr; }}
    .hero h1 {{ font-size: 18px; }}
    .hero .subtitle {{ font-size: 11px; }}
    .summary-card .sc-value {{ font-size: 22px; }}
    .summary-card .sc-label {{ font-size: 11px; }}
    .chart-wrap {{ height: 280px; }}
    .chart-card {{ padding: 8px 4px 4px; }}
    .chart-title {{ font-size: 12px; }}
    .data-table {{ font-size: 11px; }}
    .data-table thead th {{ font-size: 11px; padding: 4px 3px; }}
    .data-table tbody td {{ padding: 3px 4px; }}
    .container {{ padding: 16px 10px; }}
}}
</style>
</head>
<body>
<div class="container">

<div class="hero">
    <h1>{title}</h1>
    <div class="date">报表日期：{datetime.date.today()}</div>
</div>

<div class="summary" id="summary"></div>

<div class="toggle-bar">
    <button class="toggle-btn active" id="btnCharts" onclick="switchView('charts')">看 板</button>
    <button class="toggle-btn" id="btnTable" onclick="switchView('table')">图 表</button>
    <button class="toggle-btn" id="btnExport" onclick="ExportManager.downloadOriginal()">导出Excel</button>
</div>

<div id="view-charts">
    <div class="region-legend" id="regionLegend"></div>
    <div class="charts-grid" id="chartsGrid"></div>
</div>

<div id="view-table">
    <div class="table-wrap">
        <table class="data-table">
            <tbody id="tableBody"></tbody>
        </table>
    </div>
</div>

<div class="footer">签单排产发货看板 · 数据自动生成</div>
</div>

<script>
const DATA = {chart_data};

// ---- helpers ----
function fmtNum(v) {{
    return Math.round(v).toLocaleString();
}}

function ratioClass(r) {{
    if (r >= 1) return 'ratio-badge-high';
    if (r >= 0.5) return 'ratio-badge-mid';
    return 'ratio-badge-low';
}}

// ---- summary cards ----
(function() {{
    var el = document.getElementById('summary');
    var grand = DATA.grand;
    DATA.metrics.forEach(function(m) {{
        var d = grand[m.id];
        var pct = (d.ratio * 100).toFixed(1);
        var card = document.createElement('div');
        card.className = 'summary-card';
        card.innerHTML = '<div class="sc-label">' + m.name + '</div>'
            + '<div class="sc-value">' + fmtNum(d.ac, m.unit) + '<span class="sc-unit">' + m.unit + '</span></div>'
            + '<div><span class="sc-ratio ' + ratioClass(d.ratio) + '">已完成 ' + pct + '%</span></div>';
        el.appendChild(card);
    }});
}})();

// ---- region legend ----
(function() {{
    var el = document.getElementById('regionLegend');
    DATA.regions.forEach(function(r, i) {{
        var item = document.createElement('span');
        item.className = 'legend-item';
        item.innerHTML = '<span class="legend-dot" style="background:' + DATA.regionColors[i] + '"></span>' + r.name;
        el.appendChild(item);
    }});
}})();

// ---- charts ----
(function() {{
    // normalize values so min arc is at least 18% of max arc
    function normalizeValues(vals) {{
        var maxVal = Math.max.apply(null, vals);
        if (maxVal === 0) return vals.map(function() {{ return 1; }});
        var floor = maxVal * 0.18;
        return vals.map(function(v) {{ return Math.max(floor, v); }});
    }}

    // compute per-label line lengths with cluster stagger
    function computeLabelLines(displayVals) {{
        var total = displayVals.reduce(function(a, b) {{ return a + b; }}, 0);
        var n = displayVals.length;

        // mid-angles with original index
        var mids = [];
        var cum = 0;
        for (var i = 0; i < n; i++) {{
            var arc = displayVals[i] / total * 360;
            mids.push({{ idx: i, mid: cum + arc / 2, arc: arc }});
            cum += arc;
        }}

        // sort by angle
        var sorted = mids.slice().sort(function(a, b) {{ return a.mid - b.mid; }});

        // find clusters: consecutive labels within CLUSTER_GAP degrees
        var CLUSTER_GAP = 42;
        var clusters = [];
        var cur = [sorted[0]];
        for (var i = 1; i < n; i++) {{
            if (sorted[i].mid - sorted[i - 1].mid < CLUSTER_GAP) {{
                cur.push(sorted[i]);
            }} else {{
                if (cur.length > 1) clusters.push(cur);
                cur = [sorted[i]];
            }}
        }}
        // check wrap-around cluster
        if (clusters.length > 0 && cur.length > 1) {{
            var head = clusters[0];
            var tail = cur;
            var wrapGap = tail[tail.length - 1].mid - head[0].mid;
            if (wrapGap < 0) wrapGap += 360;
            if (wrapGap < CLUSTER_GAP) {{
                clusters[0] = tail.concat(head);
            }} else {{
                clusters.push(tail);
            }}
        }} else if (cur.length > 1) {{
            clusters.push(cur);
        }}

        // assign stagger tier within each cluster
        var tier = {{}};
        for (var c = 0; c < clusters.length; c++) {{
            for (var k = 0; k < clusters[c].length; k++) {{
                tier[clusters[c][k].idx] = k;
            }}
        }}

        // count left/right
        var leftCount = 0, rightCount = 0;
        for (var i = 0; i < n; i++) {{
            if (mids[i].mid > 180) leftCount++;
            else rightCount++;
        }}

        var lines = [];
        for (var i = 0; i < n; i++) {{
            var t = tier[i] || 0;
            var len1 = 34 + t * 6;
            var len2 = 80 + t * 16;
            var sideCount = mids[i].mid > 180 ? leftCount : rightCount;
            var otherCount = mids[i].mid > 180 ? rightCount : leftCount;

            if (sideCount > otherCount + 1) {{
                len1 += 6;
                len2 += 10;
            }}

            var nearTopBottom = Math.min(
                Math.abs(mids[i].mid - 0),
                Math.abs(mids[i].mid - 180),
                Math.abs(mids[i].mid - 360)
            );
            if (nearTopBottom < 25) {{
                len1 += 8;
                len2 += 10;
            }}

            len1 = Math.min(len1, 52);
            len2 = Math.min(len2, 130);

            lines.push({{ length: len1, length2: len2 }});
        }}
        return lines;
    }}

    function shuffle(arr) {{
        var a = arr.slice();
        for (var i = a.length - 1; i > 0; i--) {{
            var j = Math.floor(Math.random() * (i + 1));
            var t = a[i]; a[i] = a[j]; a[j] = t;
        }}
        return a;
    }}

    function makeChartOption(metric, regions, grand) {{
        var mId = metric.id;
        var unit = metric.unit;
        var g = grand[mId];
        var innerColor = DATA.innerColors[mId];

        // shuffle region order so each chart has different arrangement
        var indexed = regions.map(function(r, i) {{ return {{region: r, origIdx: i}}; }});
        var shuffled = shuffle(indexed);

        var rawVals = shuffled.map(function(item) {{ return item.region[mId].ac; }});
        var displayVals = normalizeValues(rawVals);
        var labelLines = computeLabelLines(displayVals);

        var outerData = shuffled.map(function(item, i) {{
            var r = item.region;
            var origIdx = item.origIdx;
            var d = r[mId];
            var ll = labelLines[i];
            return {{
                name: r.name,
                value: displayVals[i],
                rawAc: d.ac,
                rawFc: d.fc,
                ratio: d.ratio,
                unit: unit,
                itemStyle: {{ color: DATA.regionColors[origIdx], borderRadius: 5, borderColor: '#fff', borderWidth: 2.5 }},
                labelLine: {{
                    length: ll.length,
                    length2: ll.length2,
                    lineStyle: {{ color: DATA.regionColors[origIdx], width: 1.5 }}
                }}
            }};
        }});

        var innerData;
        if (g.fc > 0) {{
            innerData = [
                {{ value: g.ac, name: '已完成', itemStyle: {{ color: innerColor, borderRadius: 3 }} }},
                {{ value: Math.max(0, g.fc - g.ac), name: '剩余', itemStyle: {{ color: '#E8E8E8', borderRadius: 3 }} }}
            ];
        }} else {{
            innerData = [
                {{ value: 1, name: '无数据', itemStyle: {{ color: '#E8E8E8' }} }}
            ];
        }}

        return {{
            color: DATA.regionColors,
            tooltip: {{
                trigger: 'item',
                formatter: function(p) {{
                    if (p.seriesIndex === 0) {{
                        var d = p.data;
                        return d.name + '<br/>已完成 ' + fmtNum(d.rawAc) + ' / 预计 ' + fmtNum(d.rawFc) + ' ' + unit
                            + '<br/>完成比 ' + (d.ratio * 100).toFixed(0) + '%';
                    }}
                    return '';
                }}
            }},
            animation: true,
            animationDuration: 800,
            animationEasing: 'cubicOut',
            series: [
                {{
                    type: 'pie',
                    radius: ['38%', '56%'],
                    center: ['50%', '50%'],
                    top: '5%',
                    bottom: '5%',
                    padAngle: 2,
                    itemStyle: {{ borderColor: '#fff', borderWidth: 2.5 }},
                    emphasis: {{
                        focus: 'self',
                        blurScope: 'coordinateSystem',
                        scale: true,
                        scaleSize: 8,
                        label: {{ fontSize: 20, fontWeight: 'bold' }},
                        labelLine: {{ lineStyle: {{ width: 2.5 }} }}
                    }},
                    blur: {{
                        itemStyle: {{ opacity: 0.2 }},
                        labelLine: {{ lineStyle: {{ opacity: 0.15 }} }}
                    }},
                    label: {{
                        show: true,
                        position: 'outside',
                        overflow: 'truncate',
                        ellipsis: '...',
                        formatter: function(p) {{
                            var d = p.data;
                            return '{{name|' + d.name + '}}\\n{{val|已完成' + fmtNum(d.rawAc) + unit + ' ' + (d.ratio * 100).toFixed(0) + '%}}';
                        }},
                        rich: {{
                            name: {{
                                fontSize: 14,
                                fontWeight: 'bold',
                                color: '#1a2332',
                                lineHeight: 22,
                                padding: [0, 0, 2, 0]
                            }},
                            val: {{
                                fontSize: 12,
                                color: '#4a5568',
                                lineHeight: 18
                            }}
                        }}
                    }},
                    data: outerData
                }},
                {{
                    type: 'pie',
                    radius: ['22%', '34%'],
                    center: ['50%', '50%'],
                    silent: true,
                    padAngle: 2,
                    label: {{ show: false }},
                    itemStyle: {{ borderColor: '#fff', borderWidth: 2 }},
                    data: innerData
                }}
            ],
            graphic: [
                {{
                    type: 'text',
                    left: 'center',
                    top: '45%',
                    z: 100,
                    style: {{
                        text: fmtNum(g.ac) + '/' + fmtNum(g.fc) + ' ' + unit,
                        textAlign: 'center',
                        fill: '#1a2332',
                        fontSize: 19,
                        fontWeight: 'bold'
                    }}
                }},
                {{
                    type: 'text',
                    left: 'center',
                    top: '53%',
                    z: 100,
                    style: {{
                        text: '已完成 ' + (g.ratio * 100).toFixed(0) + '%',
                        textAlign: 'center',
                        fill: g.ratio >= 1 ? '#2563eb' : g.ratio >= 0.5 ? '#3b82f6' : '#60a5fa',
                        fontSize: 16,
                        fontWeight: '600'
                    }}
                }}
            ]
        }};
    }}

    function debounce(fn, delay) {{
        var timer = null;
        return function() {{
            var ctx = this, args = arguments;
            clearTimeout(timer);
            timer = setTimeout(function() {{ fn.apply(ctx, args); }}, delay);
        }};
    }}

    function buildCharts() {{
        var grid = document.getElementById('chartsGrid');
        var regions = DATA.regions;
        var grand = DATA.grand;

        DATA.metrics.forEach(function(m) {{
            var card = document.createElement('div');
            card.className = 'chart-card';

            var titleEl = document.createElement('div');
            titleEl.className = 'chart-title';
            titleEl.textContent = m.name;
            card.appendChild(titleEl);

            var wrap = document.createElement('div');
            wrap.className = 'chart-wrap';
            wrap.id = 'chart-' + m.id;
            card.appendChild(wrap);

            grid.appendChild(card);
        }});

        DATA.metrics.forEach(function(m) {{
            var dom = document.getElementById('chart-' + m.id);
            if (!dom) return;
            var chart = echarts.init(dom);
            chart.setOption(makeChartOption(m, regions, grand));
            chart.on('mouseover', function(params) {{
                if (params.seriesIndex === 0) {{
                    chart.setOption({{
                        series: [{{}}, {{ itemStyle: {{ opacity: 0.4 }} }}],
                        graphic: [{{ style: {{ opacity: 0.4 }} }}, {{ style: {{ opacity: 0.4 }} }}]
                    }});
                }}
            }});
            chart.on('mouseout', function(params) {{
                chart.setOption({{
                    series: [{{}}, {{ itemStyle: {{ opacity: 1 }} }}],
                    graphic: [{{ style: {{ opacity: 1 }} }}, {{ style: {{ opacity: 1 }} }}]
                }});
            }});
            // ResizeObserver with debounce for responsive resize
            var chartRo = new ResizeObserver(debounce(function(entries) {{
                chart.resize();
                var w = entries[0].contentRect.width;
                if (w < 640) {{
                    chart.setOption({{
                        series: [{{
                            label: {{ rich: {{ name: {{ fontSize: 11 }}, val: {{ fontSize: 10 }} }} }}
                        }}, {{}}]
                    }});
                }} else {{
                    chart.setOption({{
                        series: [{{
                            label: {{ rich: {{ name: {{ fontSize: 14 }}, val: {{ fontSize: 12 }} }} }}
                        }}, {{}}]
                    }});
                }}
            }}, 200));
            chartRo.observe(dom);
        }});
    }}

    buildCharts();

    // ---- table view ----
    window.switchView = function(view) {{
        var btnC = document.getElementById('btnCharts');
        var btnT = document.getElementById('btnTable');
        var viewC = document.getElementById('view-charts');
        var viewT = document.getElementById('view-table');
        if (view === 'charts') {{
            btnC.classList.add('active'); btnT.classList.remove('active');
            viewC.classList.remove('hide'); viewT.classList.remove('show');
            DATA.metrics.forEach(function(m) {{
                var dom = document.getElementById('chart-' + m.id);
                if (dom) {{ var c = echarts.getInstanceByDom(dom); if (c) c.resize(); }}
            }});
        }} else {{
            btnT.classList.add('active'); btnC.classList.remove('active');
            viewC.classList.add('hide'); viewT.classList.add('show');
        }}
    }};

    function buildTable() {{
        var ratioCols = {{ E:1, H:1, K:1, N:1, Q:1, T:1 }};
        var amtCols = {{ F:1, G:1, L:1, M:1, R:1, S:1 }};

        var regionHeaderHtml = '<tr class="row-region-header">'
            + '<td rowspan="2" style="width:40px">序号</td>'
            + '<td rowspan="2">模块</td>'
            + '<td colspan="6">签单情况</td>'
            + '<td colspan="6">排产情况</td>'
            + '<td colspan="6">发货情况</td>'
            + '</tr>'
            + '<tr class="row-region-header row-region-subheader">'
            + '<td>预计台数</td><td>已签约</td><td>完成比</td><td>合同额</td><td>签单额</td><td>完成比</td>'
            + '<td>预计台数</td><td>已排产</td><td>完成比</td><td>合同额</td><td>排产额</td><td>完成比</td>'
            + '<td>预计台数</td><td>已发货</td><td>完成比</td><td>合同额</td><td>发货额</td><td>完成比</td>'
            + '</tr>';

        function makeRowHtml(r) {{
            var cls = r.isGrand ? 'row-grand' : r.isSubtotal ? 'row-subtotal' : 'row-normal';
            var html = '<tr class="' + cls + '">';
            html += '<td class="col-a">' + (r.a || '') + '</td>';
            html += '<td class="col-b">' + (r.b || '') + '</td>';
            'CDEFGHIJKLMNOPQRST'.split('').forEach(function(col) {{
                var v = r.cols[col];
                var tdCls = '';
                var display = '';
                if (v === '' || v === '-') {{
                    display = v === '-' ? '-' : '';
                }} else if (ratioCols[col]) {{
                    var num = Number(v);
                    if (!isNaN(num)) {{
                        if (num >= 1) tdCls = 'ratio-high';
                        else if (num >= 0.5) tdCls = 'ratio-mid';
                        else tdCls = 'ratio-low';
                        display = num === 0 ? '0' : (num * 100).toFixed(2) + '%';
                    }} else {{
                        display = v;
                    }}
                }} else if (amtCols[col]) {{
                    var num = Number(v);
                    display = !isNaN(num) ? Math.round(num).toString() : v;
                }} else {{
                    var num = Number(v);
                    display = !isNaN(num) ? Math.round(num).toString() : v;
                }}
                html += '<td class="' + tdCls + '">' + display + '</td>';
            }});
            html += '</tr>';
            return html;
        }}

        var rows = DATA.tableRows;
        var bodyHtml = '';
        var needHeader = true;

        rows.forEach(function(r, idx) {{
            if (needHeader && !r.isGrand) {{
                bodyHtml += regionHeaderHtml;
                needHeader = false;
            }}
            bodyHtml += makeRowHtml(r);
            if (r.isSubtotal) {{
                var nextRow = rows[idx + 1];
                if (!nextRow || !nextRow.isGrand) {{
                    bodyHtml += '<tr class="row-region-spacer"><td colspan="20" style="height:18px;border:none;background:#f0f2f5;"></td></tr>';
                }}
                needHeader = true;
            }}
        }});

        document.getElementById('tableBody').innerHTML = bodyHtml;
    }}

    buildTable();

    // ---- ExportManager ----
    window.ExportManager = (function() {{
        function _base64ToBlob(b64, mimeType) {{
            var byteChars = atob(b64);
            var byteNums = new Array(byteChars.length);
            for (var i = 0; i < byteChars.length; i++) {{
                byteNums[i] = byteChars.charCodeAt(i);
            }}
            var byteArr = new Uint8Array(byteNums);
            return new Blob([byteArr], {{ type: mimeType }});
        }}

        function _downloadBlob(blob, filename) {{
            var url = URL.createObjectURL(blob);
            var a = document.createElement('a');
            a.href = url;
            a.download = filename;
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            URL.revokeObjectURL(url);
        }}

        return {{
            downloadOriginal: function() {{
                if (!DATA.excelB64) {{
                    alert('原始Excel文件数据不可用');
                    return;
                }}
                var blob = _base64ToBlob(DATA.excelB64, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');
                _downloadBlob(blob, DATA.excelFilename || 'export.xlsx');
            }}
        }};
    }})();
}})();
</script>
</body>
</html>'''

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"[HTML] 已生成: {os.path.basename(output_path)}")


# ============================================================
# HTTP 服务器 (局域网共享)
# ============================================================
def start_serve(work_dir, html_path, port=8080):
    html_name = os.path.basename(html_path)

    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return "127.0.0.1"

    local_ip = get_local_ip()

    class Handler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *a, **kw):
            super().__init__(*a, directory=work_dir, **kw)

        def log_message(self, fmt, *args):
            pass  # 静默日志

    try:
        server = socketserver.TCPServer(("0.0.0.0", port), Handler)
    except OSError as e:
        print(f"[错误] 端口 {port} 被占用: {e}")
        return

    t = threading.Thread(target=server.serve_forever, daemon=True)
    t.start()

    print(f"\n{'='*60}")
    print(f"  LAN 服务器已启动 (按 Ctrl+C 停止)")
    print(f"  本机访问: http://localhost:{port}/{html_name}")
    print(f"  局域网:   http://{local_ip}:{port}/{html_name}")
    print(f"{'='*60}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n[服务器] 已停止")
        server.shutdown()


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="签单排产发货看板生成")
    parser.add_argument("--month", default=None, help="统计月份, 如 2026-05")
    parser.add_argument("--dir", default=os.getcwd(), help="输出Excel所在目录")
    parser.add_argument("--output", default=None, help="看板HTML输出路径")
    parser.add_argument("--serve", action="store_true", help="生成后启动HTTP服务器供局域网访问")
    parser.add_argument("--port", type=int, default=8080, help="HTTP服务器端口 (默认8080)")
    args = parser.parse_args()

    if args.month:
        target_month = args.month
    else:
        today = datetime.date.today()
        target_month = f"{today.year}-{today.month:02d}"

    work_dir = args.dir
    today_str = datetime.date.today().strftime("%m-%d")

    excel_path = os.path.join(work_dir, OUTPUT_TEMPLATE.format(date=today_str))
    html_path = args.output or os.path.join(work_dir, "dashboard.html")

    if not os.path.exists(excel_path):
        import glob
        candidates = sorted(glob.glob(os.path.join(work_dir, "签单排产发货_*.xlsx")), reverse=True)
        if candidates:
            excel_path = candidates[0]
            print(f"[信息] 未找到指定日期文件，使用最新: {os.path.basename(excel_path)}")
        else:
            print(f"[错误] 未找到签单排产发货_*.xlsx，请先运行 generate_report.py")
            sys.exit(1)

    print(f"{'='*60}")
    print(f"  签单排产发货看板生成")
    print(f"  月份: {target_month}")
    print(f"{'='*60}\n")
    print(f"[文件] 数据源: {os.path.basename(excel_path)}")
    print(f"[文件] 看板: {os.path.basename(html_path)}\n")

    import openpyxl as _openpyxl
    _wb = _openpyxl.load_workbook(excel_path, data_only=True)
    try:
        region_data, grand = read_dashboard_data(excel_path, workbook=_wb)
        table_rows = read_table_rows(excel_path, workbook=_wb)
        generate_html(region_data, grand, target_month, html_path, table_rows, excel_path)
    finally:
        _wb.close()
    print(f"\n{'='*60}")
    print(f"  完成! 浏览器打开: {os.path.basename(html_path)}")
    print(f"{'='*60}")

    if args.serve:
        start_serve(work_dir, html_path, args.port)


if __name__ == "__main__":
    main()
