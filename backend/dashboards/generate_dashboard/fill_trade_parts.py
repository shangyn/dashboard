#!py -3.11
"""
fill_trade_parts.py - 商贸配件数据填充
============================================================
用法: py -3.11 fill_trade_parts.py --month 2026-05

先调用 generate_report.py 生成报表，
再从 报表a.xls / 报表b.xls 匹配填充配件-1/配件-2的签单/排产/发货金额。
后续其他Excel匹配逻辑也在此文件中扩展。
============================================================
"""
import argparse
import sys
import os
import datetime
import pandas as pd
import openpyxl
import generate_report


def safe_float(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def match_date(col_values, target_month):
    """返回布尔Series：日期以target_month开头"""
    return col_values.astype(str).str.startswith(target_month, na=False)


def compute_trade_amounts_from_a(df_a, target_month):
    """从报表a计算配件-1/配件-2的签单额(G)和排产额(M)

    报表a列:
      K(col 10): 签订日期
      P(col 15): 排产日期
      G(col 6):  合同额（人民币）
      BC(col 54): 模块
    返回: {('配件-1', 'sign'): 万元, ('配件-1', 'prod'): 万元, ...}
    """
    result = {}
    for module_name in ["配件-1", "配件-2"]:
        mask_mod = df_a.iloc[:, 54].astype(str).str.strip() == module_name

        # 签单额: K列日期匹配
        mask_sign = mask_mod & match_date(df_a.iloc[:, 10], target_month)
        sign_sum = df_a.loc[mask_sign, df_a.columns[6]].apply(safe_float).sum()
        result[(module_name, "sign")] = sign_sum / 10000

        # 排产额: P列日期匹配
        mask_prod = mask_mod & match_date(df_a.iloc[:, 15], target_month)
        prod_sum = df_a.loc[mask_prod, df_a.columns[6]].apply(safe_float).sum()
        result[(module_name, "prod")] = prod_sum / 10000

    return result


def compute_trade_amounts_from_b(df_b, target_month):
    """从报表b计算配件-1/配件-2的发货额(S)

    报表b列:
      G(col 6):  实际发运时间
      M(col 12): 总金额（人民币）
      P(col 15): 模块
    返回: {('配件-1', 'ship'): 万元, ('配件-2', 'ship'): 万元}
    """
    result = {}
    for module_name in ["配件-1", "配件-2"]:
        mask_mod = df_b.iloc[:, 15].astype(str).str.strip() == module_name
        mask_date = match_date(df_b.iloc[:, 6], target_month)
        mask = mask_mod & mask_date
        ship_sum = df_b.loc[mask, df_b.columns[12]].apply(safe_float).sum()
        result[(module_name, "ship")] = ship_sum / 10000

    return result


def find_module_rows(ws, module_names):
    """在模板B列查找模块名所在行号，返回 {模块名: 行号}"""
    result = {}
    for row in range(5, ws.max_row + 1):
        b_val = str(ws.cell(row=row, column=2).value or "").strip()
        if b_val in module_names:
            result[b_val] = row
    return result


def fill_trade_cells(report_path, source_dir, target_month):
    """打开报表，填充配件-1/配件-2的G/M/S列"""
    # 读取数据源
    df_a = pd.read_excel(os.path.join(source_dir, "报表a.xls"))
    df_b = pd.read_excel(os.path.join(source_dir, "报表b.xls"))

    amounts_a = compute_trade_amounts_from_a(df_a, target_month)
    amounts_b = compute_trade_amounts_from_b(df_b, target_month)

    # 打开报表
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active

    rows = find_module_rows(ws, {"配件-1", "配件-2"})

    # 列映射: G=签单额, M=排产额, S=发货额
    # (模块名, 类型) → 模板列字母
    col_map = {
        ("sign", "G"): "sign",
        ("prod", "M"): "prod",
        ("ship", "S"): "ship",
    }
    type_to_col = {"sign": "G", "prod": "M", "ship": "S"}

    for mod_name in ["配件-1", "配件-2"]:
        if mod_name not in rows:
            print(f"  [警告] 模板中未找到模块: {mod_name}")
            continue
        row = rows[mod_name]

        for data_type, col_letter in type_to_col.items():
            if data_type == "ship":
                amount = amounts_b.get((mod_name, data_type), 0)
            else:
                amount = amounts_a.get((mod_name, data_type), 0)

            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col_idx)
            cell.value = amount
            print(f"  {mod_name} {col_letter}列({data_type}): {amount:.2f}万元")

    # 配件数据变化后，重算模块行完成比与商贸配件合计/全国贸合计（静态值，兼容手机端）
    _refresh_totals(ws)

    wb.save(report_path)
    wb.close()
    print(f"  已保存: {os.path.basename(report_path)}")


def _refresh_totals(ws):
    """配件填充后，重算模块行完成比与商贸配件合计/全国贸合计为静态计算值。

    openpyxl 写公式不会缓存计算结果，手机端表格App多不自动重算，
    因此这里把数值直接算好写进去，保证任何端显示一致。
    """
    from openpyxl.utils import column_index_from_string

    ratio_pairs = {
        "E": ("C", "D"), "H": ("F", "G"),
        "K": ("I", "J"), "N": ("L", "M"),
        "Q": ("O", "P"), "T": ("R", "S"),
    }
    ratio_cols = list(ratio_pairs.keys())
    value_cols = "CDEFGHIJKLMNOPQRST"

    def _num(row, col):
        v = ws.cell(row=row, column=column_index_from_string(col)).value
        return float(v) if isinstance(v, (int, float)) else 0.0

    # 定位合计行
    subtotal_rows = []
    grand_row = None
    for r in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row=r, column=1).value or "").strip()
        if "全国贸" in a_val:
            grand_row = r
        elif "合计" in a_val:
            subtotal_rows.append(r)

    # 1. 模块行完成比：用所在行的预测/实际值重算（跳过"-"等文本）
    for r in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row=r, column=1).value or "").strip()
        if "合计" in a_val:
            continue
        for rc in ratio_cols:
            cell = ws.cell(row=r, column=column_index_from_string(rc))
            is_num = isinstance(cell.value, (int, float))
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if not (is_num or is_formula):
                continue
            fc_col, ac_col = ratio_pairs[rc]
            fc_val = _num(r, fc_col)
            ac_val = _num(r, ac_col)
            cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
            cell.number_format = generate_report.FMT_RATIO

    # 2. 子合计行：上方模块行求和
    prev = 4
    for sr in subtotal_rows:
        for col in value_cols:
            cell = ws.cell(row=sr, column=column_index_from_string(col))
            if col in ratio_cols:
                fc_col, ac_col = ratio_pairs[col]
                fc_val = _num(sr, fc_col)
                ac_val = _num(sr, ac_col)
                cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
                cell.number_format = generate_report.FMT_RATIO
            else:
                cell.value = sum(_num(r, col) for r in range(prev + 1, sr))
        prev = sr

    # 3. 全国贸合计：所有子合计行求和
    if grand_row is not None:
        for col in value_cols:
            cell = ws.cell(row=grand_row, column=column_index_from_string(col))
            if col in ratio_cols:
                fc_col, ac_col = ratio_pairs[col]
                fc_val = _num(grand_row, fc_col)
                ac_val = _num(grand_row, ac_col)
                cell.value = ac_val / fc_val if fc_val else (1 if ac_val else 0)
                cell.number_format = generate_report.FMT_RATIO
            else:
                cell.value = sum(_num(sr, col) for sr in subtotal_rows)


def main():
    parser = argparse.ArgumentParser(description="商贸配件数据填充")
    parser.add_argument("--month", default=None, help="统计月份, 如 2026-05")
    parser.add_argument("--dir", default=os.getcwd(), help="输出目录")
    parser.add_argument("--source", default="数据源excel", help="数据源子目录（相对于--dir）")
    args = parser.parse_args()

    if args.month:
        target_month = args.month
    else:
        today = datetime.date.today()
        target_month = f"{today.year}-{today.month:02d}"

    work_dir = args.dir
    source_dir = os.path.join(work_dir, args.source)

    print(f"{'=' * 60}")
    print(f"  商贸配件数据填充")
    print(f"  月份: {target_month}")
    print(f"{'=' * 60}\n")

    # Step 1: 生成报表 (直接调用 generate_report 模块)
    print("[Step 1] 生成报表...")

    # 匹配文件
    ledger_path = generate_report.find_file(source_dir, generate_report.LEDGER_KW)
    mapping_path = generate_report.find_file(source_dir, generate_report.MAPPING_KW)
    forecast_path = generate_report.find_file(source_dir, generate_report.FORECAST_KW)
    template_path = os.path.join(source_dir, generate_report.TEMPLATE_NAME)

    # 检查文件
    for name, path in [("台账", ledger_path), ("映射", mapping_path),
                        ("同期对比", forecast_path), ("模板", template_path)]:
        if not path or not os.path.exists(path):
            print(f"[错误] 未找到{name}文件: {path}")
            sys.exit(1)

    print(f"[文件] 台账: {os.path.basename(ledger_path)}")
    print(f"[文件] 映射: {os.path.basename(mapping_path)}")
    print(f"[文件] 同期对比: {os.path.basename(forecast_path)}")
    print(f"[文件] 模板: {os.path.basename(template_path)}")

    # 加载数据文件
    wb_map = openpyxl.load_workbook(mapping_path, data_only=True)
    wb_fc = openpyxl.load_workbook(forecast_path, data_only=True)
    wb_ledger = openpyxl.load_workbook(ledger_path, data_only=True)

    # 处理数据
    country_to_module, _, _ = generate_report.build_mapping(wb_map)
    forecast = generate_report.build_forecast_data(wb_fc, target_month)
    stats, unmatched = generate_report.process_ledger(wb_ledger, country_to_module, target_month)

    # 确定输出文件并填充模板
    today_str = datetime.date.today().strftime("%m-%d")
    report_path = os.path.join(work_dir, f"签单排产发货_{today_str}.xlsx")
    generate_report.fill_template(template_path, forecast, stats, target_month, report_path)

    # 未匹配国家
    if unmatched:
        unmatched_path = os.path.join(work_dir, f"未匹配国家_{target_month}.txt")
        with open(unmatched_path, "w", encoding="utf-8") as f:
            for c in sorted(unmatched):
                f.write(f"{c}\n")
        print(f"[信息] 未匹配国家: {unmatched_path}")

    print(f"[完成] {os.path.basename(report_path)}")

    # Step 2: 填充配件数据
    print(f"\n[Step 2] 填充配件-1/配件-2数据...")
    fill_trade_cells(report_path, source_dir, target_month)

    # Step 3: Excel重算（确保手机端公式缓存值正确）
    generate_report.recalc_with_excel(report_path)

    print(f"\n{'=' * 60}")
    print(f"  完成!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
